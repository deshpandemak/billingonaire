import asyncio
import json
import logging
import os
import posixpath
import sys
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import firebase_admin
import pandas as pd
import requests
from fastapi import Depends, FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from firebase_admin import auth, credentials, firestore
from pydantic import BaseModel

# Configure logging to show INFO level messages with timestamps for Cloud Log Viewer
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S%z",
)
logger = logging.getLogger(__name__)

_overview_stats_cache: Dict[str, Any] = {"ts": 0.0, "data": None}
_queue_status_cache: Dict[str, Any] = {"ts": 0.0, "data": None}

# Integrate with Google Cloud Logging when running on GCP (Cloud Run sets K_SERVICE)
if os.getenv("K_SERVICE"):
    try:
        import google.cloud.logging as gcp_logging
    except ImportError:
        logger.info(
            "google-cloud-logging not installed; using standard logging only.",
        )
    else:
        try:
            _gcp_log_client = gcp_logging.Client()
            _gcp_log_client.setup_logging(log_level=logging.INFO)
            logger.info("Google Cloud Logging integration enabled")
        except Exception:
            logger.warning(
                "Failed to initialize Google Cloud Logging; falling back to standard logging.",
                exc_info=True,
            )

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from Board import Board  # noqa: E402
from Board import SIMPLE_STATUS_KEYS, simple_status_for  # noqa: E402
from CourtScraper import BombayHighCourtScraper  # noqa: E402
from Dashboard import DashboardData  # noqa: E402
from OrderManager import OrderManager  # noqa: E402
from UserManager import UserManager  # noqa: E402
from UserMatterMatcher import UserRole  # noqa: E402

app = FastAPI(
    title="Billingonaire API",
    description="API for Billingonaire application",
    version="1.0.0",
    openapi_tags=[
        {"name": "Root", "description": "Root endpoint"},
        {"name": "PDF Upload", "description": "Upload PDF and extract data"},
        {"name": "Data Retrieval", "description": "Retrieve stored data"},
        {"name": "Authentication", "description": "User authentication"},
        {
            "name": "Case Status",
            "description": "Retrieve case status from Bombay High Court",
        },
        {
            "name": "Case Orders",
            "description": "Retrieve case orders from Bombay High Court",
        },
        {
            "name": "Order Management",
            "description": "Manage court order linking and states",
        },
        {
            "name": "Order Analysis",
            "description": "ML-powered analysis of court order documents",
        },
        {
            "name": "Queue Management",
            "description": "Monitor async order processing queue",
        },
        {
            "name": "User Matter Mapping",
            "description": "Link users to their legal matters using AI-powered name matching",
        },
    ],
)

# Lazy Firebase initialization - deferred until first use to avoid blocking port binding
_firebase_initialized = False
_firebase_init_error = None


def ensure_firebase():
    """Initialize Firebase Admin SDK on first use"""
    global _firebase_initialized, _firebase_init_error
    if not _firebase_initialized:
        if not firebase_admin._apps:
            import json

            # Log environment info for debugging
            logger.info("🔍 Firebase initialization - Environment check:")
            logger.info(
                f"   - Running in Cloud: {os.environ.get('K_SERVICE') is not None}"
            )
            logger.info(
                f"   - Service account key available: {bool(os.environ.get('GCLOUD_SERVICE_ACCOUNT_KEY'))}"
            )
            logger.info(
                f"   - Google credentials env: {bool(os.environ.get('GOOGLE_APPLICATION_CREDENTIALS'))}"
            )

            gcloud_key = os.environ.get("GCLOUD_SERVICE_ACCOUNT_KEY")
            if gcloud_key:
                try:
                    # Environment with service account key (local/Replit/Cloud Run with secret)
                    cred_dict = json.loads(gcloud_key)
                    cred = credentials.Certificate(cred_dict)
                    firebase_admin.initialize_app(cred)
                    logger.info(
                        "✅ Firebase Admin SDK initialized with service account key"
                    )
                except json.JSONDecodeError as e:
                    _firebase_init_error = (
                        f"Invalid JSON in GCLOUD_SERVICE_ACCOUNT_KEY: {str(e)}"
                    )
                    logger.error(f"❌ {_firebase_init_error}")
                    raise HTTPException(
                        status_code=500,
                        detail="Server configuration error: Invalid Firebase service account JSON. Contact administrator.",
                    )
                except Exception as e:
                    _firebase_init_error = f"Failed to initialize Firebase with service account key: {str(e)}"
                    logger.error(f"❌ {_firebase_init_error}")
                    raise HTTPException(
                        status_code=500,
                        detail="Server configuration error: Firebase credentials invalid. Contact administrator.",
                    )
            else:
                # Try Application Default Credentials (Cloud Run, Compute Engine, etc.)
                try:
                    logger.info(
                        "🔄 Attempting to initialize with Application Default Credentials..."
                    )
                    firebase_admin.initialize_app()
                    logger.info(
                        "✅ Firebase Admin SDK initialized with Application Default Credentials"
                    )
                except Exception as e:
                    # Final attempt: try with explicit project ID
                    try:
                        logger.info("🔄 Retrying with explicit project configuration...")
                        project_id = os.environ.get(
                            "GCP_PROJECT",
                            os.environ.get("GOOGLE_CLOUD_PROJECT", "billingonaire"),
                        )
                        config = {
                            "projectId": project_id,
                        }
                        firebase_admin.initialize_app(config)
                        logger.info(
                            f"✅ Firebase Admin SDK initialized with project ID: {project_id}"
                        )
                    except Exception as e2:
                        _firebase_init_error = (
                            f"Firebase Admin SDK initialization failed. "
                            f"ADC Error: {str(e)}. Project Config Error: {str(e2)}. "
                            f"Missing GCLOUD_SERVICE_ACCOUNT_KEY environment variable."
                        )
                        logger.error(f"❌ {_firebase_init_error}")
                        logger.error(
                            "💡 To fix: Set GCLOUD_SERVICE_ACCOUNT_KEY environment variable with Firebase service account JSON"
                        )
                        logger.error(
                            "💡 Or ensure Cloud Run service account has Firebase Admin permissions"
                        )
                        raise HTTPException(
                            status_code=500,
                            detail="Server configuration error: Firebase credentials not configured. Contact administrator.",
                        )
        _firebase_initialized = True


app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://billingonaire.web.app",
        "http://localhost:5000",
        "http://localhost:5173",
        "http://localhost:5174",
        "https://2856c3cf-582f-4f2b-a0f3-cae6a5c3b647-00-5mlgokfyfmx.pike.replit.dev",
        "http://2856c3cf-582f-4f2b-a0f3-cae6a5c3b647-00-5mlgokfyfmx.pike.replit.dev",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Lazy initialization of heavy objects to avoid blocking Cloud Run startup
user_manager = None
order_analyzer = None
auto_order_manager = None
user_matter_matcher = None


def get_user_manager():
    global user_manager
    if user_manager is None:
        user_manager = UserManager()
    return user_manager


def get_order_analyzer():
    global order_analyzer
    if order_analyzer is None:
        from order_analyzer import OrderDocumentAnalyzer

        order_analyzer = OrderDocumentAnalyzer()
    return order_analyzer


def get_auto_order_manager():
    global auto_order_manager
    if auto_order_manager is None:
        from AutoOrderManager import AutoOrderManager

        auto_order_manager = AutoOrderManager()
    return auto_order_manager


def get_user_matter_matcher():
    global user_matter_matcher
    if user_matter_matcher is None:
        from UserMatterMatcher import UserMatterMatcher

        user_matter_matcher = UserMatterMatcher()
    return user_matter_matcher


# Order fetch/analysis work is tracked entirely in Firestore (case-details.
# lifecycle_status) rather than an in-memory queue: Cloud Run runs multiple
# instances and scales to zero, so an asyncio.Queue is invisible across
# instances and is silently lost whenever an idle instance scales down.
# Every instance instead runs fetch_poll_loop/analysis_poll_loop (started at
# app startup, see the "startup" event handler below), each polling for
# lifecycle_status == *_queued (plus stale *_in_progress reclaim) and
# atomically claiming candidates via CaseDataStore.claim_for_processing.

# Thread pool executor for blocking operations (configurable via env var)
try:
    MAX_WORKERS = max(1, int(os.environ.get("ORDER_PROCESSING_WORKERS", "5")))
except (ValueError, TypeError):
    logger.warning("Invalid ORDER_PROCESSING_WORKERS value, using default of 5")
    MAX_WORKERS = 5
executor = ThreadPoolExecutor(max_workers=MAX_WORKERS)

try:
    STALE_IN_PROGRESS_MINUTES = max(
        1, int(os.environ.get("STALE_IN_PROGRESS_MINUTES", "10"))
    )
except (ValueError, TypeError):
    logger.warning("Invalid STALE_IN_PROGRESS_MINUTES value, using default of 10")
    STALE_IN_PROGRESS_MINUTES = 10

try:
    QUEUE_POLL_INTERVAL_SECONDS = max(
        1, int(os.environ.get("QUEUE_POLL_INTERVAL_SECONDS", "5"))
    )
except (ValueError, TypeError):
    logger.warning("Invalid QUEUE_POLL_INTERVAL_SECONDS value, using default of 5")
    QUEUE_POLL_INTERVAL_SECONDS = 5

QUEUE_POLL_BATCH_SIZE = MAX_WORKERS * 2

_fetch_semaphore = asyncio.Semaphore(MAX_WORKERS)
_analysis_semaphore = asyncio.Semaphore(MAX_WORKERS)
# Set by /queue/restart to make a poll loop check immediately instead of
# waiting out its interval; cleared at the top of every tick.
_wake_fetch_poll = asyncio.Event()
_wake_analysis_poll = asyncio.Event()
# Keeps strong references to spawned tasks: asyncio.create_task() only holds
# a weak reference, so an unreferenced task can be silently garbage
# collected mid-run -- a second, independent way work could vanish.
_background_tasks: set = set()


def _track_task(task: "asyncio.Task") -> None:
    _background_tasks.add(task)
    task.add_done_callback(_background_tasks.discard)


def get_current_user(request: Request):
    ensure_firebase()  # Initialize Firebase before auth operations
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(
            status_code=401, detail="Missing or invalid authentication token"
        )

    id_token = auth_header.split("Bearer ")[1]

    try:
        # Verify the Firebase ID token with more detailed error logging
        logger.info("Attempting to verify ID token for authentication")
        decoded_token = auth.verify_id_token(id_token)
        logger.info(f"Token verified successfully for user: {decoded_token.get('uid')}")
        return decoded_token
    except Exception as e:
        logger.error(f"Token verification failed: {str(e)}")
        # SECURITY: Do not log token details to prevent leakage
        raise HTTPException(status_code=401, detail="Invalid authentication token")


def require_active_user(current_user: dict = Depends(get_current_user)):
    """Dependency to require active user account"""
    uid = current_user.get("uid")
    profile = get_user_manager().get_user_profile(uid)

    if not profile.get("is_active", True):
        raise HTTPException(
            status_code=403, detail="Account is disabled. Contact administrator."
        )

    return {**current_user, "profile": profile}


def get_user_with_profile(current_user: dict = Depends(require_active_user)):
    """Dependency to get current user with profile (active users only)"""
    return current_user


def require_admin(current_user: dict = Depends(get_current_user)):
    """Dependency to require admin role"""
    if not get_user_manager().is_admin(current_user.get("uid")):
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user


def require_admin_active(current_user: dict = Depends(require_active_user)):
    """Dependency to require active admin user"""
    if not get_user_manager().is_admin(current_user.get("uid")):
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user


# Async Order Processing Functions
async def trigger_async_order_processing(df: pd.DataFrame):
    """Mark uploaded cases fetch_queued so a fetch_poll_loop picks them up."""
    try:
        records = df.to_dict(orient="records")
        case_store = get_auto_order_manager().case_store

        for record in records:
            case_ref = (
                f"{record['case_type']}/{record['case_no']}/{record['case_year']}"
            )
            case_store.transition_lifecycle(
                case_ref,
                "fetch_queued",
                metadata={
                    "source": "board_upload",
                    "board_date": record.get("board_date"),
                },
                event_type="fetch_queued_from_upload",
                extra_fields={
                    "latest_board_date": case_store._to_iso_date(
                        record.get("board_date")
                    )
                },
            )
            logger.info(f"Marked case {case_ref} fetch_queued after board upload")

        _wake_fetch_poll.set()

    except Exception as e:
        logger.error(f"Error queueing uploaded cases for fetch: {e}")


def _run_fetch_case(case_info: Dict) -> Dict:
    """Blocking fetch+inline-analysis job used by fetch_poll_loop's executor."""
    return get_auto_order_manager()._process_single_case(case_info)


def _resolve_board_doc_id(case_info: Dict) -> Optional[str]:
    """auto_map_case_to_users (and everything downstream of it --
    user-case-mappings, user-matter-pending-confirmations, and the
    bill-generation read path at GET /bills/generate) all expect a
    daily-boards doc id ("YYYY-MM-DD-TYPE-NO-YEAR"). The poll loops key
    candidates off case-details instead ("TYPE-NO-YEAR", no date), so
    case_info["id"] is the wrong shape to pass straight through --
    resolve the real daily-boards id from board_assignment_ids
    (case_data_store.py's link back to the board rows a case appeared
    on), preferring the entry for the board_date being processed right
    now since a case can appear on the board more than once."""
    board_ids = case_info.get("board_assignment_ids") or []
    if not board_ids:
        return None
    board_date = case_info.get("board_date")
    if board_date:
        for board_id in board_ids:
            if board_id.startswith(f"{board_date}-"):
                return board_id
    return board_ids[-1]


async def _process_claimed_fetch_case(case_info: Dict) -> None:
    case_ref = case_info["case_ref"]
    try:
        loop = asyncio.get_event_loop()
        result = await asyncio.wait_for(
            loop.run_in_executor(executor, _run_fetch_case, case_info),
            timeout=300.0,
        )

        if result.get("analysis_success"):
            logger.info(f"✅ Fetch+analysis succeeded for {case_ref}")
            board_doc_id = _resolve_board_doc_id(case_info)
            if board_doc_id:
                try:
                    await auto_map_case_to_users(board_doc_id, case_info)
                except Exception as mapping_error:
                    logger.error(
                        f"Error mapping users after fetch for {case_ref}: {mapping_error}"
                    )
            else:
                logger.warning(
                    f"No board_assignment_ids for {case_ref} -- skipping user mapping"
                )
        elif result.get("download_success"):
            # Order downloaded but inline analysis didn't complete (rare --
            # _process_single_case normally analyses inline). Queue it for
            # the analysis poll loop rather than leaving it stranded.
            get_auto_order_manager().case_store.transition_lifecycle(
                case_ref,
                "analysis_queued",
                reason="Order downloaded but not analysed inline",
                metadata={"source": "fetch_poll_loop"},
                event_type="analysis_queued_after_fetch",
            )
            _wake_analysis_poll.set()
        else:
            logger.warning(
                f"⚠️ Fetch failed for {case_ref}: {result.get('error', 'Unknown error')}"
            )
    except asyncio.TimeoutError:
        logger.error(f"❌ Timeout after 5 minutes fetching {case_ref}")
        try:
            get_auto_order_manager().case_store.transition_lifecycle(
                case_ref,
                "fetch_failed_terminal",
                reason="Worker timeout after 5 minutes",
                force=True,
                metadata={"source": "fetch_poll_loop_timeout"},
                event_type="fetch_timeout",
            )
        except Exception as lc_err:
            logger.error(f"Failed to mark lifecycle failed after timeout: {lc_err}")
    except Exception as e:
        logger.error(f"❌ Error fetching {case_ref}: {e}")
        try:
            get_auto_order_manager().case_store.transition_lifecycle(
                case_ref,
                "fetch_failed_terminal",
                reason=f"Worker error: {str(e)[:200]}",
                force=True,
                metadata={"source": "fetch_poll_loop_exception"},
                event_type="fetch_error",
            )
        except Exception as lc_err:
            logger.error(f"Failed to mark lifecycle failed after exception: {lc_err}")
    finally:
        _fetch_semaphore.release()


def _query_claim_candidates(
    case_store,
    queued_status: str,
    in_progress_status: str,
    batch_size: int,
    backlog_statuses: Optional[tuple] = None,
) -> List[Dict]:
    """Single-field equality queries only (already auto-indexed, no new
    composite index needed): cases waiting to start, cases stuck at the
    in-progress status past the staleness window (a worker that died
    mid-run without reaching a terminal status), and -- only if there's
    still room left in the batch -- cases sitting in an un-queued backlog
    state (e.g. board_ingested) that nothing has ever explicitly queued.
    This last tier is what keeps the pipeline self-feeding: without it,
    cases that were never queued by an upload or a manual "Fetch Orders"
    click sit forever, invisible to this poll loop, no matter how idle it
    is. Every candidate is tagged with the lifecycle_status it was actually
    found at (`_claim_from_status`) so the caller claims from the right
    state instead of assuming it's always `queued_status`."""
    db = firestore.client()
    candidates: List[Dict] = []
    seen_ids = set()

    queued_query = (
        db.collection("case-details")
        .where("lifecycle_status", "==", queued_status)
        .limit(batch_size)
    )
    for doc in queued_query.stream():
        if doc.id in seen_ids:
            continue
        seen_ids.add(doc.id)
        data = doc.to_dict() or {}
        data["_doc_id"] = doc.id
        data["_claim_from_status"] = queued_status
        candidates.append(data)

    if len(candidates) < batch_size:
        stale_query = (
            db.collection("case-details")
            .where("lifecycle_status", "==", in_progress_status)
            .limit(batch_size * 4)
        )
        for doc in stale_query.stream():
            if len(candidates) >= batch_size or doc.id in seen_ids:
                continue
            data = doc.to_dict() or {}
            if not case_store._is_stale(
                data.get("lifecycle_status_updated_at"), STALE_IN_PROGRESS_MINUTES
            ):
                continue
            seen_ids.add(doc.id)
            data["_doc_id"] = doc.id
            data["_claim_from_status"] = in_progress_status
            candidates.append(data)

    if backlog_statuses and len(candidates) < batch_size:
        for backlog_status in backlog_statuses:
            if len(candidates) >= batch_size:
                break
            backlog_query = (
                db.collection("case-details")
                .where("lifecycle_status", "==", backlog_status)
                .limit(batch_size - len(candidates))
            )
            for doc in backlog_query.stream():
                if len(candidates) >= batch_size or doc.id in seen_ids:
                    continue
                data = doc.to_dict() or {}
                seen_ids.add(doc.id)
                data["_doc_id"] = doc.id
                data["_claim_from_status"] = backlog_status
                candidates.append(data)

    return candidates


async def fetch_poll_loop():
    """Runs for the lifetime of the process, one per Cloud Run instance
    (started at app startup). Polls for fetch_queued cases plus stale
    fetch_in_progress cases, atomically claims each, and runs it in the
    thread pool bounded by _fetch_semaphore."""
    logger.info("🚀 Fetch poll loop started")

    while True:
        _write_poll_heartbeat("fetch_last_tick")
        try:
            # Constructed fresh every tick (not once before the loop): Firebase
            # is initialized lazily on first authenticated request, so on a
            # cold start this can fail on the very first tick. Keeping it
            # inside the try lets the loop retry next tick instead of dying
            # for the rest of the process's life.
            ensure_firebase()
            case_store = get_auto_order_manager().case_store
            candidates = _query_claim_candidates(
                case_store,
                "fetch_queued",
                "fetch_in_progress",
                QUEUE_POLL_BATCH_SIZE,
                # Cases nothing has ever explicitly queued -- a bulk-imported
                # or historically-ingested board row whose case-details doc
                # never went through the upload/"Fetch Orders" queueing path.
                # Without this the pipeline goes idle the moment fetch_queued
                # and stale fetch_in_progress both run dry, even with tens of
                # thousands of un-fetched cases still sitting at board_ingested.
                backlog_statuses=("board_ingested", "not_linked"),
            )
            for case_data in candidates:
                case_ref = case_data.get("case_ref")
                if not case_ref:
                    continue
                claim = case_store.claim_for_processing(
                    case_ref,
                    "fetch_in_progress",
                    from_statuses={case_data.get("_claim_from_status", "fetch_queued")},
                    stale_after_minutes=STALE_IN_PROGRESS_MINUTES,
                    reason="Claimed by fetch poll loop",
                    event_type="fetch_claimed",
                )
                if not claim["applied"]:
                    continue
                case_info = {
                    "id": case_data.get("_doc_id"),
                    "case_ref": case_ref,
                    "board_date": case_data.get("board_date")
                    or case_data.get("latest_board_date"),
                    "board_assignment_ids": case_data.get("board_assignment_ids") or [],
                }
                await _fetch_semaphore.acquire()
                _track_task(asyncio.create_task(_process_claimed_fetch_case(case_info)))
        except Exception as e:
            logger.error(f"Fetch poll loop error: {e}")

        try:
            await asyncio.wait_for(
                _wake_fetch_poll.wait(), timeout=QUEUE_POLL_INTERVAL_SECONDS
            )
        except asyncio.TimeoutError:
            pass
        _wake_fetch_poll.clear()


def _run_case_analysis_job(case_info: Dict) -> Dict:
    """Blocking analysis job used by async worker executor."""
    manager = get_auto_order_manager()
    case_ref = case_info.get("case_ref")
    case_id = case_info.get("id")
    board_date = case_info.get("board_date")

    order_context = manager._get_case_order_context(case_ref)
    order_link = order_context.get("order_link")

    if not order_link:
        manager.case_store.transition_lifecycle(
            case_ref,
            "analysis_failed_retryable",
            reason="No order link available for analysis",
            metadata={"source": "analysis_queue", "case_id": case_id},
            event_type="analysis_queue_no_link",
        )
        return {
            "case_ref": case_ref,
            "analysis_success": False,
            "error": "No order link available for analysis",
        }

    case_data = {
        "id": case_id,
        "case_ref": case_ref,
        "order_link": order_link,
        "board_date": board_date,
        "order_status": "linked",
    }
    result_template = {
        "case_id": case_id,
        "case_ref": case_ref,
        "download_success": True,
        "analysis_success": False,
        "order_link": order_link,
        "analysis_data": None,
        "error": None,
        "retry_attempts": [],
        "has_existing_order": True,
    }
    return manager._analyze_existing_order(case_data, result_template)


async def _process_claimed_analysis_case(case_info: Dict) -> None:
    case_ref = case_info["case_ref"]
    case_id = case_info.get("id")
    try:
        loop = asyncio.get_event_loop()
        result = await asyncio.wait_for(
            loop.run_in_executor(executor, _run_case_analysis_job, case_info),
            timeout=300.0,
        )

        if result.get("analysis_success"):
            logger.info(f"✅ Analysis completed for {case_ref}")
            board_doc_id = _resolve_board_doc_id(case_info)
            if board_doc_id:
                try:
                    await auto_map_case_to_users(board_doc_id, case_info)
                except Exception as mapping_error:
                    logger.error(
                        f"Error mapping users after analysis for {case_ref}: {mapping_error}"
                    )
            else:
                logger.warning(
                    f"No board_assignment_ids for {case_ref} -- skipping user mapping"
                )
        else:
            error_msg = result.get("error") or "Analysis failed"
            get_auto_order_manager().case_store.transition_lifecycle(
                case_ref,
                "analysis_failed_retryable",
                reason=error_msg,
                metadata={"source": "analysis_poll_loop", "case_id": case_id},
                event_type="analysis_queue_failed",
            )
            logger.warning(f"⚠️ Analysis failed for {case_ref}: {error_msg}")

    except asyncio.TimeoutError:
        get_auto_order_manager().case_store.transition_lifecycle(
            case_ref,
            "analysis_failed_retryable",
            reason="Analysis worker timeout after 5 minutes",
            metadata={"source": "analysis_poll_loop", "case_id": case_id},
            event_type="analysis_queue_timeout",
        )
        logger.error(f"❌ Timeout while analyzing {case_ref}")
    except Exception as e:
        get_auto_order_manager().case_store.transition_lifecycle(
            case_ref,
            "analysis_failed_retryable",
            reason=str(e),
            metadata={"source": "analysis_poll_loop", "case_id": case_id},
            event_type="analysis_queue_exception",
        )
        logger.error(f"❌ Error analyzing {case_ref}: {e}")
    finally:
        _analysis_semaphore.release()


async def analysis_poll_loop():
    """Analysis counterpart of fetch_poll_loop: polls analysis_queued plus
    stale analysis_in_progress cases."""
    logger.info("🚀 Analysis poll loop started")

    while True:
        _write_poll_heartbeat("analysis_last_tick")
        try:
            ensure_firebase()
            case_store = get_auto_order_manager().case_store
            candidates = _query_claim_candidates(
                case_store,
                "analysis_queued",
                "analysis_in_progress",
                QUEUE_POLL_BATCH_SIZE,
                # Orders that were downloaded (manual upload, or a fetch path
                # that doesn't inline-analyse) but never explicitly queued
                # for analysis -- same self-feeding rationale as fetch_poll_loop's
                # backlog_statuses above.
                backlog_statuses=("fetch_succeeded",),
            )
            for case_data in candidates:
                case_ref = case_data.get("case_ref")
                if not case_ref:
                    continue
                claim = case_store.claim_for_processing(
                    case_ref,
                    "analysis_in_progress",
                    from_statuses={
                        case_data.get("_claim_from_status", "analysis_queued")
                    },
                    stale_after_minutes=STALE_IN_PROGRESS_MINUTES,
                    reason="Claimed by analysis poll loop",
                    event_type="analysis_claimed",
                )
                if not claim["applied"]:
                    continue
                case_info = {
                    "id": case_data.get("_doc_id"),
                    "case_ref": case_ref,
                    "board_date": case_data.get("board_date")
                    or case_data.get("latest_board_date"),
                    "board_assignment_ids": case_data.get("board_assignment_ids") or [],
                }
                await _analysis_semaphore.acquire()
                _track_task(
                    asyncio.create_task(_process_claimed_analysis_case(case_info))
                )
        except Exception as e:
            logger.error(f"Analysis poll loop error: {e}")

        try:
            await asyncio.wait_for(
                _wake_analysis_poll.wait(), timeout=QUEUE_POLL_INTERVAL_SECONDS
            )
        except asyncio.TimeoutError:
            pass
        _wake_analysis_poll.clear()


async def auto_map_case_to_users(case_id: str, case_info: Dict):
    """Automatically map case to users after order analysis completion"""
    try:
        # Initialize Firestore client
        db = firestore.client()

        # Get all users who have configured roles
        users_ref = db.collection("user-roles")
        user_docs = users_ref.stream()

        mapped_users = []

        for user_doc in user_docs:
            try:
                user_id = user_doc.id
                user_data = user_doc.to_dict()

                # Create UserRole object from stored data. Default matches
                # UserMatterMatcher.get_user_role_config's canonical default
                # (0.50, "lowered from 0.75 to match bill generation logic")
                # -- this call site had drifted from that and was silently
                # requiring a much higher bar for every user without an
                # explicit confidence_threshold stored.
                user_role = UserRole(
                    role_type=user_data.get("role_type"),
                    full_name=user_data.get("full_name"),
                    name_variations=user_data.get("name_variations", []),
                    pattern_keywords=user_data.get("pattern_keywords", []),
                    confidence_threshold=user_data.get("confidence_threshold", 0.50),
                )

                # Check if this case matches the user
                matcher = get_user_matter_matcher()
                user_matches = matcher.find_user_matters_for_case(
                    user_id, user_role, case_id
                )

                # Roadmap #9: matches that fell just short of the threshold
                # used to be discarded with no trace -- ask the user instead
                # of silently missing the matter assignment. One pending
                # confirmation per (user, case, source, field), same dedup
                # key shape as the accepted mapping below.
                near_misses = matcher.find_near_miss_matters_for_case(
                    user_id, user_role, case_id
                )
                for near_miss in near_misses:
                    pending_key = (
                        f"{user_id}_{case_id}_{near_miss.match_source}_"
                        f"{near_miss.match_field}"
                    )
                    db.collection("user-matter-pending-confirmations").document(
                        pending_key
                    ).set(
                        {
                            "user_id": user_id,
                            "case_id": case_id,
                            "case_ref": case_info.get("case_ref"),
                            "match_source": near_miss.match_source,
                            "match_field": near_miss.match_field,
                            "matched_text": near_miss.matched_text,
                            "confidence_score": near_miss.confidence_score,
                            "role_type": near_miss.role_type,
                            "board_date": near_miss.board_date,
                            "created_at": firestore.SERVER_TIMESTAMP,
                            "status": "pending",
                        },
                        merge=True,
                    )

                if user_matches:
                    # Store the mapping in user-case-mappings collection
                    for match in user_matches:
                        mapping_data = {
                            "user_id": user_id,
                            "case_id": case_id,
                            "case_ref": case_info.get("case_ref"),
                            "match_source": match.match_source,
                            "match_field": match.match_field,
                            "matched_text": match.matched_text,
                            "confidence_score": match.confidence_score,
                            "role_type": match.role_type,
                            "board_date": match.board_date,
                            "mapped_at": firestore.SERVER_TIMESTAMP,
                            "auto_mapped": True,
                        }

                        # Use composite key to prevent duplicates
                        mapping_key = f"{user_id}_{case_id}_{match.match_source}_{match.match_field}"
                        db.collection("user-case-mappings").document(mapping_key).set(
                            mapping_data, merge=True
                        )

                        mapped_users.append(
                            {
                                "user_id": user_id,
                                "role_type": user_role.role_type,
                                "confidence": match.confidence_score,
                            }
                        )

            except Exception as user_error:
                logger.error(
                    f"Error processing user {user_doc.id} for case mapping: {user_error}"
                )
                continue

        if mapped_users:
            logger.info(
                f"Case {case_info.get('case_ref')} mapped to {len(mapped_users)} users: {[u['user_id'] for u in mapped_users]}"
            )
        else:
            logger.info(f"No user matches found for case {case_info.get('case_ref')}")

    except Exception as e:
        logger.error(f"Error in auto_map_case_to_users: {e}")
        raise


@app.on_event("startup")
async def _start_poll_loops():
    """Every Cloud Run instance actively polls from the moment it comes up,
    rather than lazily starting workers on first enqueue (the old design's
    workers never actually started on an instance that only ever received
    enqueue requests routed to a different instance)."""
    _track_task(asyncio.create_task(fetch_poll_loop()))
    _track_task(asyncio.create_task(analysis_poll_loop()))


# Login/logout endpoints removed - using Firebase client-side authentication


@app.get("/", tags=["Root"])
async def read_root():
    return {
        "message": "Hello, World! 🚀 Billingonaire API is running with async order processing."
    }


@app.post("/upload-pdf", tags=["PDF Upload"])
async def upload_pdf(
    files: List[UploadFile] = File(...), current_user=Depends(require_admin)
):
    results: List[Dict[str, Any]] = []
    for file in files:
        if file.content_type != "application/pdf":
            results.append(
                {
                    "filename": file.filename,
                    "error": "Invalid file type. Only PDF files are allowed.",
                }
            )
            continue
        max_retries = 3
        for attempt in range(max_retries):
            try:
                logger.info(f"Starting upload processing for file: {file.filename}")
                board = Board()
                df = board.readFile(file.filename, file.file)
                record_count = len(df) if df is not None else 0
                logger.info(
                    f"PDF processed successfully. Records found: {record_count}"
                )

                if record_count > 0:
                    board.saveData(df)
                    logger.info(f"Data saved successfully for {file.filename}")

                    # Trigger async order processing for uploaded cases
                    await trigger_async_order_processing(df)

                    board_date = None
                    try:
                        board_date = (
                            str(df["board_date"].iloc[0])
                            if "board_date" in df.columns and len(df) > 0
                            else None
                        )
                    except Exception:
                        pass

                    results.append(
                        {
                            "filename": file.filename,
                            "message": "Data saved successfully - Order processing started in background",
                            "records_processed": record_count,
                            "board_date": board_date,
                        }
                    )
                else:
                    logger.warning(f"No records found in {file.filename}")
                    results.append(
                        {
                            "filename": file.filename,
                            "message": "No records found in PDF",
                            "records_processed": 0,
                        }
                    )
                break
            except ConnectionResetError as e:
                logger.error(f"ConnectionResetError on attempt {attempt + 1}: {str(e)}")
                if attempt < max_retries - 1:
                    await asyncio.sleep(1)
                    continue
                else:
                    results.append(
                        {
                            "filename": file.filename,
                            "error": "Connection was reset by the remote host. Please try again later.",
                        }
                    )
                    break
            except Exception as e:
                logger.error(f"Error processing {file.filename}: {str(e)}")
                logger.error("Stack trace:", exc_info=True)
                results.append({"filename": file.filename, "error": str(e)})
                break
    return {"results": results}


@app.post("/save-data", tags=["PDF Upload"])
async def save_data(data: dict, current_user=Depends(require_admin)):
    try:
        board = Board()
        df = pd.DataFrame(data["data"])
        board.saveData(df)
        return {"message": "Data saved successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/get-data", tags=["Data Retrieval"])
async def get_data(
    request: Request, current_user_with_profile=Depends(get_user_with_profile)
):
    try:
        search_criteria = await request.json()

        board = Board()

        # SECURITY: Apply AGP filter for non-admin users - strict enforcement
        uid = current_user_with_profile.get("uid")
        agp_filter = get_user_manager().get_user_agp_filter(
            uid
        )  # This will raise 403 if invalid

        data = board.getData(search_criteria, agp_filter)
        return data
    except Exception as e:
        logger.error(f"Error in data retrieval: {str(e)}")
        raise HTTPException(status_code=500, detail="Error retrieving data")


@app.get("/cases/lifecycle", tags=["Data Retrieval"])
async def get_cases_lifecycle(
    case_type: Optional[str] = Query(None),
    case_number: Optional[str] = Query(None),
    case_year: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    advocate_name: Optional[str] = Query(None),
    order_status: Optional[str] = Query(None),
    order_category: Optional[str] = Query(None),
    lifecycle_status: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    timeline_limit: int = Query(5, ge=0, le=50),
    current_user_with_profile=Depends(get_user_with_profile),
):
    """Return unified board, case, order, and lifecycle sections for each matter."""
    try:
        board = Board()
        search_criteria = {
            "caseType": case_type,
            "caseNumber": case_number,
            "caseYear": case_year,
            "startDate": start_date,
            "endDate": end_date,
            "advocateName": advocate_name,
            "orderStatus": order_status,
            "orderCategory": order_category,
        }

        uid = current_user_with_profile.get("uid")
        agp_filter = get_user_manager().get_user_agp_filter(uid)

        rows = board.getData(search_criteria, agp_filter)

        case_store = get_auto_order_manager().case_store
        items = []
        for row in rows:
            case_ref = row.get("case_ref")
            if not case_ref:
                case_ref = case_store.build_case_ref(
                    row.get("case_type"), row.get("case_no"), row.get("case_year")
                )

            case_details = case_store.get_case_details(case_ref) or {}
            resolved_lifecycle_status = (
                case_details.get("lifecycle_status")
                or case_store.map_legacy_order_status(
                    case_details.get("latest_order_status") or row.get("order_status")
                )
                or "board_ingested"
            )

            if lifecycle_status and resolved_lifecycle_status != lifecycle_status:
                continue

            timeline = list(case_details.get("lifecycle_events") or [])
            timeline_preview = timeline[-timeline_limit:] if timeline_limit else []

            items.append(
                {
                    "board": {
                        "id": row.get("id"),
                        "board_date": row.get("board_date"),
                        "serial_number": row.get("serial_number"),
                        "file_name": row.get("file_name"),
                        "petitioner_lawyer": row.get("petitioner_lawyer"),
                        "respondent_lawyer": row.get("respondent_lawyer"),
                    },
                    "case": {
                        "case_ref": case_ref,
                        "case_type": row.get("case_type"),
                        "case_no": row.get("case_no"),
                        "case_year": row.get("case_year"),
                        "petitioner": case_details.get("petitioner")
                        or row.get("order_petitioner"),
                        "respondent": case_details.get("respondent")
                        or row.get("order_respondent"),
                        "government_pleader": case_details.get("government_pleader")
                        or row.get("government_pleader")
                        or [],
                    },
                    "order": {
                        "status": case_details.get("latest_order_status")
                        or row.get("order_status")
                        or "not_linked",
                        "link": case_details.get("latest_order_link")
                        or row.get("order_link"),
                        "category": case_details.get("latest_order_category")
                        or row.get("order_category"),
                        "date": case_details.get("latest_order_date")
                        or row.get("order_date"),
                    },
                    "lifecycle": {
                        "status": resolved_lifecycle_status,
                        "updated_at": case_details.get("lifecycle_status_updated_at")
                        or row.get("lifecycle_status_updated_at"),
                        "timeline_preview": timeline_preview,
                        "event_count": len(timeline),
                    },
                }
            )

            if len(items) >= limit:
                break

        return {
            "items": items,
            "count": len(items),
            "filters": {
                "case_type": case_type,
                "case_number": case_number,
                "case_year": case_year,
                "order_status": order_status,
                "order_category": order_category,
                "lifecycle_status": lifecycle_status,
            },
        }

    except Exception as e:
        logger.error(f"Error building lifecycle view: {e}")
        raise HTTPException(status_code=500, detail="Error retrieving lifecycle data")


@app.get("/cases/{case_ref:path}/timeline", tags=["Data Retrieval"])
async def get_case_timeline(
    case_ref: str,
    limit: int = Query(50, ge=1, le=500),
    current_user=Depends(get_current_user),
):
    """Return full case details + lifecycle timeline for the case detail modal.

    Returns petitioner, respondent, orders, board_dates, and lifecycle_events
    so the frontend CaseDetailModal can display all sections without additional
    API calls.
    """
    try:
        normalized_case_ref = str(case_ref or "").strip().upper()
        if not normalized_case_ref:
            raise HTTPException(status_code=400, detail="case_ref is required")

        ensure_firebase()
        db = firestore.client()
        case_store = get_auto_order_manager().case_store
        case_details = case_store.get_case_details(normalized_case_ref)
        if not case_details:
            raise HTTPException(status_code=404, detail="Case not found")

        # Lifecycle events (paginated)
        all_events = list(case_details.get("lifecycle_events") or [])
        lifecycle_events = all_events[-limit:] if limit and limit > 0 else all_events

        # Board date records — batch-fetch from daily-boards using stored assignment IDs
        board_dates: list = []
        board_ids = case_details.get("board_assignment_ids") or []
        if board_ids:
            try:
                doc_refs = [
                    db.collection("daily-boards").document(bid)
                    for bid in board_ids[:50]
                ]
                for snap in db.get_all(doc_refs):
                    if snap.exists:
                        d = snap.to_dict() or {}
                        raw_bd = d.get("board_date")
                        if hasattr(raw_bd, "strftime"):
                            # Firestore Timestamp / DatetimeWithNanoseconds
                            bd = raw_bd.strftime("%Y-%m-%d")
                        else:
                            bd = str(raw_bd or "")
                            # Handle both ISO ("T") and space-separated formats
                            if "T" in bd:
                                bd = bd.split("T", 1)[0]
                            elif " " in bd:
                                bd = bd.split(" ", 1)[0]
                        board_dates.append(
                            {
                                "board_date": bd,
                                "board_doc_id": snap.id,
                                "respondent_lawyer": d.get("respondent_lawyer") or "",
                                "additional_respondent_lawyers": d.get(
                                    "additional_respondent_lawyers"
                                )
                                or [],
                                "petitioner_lawyer": d.get("petitioner_lawyer") or "",
                            }
                        )
            except Exception as _bd_err:
                logger.warning(
                    "get_case_timeline: board_dates fetch failed for %s: %s",
                    normalized_case_ref,
                    _bd_err,
                )

        lifecycle_status = (
            case_details.get("lifecycle_status")
            or case_store.map_legacy_order_status(
                case_details.get("latest_order_status")
            )
            or "board_ingested"
        )

        # Normalise each order's board_date field (may be a Timestamp from old data)
        raw_orders = case_details.get("orders") or []
        orders_out = []
        for o in raw_orders:
            if not isinstance(o, dict):
                continue
            o = dict(o)
            # Skip status-only entries that have neither an order_date nor an
            # order_link — these are internal tracking markers (e.g. order_failed
            # after exhausting sequence retries) that should not appear as rows
            # in the modal's appearances table.
            if not o.get("order_date") and not o.get("order_link"):
                continue
            raw_bd = o.get("board_date")
            if raw_bd is not None:
                if hasattr(raw_bd, "strftime"):
                    o["board_date"] = raw_bd.strftime("%Y-%m-%d")
                else:
                    bd_str = str(raw_bd)
                    if "T" in bd_str:
                        bd_str = bd_str.split("T", 1)[0]
                    elif " " in bd_str:
                        bd_str = bd_str.split(" ", 1)[0]
                    o["board_date"] = bd_str
            orders_out.append(o)

        return {
            "case_ref": normalized_case_ref,
            "lifecycle_status": lifecycle_status,
            "petitioner": case_details.get("petitioner") or "",
            "respondent": case_details.get("respondent") or "",
            "government_pleader": case_details.get("government_pleader") or [],
            "orders": orders_out,
            "board_dates": board_dates,
            "lifecycle_events": lifecycle_events,
            # backward-compat aliases kept for any callers expecting the old shape
            "timeline": lifecycle_events,
            "count": len(lifecycle_events),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error fetching case timeline for %s: %s", case_ref, e)
        raise HTTPException(status_code=500, detail="Error retrieving case timeline")


@app.get("/debug/auth-test")
async def auth_test(current_user=Depends(get_current_user)):
    return {"message": "Authentication successful", "user_id": current_user.get("uid")}


@app.get("/debug/simple-db-check", tags=["Admin"])
async def simple_database_check(current_user=Depends(require_admin)):
    try:
        board = Board()

        # Get all documents
        all_docs = list(board.db.collection("daily-boards").limit(10).stream())

        # Get sample documents
        sample_docs = []
        case_years_found = []

        for doc in all_docs:
            doc_data = doc.to_dict()
            case_year = doc_data.get("case_year")
            case_years_found.append(case_year)

            # Convert datetime to string for JSON serialization
            if "board_date" in doc_data and hasattr(doc_data["board_date"], "strftime"):
                doc_data["board_date"] = doc_data["board_date"].strftime("%Y-%m-%d")

            sample_docs.append(
                {
                    "document_id": doc.id,
                    "case_year": case_year,
                    "case_year_type": str(type(case_year)),
                    "board_date": doc_data.get("board_date"),
                    "all_fields": list(doc_data.keys()),
                }
            )

        # Test query for case_year = "2025"
        test_query = board.db.collection("daily-boards").where(
            "case_year", "==", "2025"
        )
        test_results = list(test_query.stream())

        return {
            "total_documents": len(all_docs),
            "case_years_found": case_years_found,
            "test_query_for_2025_results": len(test_results),
            "sample_documents": sample_docs[:3],
            "database_status": "connected" if all_docs else "empty",
        }
    except Exception as e:
        return {"error": str(e), "database_status": "error"}


# User management endpoints
@app.get("/user/profile", tags=["User Management"])
async def get_user_profile(current_user_with_profile=Depends(get_user_with_profile)):
    """Get current user's profile"""
    return current_user_with_profile["profile"]


@app.post("/user/profile", tags=["User Management"])
async def create_or_update_profile(
    profile_data: dict, current_user=Depends(get_current_user)
):
    """Create or update user profile (self-service - no role changes)"""
    uid = current_user.get("uid")
    email = current_user.get("email")

    # Check if this is the initial admin user
    if email == "deshpande.mak@gmail.com":
        # Create admin profile directly
        return get_user_manager().create_user_profile(
            uid=uid, email=email, role="admin", full_name=profile_data.get("full_name")
        )

    # SECURITY: Remove role from self-service updates to prevent privilege escalation
    safe_updates = {"full_name": profile_data.get("full_name")}

    # Check if profile exists
    try:
        existing_profile = get_user_manager().get_user_profile(uid)
        if existing_profile.get("needs_setup"):
            # For new profiles, create user with legal category
            return get_user_manager().create_user_profile(
                uid=uid,
                email=email,
                role="user",
                legal_category="assistant_government_pleader",
                full_name=profile_data.get("full_name"),
            )
        else:
            # Update existing profile with safe fields only
            return get_user_manager().update_user_profile(uid, safe_updates)
    except Exception as e:
        logger.warning(f"Profile update failed, creating new profile: {e}")
        # Create new profile with user role and legal category
        return get_user_manager().create_user_profile(
            uid=uid,
            email=email,
            role="user",
            legal_category="assistant_government_pleader",
            full_name=profile_data.get("full_name"),
        )


@app.post("/user/change-password", tags=["User Management"])
async def change_password(password_data: dict, current_user=Depends(get_current_user)):
    """Change user password"""
    try:
        uid = current_user.get("uid")
        new_password = password_data.get("new_password")

        if not new_password or len(new_password) < 6:
            raise HTTPException(
                status_code=400, detail="Password must be at least 6 characters"
            )

        # Update password in Firebase Auth
        auth.update_user(uid, password=new_password)

        logger.info(f"Password changed for user {uid}")
        return {"message": "Password changed successfully"}

    except Exception as e:
        logger.error(f"Error changing password: {str(e)}")
        raise HTTPException(status_code=500, detail="Error changing password")


@app.get("/admin/users", tags=["Admin"])
async def list_users(
    role_filter: str = Query(None, description="Filter by role: admin or agp"),
    current_user=Depends(require_admin_active),
):
    """List all users (admin only)"""
    return get_user_manager().list_users(role_filter)


@app.post("/admin/user/{target_uid}/role", tags=["Admin"])
async def update_user_role(
    target_uid: str, role_data: dict, current_user=Depends(require_admin_active)
):
    """Update user role and profile information (admin only)"""
    admin_uid = current_user.get("uid")
    return get_user_manager().admin_update_user_profile(
        target_uid, role_data, admin_uid
    )


@app.post("/admin/setup-initial-admin", tags=["Admin"])
async def setup_initial_admin():
    """Set up the hardcoded INITIAL_ADMIN_EMAIL account as administrator.

    Not an escalation vector (it can only ever promote that one address, and
    only if the account already exists), but it used to return the admin's
    full user document to any anonymous caller. Bootstrap only needs to say
    whether it worked.
    """
    result = get_user_manager().setup_initial_admin() or {}
    return {
        "success": True,
        "message": result.get("message", "Initial admin is configured."),
    }


@app.get("/admin/active-users", tags=["Admin"])
async def get_active_users_for_bills(current_user=Depends(require_admin_active)):
    """Get list of active user names for bill generation (admin only)"""
    return {"user_names": get_user_manager().get_active_user_names()}


@app.get("/admin/available-roles", tags=["Admin"])
async def get_available_roles(current_user=Depends(require_admin_active)):
    """Get available user roles for admin interface"""
    return {"roles": get_user_manager().get_available_roles()}


@app.get("/admin/available-legal-categories", tags=["Admin"])
async def get_available_legal_categories(current_user=Depends(require_admin_active)):
    """Get available legal categories for admin interface"""
    return {"legal_categories": get_user_manager().get_available_legal_categories()}


@app.get("/admin/firebase-users", tags=["Admin"])
async def list_firebase_auth_users(current_user=Depends(require_admin_active)):
    """List all users from Firebase Authentication"""
    return get_user_manager().list_firebase_auth_users()


@app.get("/admin/unsynced-users", tags=["Admin"])
async def get_unsynced_firebase_users(current_user=Depends(require_admin_active)):
    """Get Firebase Auth users that don't have Firestore profiles"""
    return get_user_manager().get_firebase_auth_users_not_in_firestore()


@app.post("/admin/sync-firebase-users", tags=["Admin"])
async def sync_firebase_users(current_user=Depends(require_admin_active)):
    """Sync Firebase Auth users to Firestore database"""
    uid = current_user.get("uid")
    return get_user_manager().sync_firebase_users_to_firestore(uid)


@app.post("/admin/create-user", tags=["Admin"])
async def create_new_user(user_data: dict, current_user=Depends(require_admin_active)):
    """Create a new user with default password (admin only)"""
    try:
        admin_uid = current_user.get("uid")
        email = user_data.get("email")
        role = user_data.get("role", "user")
        legal_category = user_data.get("legal_category", "assistant_government_pleader")
        full_name = user_data.get("full_name", "")

        if not email:
            raise HTTPException(status_code=400, detail="Email is required")

        # Create user in Firebase Auth with default password
        try:
            firebase_user = auth.create_user(
                email=email,
                password="password123",  # Default password
                email_verified=False,
            )

            # Create user profile in Firestore
            user_profile = get_user_manager().create_user_profile(
                uid=firebase_user.uid,
                email=email,
                role=role,
                legal_category=(
                    legal_category
                    if get_user_manager().is_legal_professional(role)
                    else None
                ),
                full_name=full_name,
            )

            logger.info(f"Admin {admin_uid} created new user {email} with role {role}")

            return {
                "message": "User created successfully",
                "user": user_profile,
                "default_password": "password123",
                "note": "User should change password on first login",
            }

        except auth.EmailAlreadyExistsError:
            raise HTTPException(
                status_code=400, detail="Email already exists in the system"
            )
        except Exception as e:
            logger.error(f"Error creating Firebase user: {str(e)}")
            raise HTTPException(
                status_code=500, detail=f"Error creating user account: {str(e)}"
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in create_new_user: {str(e)}")
        raise HTTPException(status_code=500, detail="Error creating user")


# Dashboard endpoints (with authentication)
dashboard_data = None


def get_dashboard_data():
    global dashboard_data
    if dashboard_data is None:
        ensure_firebase()  # Ensure Firebase is initialized before creating DashboardData
        dashboard_data = DashboardData()
    return dashboard_data


@app.get("/dashboard/weekly-status")
async def dashboard_weekly_status(
    start_date: str = Query(None),
    end_date: str = Query(None),
    current_user_with_profile=Depends(get_user_with_profile),
    response: Response = None,
):
    # SECURITY: Get AGP filter for the user - strict enforcement
    uid = current_user_with_profile.get("uid")
    agp_filter = get_user_manager().get_user_agp_filter(
        uid
    )  # This will raise 403 if invalid

    data = await get_dashboard_data().get_weekly_status(
        start_date, end_date, agp_filter
    )
    if response is not None:
        response.headers["Cache-Control"] = "public, max-age=120"
    return JSONResponse(content=data)


@app.get("/dashboard/agp-stats")
async def dashboard_agp_stats(
    agp_name: str = Query(None),
    current_user_with_profile=Depends(get_user_with_profile),
    response: Response = None,
):
    # SECURITY: Get AGP filter for the user - strict enforcement
    uid = current_user_with_profile.get("uid")
    agp_filter = get_user_manager().get_user_agp_filter(
        uid
    )  # This will raise 403 if invalid

    # For AGP users, use their assigned AGP name; for admins, use query parameter
    target_agp = agp_filter or agp_name

    data = await get_dashboard_data().get_agp_stats(target_agp, agp_filter)
    if response is not None:
        response.headers["Cache-Control"] = "public, max-age=120"
    return JSONResponse(content=data)


@app.get("/dashboard/monthly-avg")
async def dashboard_monthly_avg(
    year: str = Query(None),
    current_user_with_profile=Depends(get_user_with_profile),
    response: Response = None,
):
    # SECURITY: Get AGP filter for the user - strict enforcement
    uid = current_user_with_profile.get("uid")
    agp_filter = get_user_manager().get_user_agp_filter(
        uid
    )  # This will raise 403 if invalid

    data = await get_dashboard_data().get_monthly_avg(year, agp_filter)
    if response is not None:
        response.headers["Cache-Control"] = "public, max-age=120"
    return JSONResponse(content=data)


@app.get("/dashboard/matters-by-date-range")
async def dashboard_matters_by_date_range(
    start_date: str = Query(
        None, description="Start date (YYYY-MM-DD) - defaults to last 5 days"
    ),
    end_date: str = Query(
        None, description="End date (YYYY-MM-DD) - defaults to today"
    ),
    current_user_with_profile=Depends(get_user_with_profile),
    response: Response = None,
):
    """Get total matters by date range with average for bar chart + line visualization"""
    # SECURITY: Get AGP filter for the user - strict enforcement
    uid = current_user_with_profile.get("uid")
    agp_filter = get_user_manager().get_user_agp_filter(
        uid
    )  # This will raise 403 if invalid

    data = await get_dashboard_data().get_matters_by_date_range(
        start_date, end_date, agp_filter
    )
    if response is not None:
        response.headers["Cache-Control"] = "public, max-age=120"
    return JSONResponse(content=data)


@app.get("/dashboard/agp-distribution-weekly")
async def dashboard_agp_distribution_weekly(
    current_user_with_profile=Depends(get_user_with_profile),
):
    """Get AGP distribution for current week (Monday to current date)"""
    # SECURITY: Get AGP filter for the user - strict enforcement
    uid = current_user_with_profile.get("uid")
    agp_filter = get_user_manager().get_user_agp_filter(
        uid
    )  # This will raise 403 if invalid

    data = await get_dashboard_data().get_agp_distribution_weekly(agp_filter)
    return JSONResponse(content=data)


@app.get("/dashboard/agp-distribution-monthly")
async def dashboard_agp_distribution_monthly(
    current_user_with_profile=Depends(get_user_with_profile),
):
    """Get AGP distribution for current month to date"""
    # SECURITY: Get AGP filter for the user - strict enforcement
    uid = current_user_with_profile.get("uid")
    agp_filter = get_user_manager().get_user_agp_filter(
        uid
    )  # This will raise 403 if invalid

    data = await get_dashboard_data().get_agp_distribution_monthly(agp_filter)
    return JSONResponse(content=data)


@app.get("/dashboard/agp-distribution-yearly")
async def dashboard_agp_distribution_yearly(
    current_user_with_profile=Depends(get_user_with_profile),
):
    """Get AGP distribution for current year to date"""
    # SECURITY: Get AGP filter for the user - strict enforcement
    uid = current_user_with_profile.get("uid")
    agp_filter = get_user_manager().get_user_agp_filter(
        uid
    )  # This will raise 403 if invalid

    data = await get_dashboard_data().get_agp_distribution_yearly(agp_filter)
    return JSONResponse(content=data)


@app.get("/dashboard/board-date-summary")
async def dashboard_board_date_summary(
    start_date: str = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(None, description="End date (YYYY-MM-DD)"),
    year: Optional[int] = Query(None, description="Year filter, e.g. 2026"),
    quarter: Optional[int] = Query(None, description="Quarter filter (1-4)"),
    limit: int = Query(180, ge=1, le=1000),
    current_user_with_profile=Depends(get_user_with_profile),
    response: Response = None,
):
    """Get board-date summary with case counts and distinct pleader counts."""
    uid = current_user_with_profile.get("uid")
    agp_filter = get_user_manager().get_user_agp_filter(uid)

    data = await get_dashboard_data().get_board_date_summary(
        start_date=start_date,
        end_date=end_date,
        year=year,
        quarter=quarter,
        limit=limit,
        agp_filter=agp_filter,
    )
    if response is not None:
        response.headers["Cache-Control"] = "public, max-age=120"
    return JSONResponse(content=data)


@app.get("/dashboard/board-date-agp-distribution")
async def dashboard_board_date_agp_distribution(
    board_dates: List[str] = Query(..., description="One or more board_date values"),
    current_user_with_profile=Depends(get_user_with_profile),
):
    """Get AGP-wise case distribution for selected board dates."""
    uid = current_user_with_profile.get("uid")
    agp_filter = get_user_manager().get_user_agp_filter(uid)

    data = await get_dashboard_data().get_agp_distribution_for_board_dates(
        board_dates=board_dates,
        agp_filter=agp_filter,
    )
    return JSONResponse(content=data)


@app.get("/dashboard/board-date-cases")
async def dashboard_board_date_cases(
    board_dates: List[str] = Query(..., description="One or more board_date values"),
    limit: int = Query(2000, ge=1, le=5000),
    current_user_with_profile=Depends(get_user_with_profile),
):
    """Get case rows for selected board dates."""
    uid = current_user_with_profile.get("uid")
    agp_filter = get_user_manager().get_user_agp_filter(uid)

    data = await get_dashboard_data().get_cases_for_board_dates(
        board_dates=board_dates,
        limit=limit,
        agp_filter=agp_filter,
    )
    return JSONResponse(content=data)


# ML Enhancement endpoints
@app.get("/ml/status")
async def get_ml_enhancement_status(current_user=Depends(get_current_user)):
    """Get status of ML enhancement capabilities"""
    try:
        board = Board()
        if hasattr(board, "ml_parser") and board.ml_parser:
            status = board.ml_parser.get_enhancement_status()
            status["ml_parser_available"] = True
            status[
                "message"
            ] = "ML Enhanced Parser is active and improving PDF processing quality"
        else:
            status = {
                "ml_parser_available": False,
                "capabilities": {
                    "enhanced_preprocessing": False,
                    "ner": False,
                    "fuzzy_matching": False,
                    "learning": False,
                    "advanced_fuzzy": False,
                },
                "message": "ML Enhanced Parser not available - using standard PDF processing",
            }
        return JSONResponse(content=status)
    except Exception as e:
        logger.error(f"Error fetching ML status: {e}")
        raise HTTPException(
            status_code=500, detail="Failed to fetch ML enhancement status"
        )


@app.post("/ml/learn-from-correction")
async def learn_from_correction(
    correction_data: dict, current_user=Depends(get_current_user)
):
    """Allow users to provide corrections for ML learning"""
    try:
        board = Board()
        if hasattr(board, "ml_parser") and board.ml_parser:
            board.ml_parser.learn_from_correction(
                filename=correction_data.get("filename", ""),
                original_extraction=correction_data.get("original_extraction", ""),
                corrected_extraction=correction_data.get("corrected_extraction", ""),
                user_feedback=correction_data.get("user_feedback", {}),
            )
            return JSONResponse(
                content={"message": "Learning data stored successfully"}
            )
        else:
            return JSONResponse(
                content={"message": "ML Enhanced Parser not available for learning"}
            )
    except Exception as e:
        logger.error(f"Error storing learning data: {e}")
        raise HTTPException(status_code=500, detail="Failed to store learning data")


# Court integration endpoints
court_scraper = None
order_manager = None


def get_court_scraper():
    global court_scraper
    if court_scraper is None:
        court_scraper = BombayHighCourtScraper()
    return court_scraper


def get_order_manager():
    global order_manager
    if order_manager is None:
        ensure_firebase()  # Ensure Firebase is initialized before creating OrderManager
        order_manager = OrderManager()
    return order_manager


def _normalize_iso_date(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    if "T" in raw:
        raw = raw.split("T", 1)[0]
    return raw


class OrderLinkAnalysisRequest(BaseModel):
    url: str
    persist_result: bool = False


def _serialize_order_analysis_result(
    filename: str,
    analysis_result,
    analysis_id: Optional[str] = None,
    source_url: Optional[str] = None,
    download_metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    response_data: Dict[str, Any] = {
        "filename": filename,
        "order_category": analysis_result.order_category,
        "category_confidence": round(analysis_result.category_confidence, 3),
        "order_date": analysis_result.order_date,
        "cases": [
            {
                "case_type": case.case_type,
                "case_number": case.case_number,
                "case_year": case.case_year,
                "petitioner": case.petitioner,
                "respondent": case.respondent,
                "government_pleader": case.government_pleader,
            }
            for case in analysis_result.cases
        ],
        "summary": {
            "total_cases": len(analysis_result.cases),
        },
    }

    if analysis_id is not None:
        response_data["analysis_id"] = analysis_id
    if source_url is not None:
        response_data["source_url"] = source_url
    if download_metadata is not None:
        response_data["download_metadata"] = download_metadata

    return response_data


def _derive_pdf_filename(source_url: str) -> str:
    parsed = urlparse(source_url)
    basename = posixpath.basename(parsed.path or "") or "order.pdf"
    if not basename.lower().endswith(".pdf"):
        basename = f"{basename}.pdf"
    return basename


def _download_pdf_from_url(source_url: str) -> Dict[str, Any]:
    response = requests.get(source_url, timeout=60)
    response.raise_for_status()

    file_content = response.content or b""
    if not file_content:
        raise ValueError("Downloaded file is empty")

    content_type = (response.headers.get("content-type") or "").lower()
    is_pdf = (
        source_url.lower().endswith(".pdf")
        or "application/pdf" in content_type
        or file_content.startswith(b"%PDF")
    )
    if not is_pdf:
        raise ValueError(
            f"URL did not return a PDF document. Content-Type was '{content_type or 'unknown'}'"
        )

    return {
        "filename": _derive_pdf_filename(source_url),
        "file_content": file_content,
        "metadata": {
            "source_url": source_url,
            "resolved_url": response.url,
            "content_type": content_type,
            "content_length": len(file_content),
            "status_code": response.status_code,
        },
    }


def _get_cached_case_details_payload(case_ref: str) -> Optional[Dict]:
    normalized_case_ref = str(case_ref or "").strip().upper()
    if not normalized_case_ref:
        return None

    case_details = (
        get_auto_order_manager().case_store.get_case_details(normalized_case_ref) or {}
    )
    if not case_details:
        return None

    petitioner = str(case_details.get("petitioner") or "").strip()
    respondent = str(case_details.get("respondent") or "").strip()
    orders = case_details.get("orders") or []

    if not petitioner and not respondent and not orders:
        return None

    return {
        "status": "found",
        "source": "case_store_cached",
        "case_ref": normalized_case_ref,
        "case_number": normalized_case_ref,
        "petitioner": petitioner,
        "respondent": respondent,
        "latest_board_date": case_details.get("latest_board_date"),
        "latest_order_link": case_details.get("latest_order_link"),
        "orders_count": len(orders),
    }


def _get_cached_case_orders_payload(
    case_ref: str, date: Optional[str]
) -> Optional[Dict]:
    normalized_case_ref = str(case_ref or "").strip().upper()
    if not normalized_case_ref:
        return None

    case_details = (
        get_auto_order_manager().case_store.get_case_details(normalized_case_ref) or {}
    )
    orders = case_details.get("orders") or []
    if not isinstance(orders, list) or not orders:
        return None

    requested_date = _normalize_iso_date(date)
    normalized_orders = []
    for item in orders:
        if not isinstance(item, dict):
            continue

        board_date = _normalize_iso_date(item.get("board_date"))
        order_date = _normalize_iso_date(item.get("order_date"))
        if requested_date and requested_date not in {board_date, order_date}:
            continue

        order_link = str(item.get("order_link") or "").strip()
        if not order_link:
            continue

        normalized_orders.append(
            {
                "listing_date": board_date or order_date,
                "download_url": order_link,
                "order_description": item.get("order_filename")
                or item.get("order_category")
                or "Cached order",
                "order_status": item.get("order_status"),
                "order_source": item.get("order_source")
                or item.get("cache_validation_source")
                or "case_store_cached",
            }
        )

    if not normalized_orders:
        return None

    petitioner = str(case_details.get("petitioner") or "").strip() or None
    respondent = str(case_details.get("respondent") or "").strip() or None
    title: Optional[str] = None
    if petitioner and respondent:
        title = f"{petitioner} against {respondent}"
    elif petitioner or respondent:
        title = petitioner or respondent

    case_orders = [
        {
            "date": o.get("listing_date"),
            "download_link": o.get("download_url"),
        }
        for o in normalized_orders
        if o.get("download_url")
    ]

    return {
        "status": "found",
        "source": "case_store_cached",
        "case_ref": normalized_case_ref,
        "date": requested_date,
        "case_summary": None,
        "petitioner": petitioner,
        "respondent": respondent,
        "title": title,
        "case_orders": case_orders,
        "case_details": {
            "case_number": normalized_case_ref,
            "petitioner_name": petitioner,
            "respondent_name": respondent,
        },
        "court_orders": normalized_orders,
    }


@app.get("/court/case-details", tags=["Case Status"])
async def get_case_details(
    case_ref: str = Query(..., description="Case reference like 'WP/294/2025'"),
    bench: str = Query(
        "mumbai",
        description="Court bench: mumbai, mumbai_appellate, aurangabad, nagpur, goa",
    ),
    current_user=Depends(get_current_user),
):
    """
    Fetch case details from Bombay High Court
    Example: /court/case-details?case_ref=WP/294/2025&bench=mumbai
    """
    try:
        cached_payload = _get_cached_case_details_payload(case_ref)
        if cached_payload:
            return JSONResponse(content=cached_payload)

        case_details = get_court_scraper().get_case_details(case_ref, bench)
        return JSONResponse(content=case_details)
    except Exception as e:
        logger.error(f"Error fetching case details for {case_ref}: {e}")
        return JSONResponse(
            status_code=500,
            content={
                "error": f"Failed to fetch case details: {str(e)}",
                "case_ref": case_ref,
            },
        )


@app.get("/court/case-orders", tags=["Case Orders"])
async def get_case_orders(
    case_ref: str = Query(..., description="Case reference like 'WP/294/2025'"),
    date: str = Query(None, description="Specific date in YYYY-MM-DD format"),
    bench: str = Query(
        "mumbai",
        description="Court bench: mumbai, mumbai_appellate, aurangabad, nagpur, goa",
    ),
    current_user=Depends(get_current_user),
):
    """
    Fetch case orders from Bombay High Court for a specific case and date
    Example: /court/case-orders?case_ref=WP/294/2025&date=2025-01-03
    """
    try:
        cached_payload = _get_cached_case_orders_payload(case_ref, date)
        if cached_payload:
            response_payload = {
                "case_ref": case_ref,
                "date": date,
                "orders": cached_payload.get("court_orders", []),
            }
            response_payload.update(cached_payload)
            return JSONResponse(content=response_payload)

        case_orders = get_court_scraper().get_case_orders(case_ref, date, bench)
        if isinstance(case_orders, dict):
            response_payload = {
                "case_ref": case_ref,
                "date": date,
                "orders": case_orders.get("court_orders", []),
            }
            response_payload.update(case_orders)
            return JSONResponse(content=response_payload)

        return JSONResponse(
            content={"case_ref": case_ref, "date": date, "orders": case_orders}
        )
    except Exception as e:
        logger.error(f"Error fetching case orders for {case_ref}: {e}")
        return JSONResponse(
            status_code=500,
            content={
                "error": f"Failed to fetch case orders: {str(e)}",
                "case_ref": case_ref,
            },
        )


@app.post("/court/batch-case-lookup", tags=["Case Status"])
async def batch_case_lookup(
    case_refs: List[str],
    bench: str = Query(
        "mumbai",
        description="Court bench: mumbai, mumbai_appellate, aurangabad, nagpur, goa",
    ),
    current_user=Depends(get_current_user),
):
    """
    Fetch case details for multiple cases in batch
    Useful for getting court data for multiple cases from your billing records
    """
    try:
        results = []
        for case_ref in case_refs:
            case_details = _get_cached_case_details_payload(case_ref)
            if not case_details:
                case_details = get_court_scraper().get_case_details(case_ref, bench)
            results.append(
                {
                    "case_ref": case_ref,
                    "details": case_details,
                    "timestamp": pd.Timestamp.now().isoformat(),
                }
            )

        return JSONResponse(
            content={"total_cases": len(case_refs), "results": results, "bench": bench}
        )
    except Exception as e:
        logger.error(f"Error in batch case lookup: {e}")
        return JSONResponse(
            status_code=500,
            content={
                "error": f"Batch lookup failed: {str(e)}",
                "total_cases": len(case_refs),
            },
        )


# Order Management endpoints
@app.get("/orders/cases-without-orders", tags=["Order Management"])
async def get_cases_without_orders(
    limit: int = Query(100, description="Number of cases to return"),
    offset: int = Query(0, description="Pagination offset"),
    current_user=Depends(get_current_user),
):
    """
    Get cases from board data that don't have linked orders
    Used for order management interface
    """
    try:
        result = get_order_manager().get_cases_without_orders(limit, offset)
        return JSONResponse(content=result)
    except Exception as e:
        logger.error(f"Error fetching cases without orders: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to fetch cases: {str(e)}"}
        )


@app.post("/orders/create-link", tags=["Order Management"])
async def create_order_link(request: Request, current_user=Depends(get_current_user)):
    """
    Create or update an order link for a case.

    Body: case_id, status, order_link, order_text, court_bench, notes.
    When an order_link is supplied the order is analysed automatically, via the
    same _analyze_existing_order path used by the analysis queue and the
    per-row Analyse button.
    """
    try:
        order_data = await request.json()
        case_id = order_data.get("case_id")

        if not case_id:
            return JSONResponse(
                status_code=400, content={"error": "case_id is required"}
            )

        result = get_order_manager().create_order_link(case_id, order_data)

        order_link = order_data.get("order_link")
        if result.get("success") and order_link:
            try:
                db = firestore.client()
                case_doc = db.collection("daily-boards").document(case_id).get()
                if case_doc.exists:
                    case_data = case_doc.to_dict()
                    case_ref = (
                        f"{case_data.get('case_type')}/"
                        f"{case_data.get('case_no')}/{case_data.get('case_year')}"
                    )
                    analysis = get_auto_order_manager()._analyze_existing_order(
                        {
                            "id": case_id,
                            "case_ref": case_ref,
                            "order_link": order_link,
                            "board_date": case_data.get("board_date"),
                        },
                        {
                            "case_id": case_id,
                            "case_ref": case_ref,
                            "download_success": True,
                            "analysis_success": False,
                            "order_link": order_link,
                            "analysis_data": None,
                            "error": None,
                            "retry_attempts": [],
                            "has_existing_order": True,
                        },
                    )
                    if analysis.get("analysis_success"):
                        result["analysis_completed"] = True
                        result["analysis_message"] = "Order analysed successfully"
                    else:
                        result["analysis_completed"] = False
                        result["analysis_message"] = analysis.get("error")
            except Exception as analysis_error:
                logger.warning(
                    "create-link: auto-analysis failed for case_id=%s: %s",
                    case_id,
                    analysis_error,
                )
                result["analysis_completed"] = False
                result["analysis_message"] = str(analysis_error)

        return JSONResponse(content=result)
    except Exception as e:
        logger.error(f"Error creating order link: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to create order link: {str(e)}"}
        )


@app.put("/orders/update-status", tags=["Order Management"])
async def update_order_status(
    case_id: str = Query(..., description="Case document ID"),
    status: str = Query(
        ...,
        description="Order status: linked, analysed, order_failed, order_analysis_failed, manually_uploaded, not_linked",
    ),
    notes: str = Query("", description="Optional notes"),
    current_user=Depends(require_admin),
):
    """Update the status of an order (admin only).

    Was gated on get_current_user, so any signed-in AGP could flip any case
    in the system to analysed/not_linked -- which changes what appears on
    other people's bills. Every sibling mutation (/cases/{ref}/reset,
    /admin/orders/{id}/override, /cases/{ref}/manual-override) already
    requires admin; this one was simply missed.
    """
    try:
        result = get_order_manager().update_order_status(case_id, status, notes)
        return JSONResponse(content=result)
    except Exception as e:
        logger.error(f"Error updating order status: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to update status: {str(e)}"}
        )


@app.get("/orders/by-status", tags=["Order Management"])
async def get_orders_by_status(
    status: str = Query(..., description="Order status to filter by"),
    limit: int = Query(100, description="Maximum number of orders"),
    current_user=Depends(get_current_user),
):
    """Get all orders with a specific status"""
    try:
        orders = get_order_manager().get_orders_by_status(status, limit)
        return JSONResponse(content={"orders": orders, "count": len(orders)})
    except Exception as e:
        logger.error(f"Error fetching orders by status: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to fetch orders: {str(e)}"}
        )


@app.get("/orders/case-details/{case_id}", tags=["Order Management"])
async def get_case_with_order_info(
    case_id: str, current_user=Depends(get_current_user)
):
    """Get complete case information including order status"""
    try:
        result = get_order_manager().get_case_with_order_info(case_id)
        return JSONResponse(content=result)
    except Exception as e:
        logger.error(f"Error fetching case with order info: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to fetch case details: {str(e)}"},
        )


# ============================================
# ORDER DOCUMENT ANALYSIS ENDPOINTS
# ============================================


@app.post("/analyze-order", tags=["Order Analysis"])
async def analyze_order_document(
    file: UploadFile = File(...), current_user=Depends(get_current_user)
):
    """
    Analyze a court order document to extract:
    - Order category (ADJOURNED/HEARD & ADJOURNED/DISPOSED OFF)
    - Petitioner and respondent names
    - AGP names and dates
    - Key phrases and next hearing dates
    """
    try:
        filename = file.filename or "uploaded-order.pdf"

        # Validate file type
        if not filename.lower().endswith(".pdf"):
            return JSONResponse(
                status_code=400,
                content={"error": "Only PDF files are supported for order analysis"},
            )

        # Read file content
        file_content = await file.read()

        if len(file_content) == 0:
            return JSONResponse(
                status_code=400, content={"error": "Uploaded file is empty"}
            )

        logger.info(f"Starting order analysis for file: {filename}")

        # Analyze the order document
        analysis_result = get_order_analyzer().analyze_order_document(
            filename, file_content
        )

        # DEBUG: Log the actual category being returned
        logger.info(f"🔍 CATEGORY DEBUG for {filename}:")
        logger.info(f"   order_category: '{analysis_result.order_category}'")
        logger.info(f"   category_confidence: {analysis_result.category_confidence}")

        # Save analysis result to database
        doc_id = get_order_analyzer().save_analysis_result(filename, analysis_result)

        response_data = _serialize_order_analysis_result(
            filename=filename,
            analysis_result=analysis_result,
            analysis_id=doc_id,
        )

        logger.info(f"Order analysis completed successfully for {filename}")
        return JSONResponse(content=response_data)

    except HTTPException as he:
        logger.error(f"HTTP error in order analysis: {he.detail}")
        return JSONResponse(status_code=he.status_code, content={"error": he.detail})
    except Exception as e:
        logger.error(f"Unexpected error in order analysis: {str(e)}")
        return JSONResponse(
            status_code=500, content={"error": f"Order analysis failed: {str(e)}"}
        )


@app.post("/admin/order-analysis/from-link", tags=["Order Analysis"])
async def analyze_order_document_from_link(
    request: OrderLinkAnalysisRequest,
    current_user: dict = Depends(require_admin_active),
):
    """Download a PDF from the provided URL and run the existing order analyzer on it."""
    _ = current_user
    try:
        download = _download_pdf_from_url(request.url)
        filename = download["filename"]
        file_content = download["file_content"]

        analysis_result = get_order_analyzer().analyze_order_document(
            filename, file_content
        )
        analysis_id = None
        if request.persist_result:
            analysis_id = get_order_analyzer().save_analysis_result(
                filename, analysis_result
            )

        response_data = _serialize_order_analysis_result(
            filename=filename,
            analysis_result=analysis_result,
            analysis_id=analysis_id,
            source_url=request.url,
            download_metadata=download["metadata"],
        )
        response_data["persisted"] = bool(request.persist_result)
        return JSONResponse(content=response_data)
    except requests.exceptions.RequestException as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to download order PDF from URL: {str(exc)}",
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logger.error(f"Unexpected error analyzing order from link: {exc}")
        raise HTTPException(
            status_code=500,
            detail=f"Order analysis failed for link: {str(exc)}",
        )


@app.get("/analysis-history", tags=["Order Analysis"])
def get_analysis_history(
    limit: int = Query(50, description="Maximum number of analyses to return"),
    current_user=Depends(get_current_user),
):
    """Get history of order document analyses from case-details."""
    try:
        db = firestore.client()
        case_docs = (
            db.collection("case-details")
            .order_by("updated_at", direction=firestore.Query.DESCENDING)
            .limit(limit * 3)
            .stream()
        )

        analyses = []
        for doc in case_docs:
            case_data = doc.to_dict() or {}
            latest_status = case_data.get("latest_order_status")
            if latest_status != "analysed":
                continue

            orders = case_data.get("orders") or []
            latest_order = orders[-1] if orders and isinstance(orders[-1], dict) else {}

            board_date = case_data.get("latest_board_date")
            board_row = None
            if board_date:
                board_row = (
                    db.collection("daily-boards")
                    .where("case_ref", "==", case_data.get("case_ref"))
                    .where("board_date", "==", board_date)
                    .limit(1)
                    .get()
                )
            board_data = board_row[0].to_dict() if board_row else {}

            analysis_data = {
                "id": doc.id,
                "case_id": doc.id,
                "case_ref": case_data.get("case_ref"),
                "case_type": case_data.get("case_type"),
                "case_no": case_data.get("case_no"),
                "case_year": case_data.get("case_year"),
                "board_date": case_data.get("latest_board_date"),
                "petitioner_lawyer": board_data.get("petitioner_lawyer"),
                "respondent_lawyer": board_data.get("respondent_lawyer"),
                "order_category": case_data.get("latest_order_category")
                or latest_order.get("order_category"),
                "category_confidence": latest_order.get("order_category_confidence"),
                "order_date": case_data.get("latest_order_date")
                or latest_order.get("order_date"),
                "date_validation": latest_order.get("order_date_validation"),
                "order_link": case_data.get("latest_order_link")
                or latest_order.get("order_link"),
                "analysis_timestamp": latest_order.get("order_analysis_timestamp"),
            }

            analyses.append(analysis_data)
            if len(analyses) >= limit:
                break

        return JSONResponse(
            content={
                "analyses": analyses,
                "count": len(analyses),
                "total_fetched": len(analyses),
            }
        )

    except Exception as e:
        logger.error(f"Error fetching analysis history: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to fetch analysis history: {str(e)}"},
        )


@app.get("/analysis/{analysis_id}", tags=["Order Analysis"])
async def get_analysis_details(
    analysis_id: str, current_user=Depends(get_current_user)
):
    """Get detailed analysis results for a specific case from case-details."""
    try:
        db = firestore.client()

        doc_ref = db.collection("daily-boards").document(analysis_id)
        doc = doc_ref.get()

        if not doc.exists:
            return JSONResponse(status_code=404, content={"error": "Case not found"})

        board_data = doc.to_dict() or {}
        case_ref = f"{board_data.get('case_type', '')}/{board_data.get('case_no', '')}/{board_data.get('case_year', '')}"
        case_data = get_auto_order_manager().case_store.get_case_details(case_ref) or {}
        latest_status = case_data.get("latest_order_status", "not_linked")
        orders = case_data.get("orders") or []
        latest_order = orders[-1] if orders and isinstance(orders[-1], dict) else {}

        if latest_status != "analysed":
            return JSONResponse(
                status_code=404,
                content={"error": "Order analysis not completed for this case"},
            )

        analysis_data = {
            "id": doc.id,
            "case_id": doc.id,
            "case_ref": case_ref,
            "case_type": board_data.get("case_type"),
            "case_no": board_data.get("case_no"),
            "case_year": board_data.get("case_year"),
            "board_date": board_data.get("board_date"),
            "petitioner_lawyer": board_data.get("petitioner_lawyer"),
            "respondent_lawyer": board_data.get("respondent_lawyer"),
            "serial_number": board_data.get("serial_number"),
            "additional_cases": board_data.get("additional_cases"),
            "order_category": case_data.get("latest_order_category")
            or latest_order.get("order_category"),
            "category_confidence": latest_order.get("order_category_confidence"),
            "order_date": case_data.get("latest_order_date")
            or latest_order.get("order_date"),
            "date_validation": latest_order.get("order_date_validation"),
            "order_link": case_data.get("latest_order_link")
            or latest_order.get("order_link"),
            "analysis_timestamp": latest_order.get("order_analysis_timestamp"),
            "last_updated": latest_order.get("updated_at"),
        }

        return JSONResponse(content=analysis_data)

    except Exception as e:
        logger.error(f"Error fetching analysis details: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to fetch analysis details: {str(e)}"},
        )


@app.get("/analysis-stats", tags=["Order Analysis"])
def get_analysis_statistics(current_user=Depends(get_current_user)):
    """Get statistics about order document analyses from case-details.

    Streams every analysed case-details doc, unbounded.

    Deliberately a sync `def`, not `async def`: it does blocking Firestore
    I/O, and FastAPI runs sync handlers in a threadpool, so a slow query
    here cannot stall the event loop (and with it every other request
    plus the fetch/analysis poll loops).
    """
    try:
        db = firestore.client()

        analyses_ref = db.collection("case-details").where(
            "latest_order_status", "==", "analysed"
        )
        docs = analyses_ref.stream()

        stats = {
            "total_analyses": 0,
            "category_distribution": {
                "ADJOURNED": 0,
                "HEARD_AND_ADJOURNED": 0,
                "DISPOSED_OFF": 0,
            },
            "avg_confidence": 0.0,
            "recent_analyses": 0,  # Last 30 days
        }

        confidences = []
        recent_cutoff = datetime.now().timestamp() - (30 * 24 * 60 * 60)  # 30 days ago

        for doc in docs:
            data = doc.to_dict() or {}
            latest_order = {}
            orders = data.get("orders") or []
            if orders and isinstance(orders[-1], dict):
                latest_order = orders[-1]
            stats["total_analyses"] += 1

            # Category distribution
            category = data.get("latest_order_category") or latest_order.get(
                "order_category", "UNKNOWN"
            )
            if category in stats["category_distribution"]:
                stats["category_distribution"][category] += 1

            # Confidence scores
            confidence = latest_order.get("order_category_confidence", 0)
            if confidence > 0:
                confidences.append(confidence)

            # Recent analyses - use order_analysis_timestamp
            timestamp_str = latest_order.get("order_analysis_timestamp", "")
            if timestamp_str:
                try:
                    timestamp = datetime.fromisoformat(
                        timestamp_str.replace("Z", "+00:00")
                    ).timestamp()
                    if timestamp > recent_cutoff:
                        stats["recent_analyses"] += 1
                except (ValueError, TypeError):
                    pass

        # Calculate average confidence
        if confidences:
            stats["avg_confidence"] = round(sum(confidences) / len(confidences), 3)

        return JSONResponse(content=stats)

    except Exception as e:
        logger.error(f"Error fetching analysis statistics: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to fetch statistics: {str(e)}"}
        )


# Auto Order Management Endpoints
@app.post("/auto-orders/process-cases", tags=["Auto Order Management"])
async def auto_process_orders(request: Request, current_user=Depends(get_current_user)):
    """Automatically process cases for order download and analysis"""
    try:
        body = await request.json()
        filters = body.get("filters", {})
        limit = body.get("limit", 50)

        result = get_auto_order_manager().get_orders_for_cases(filters, limit)

        if result.get("success"):
            return JSONResponse(content=result)
        else:
            return JSONResponse(
                status_code=500, content={"error": result.get("error", "Unknown error")}
            )

    except Exception as e:
        logger.error(f"Error in auto-process-orders: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to process orders: {str(e)}"}
        )


@app.post("/jobs/fetch-orders", tags=["Auto Order Management"])
async def queue_fetch_orders_jobs(
    request: Request, current_user=Depends(require_admin_active)
):
    """Queue fetch jobs for eligible cases based on filters."""
    try:
        body = await request.json()
        filters = body.get("filters", {})
        board_dates = body.get("board_dates") or []
        case_refs = body.get("case_refs") or []
        limit = int(body.get("limit", 100))
        scope = body.get("scope", "missing_only")

        if limit < 1 or limit > 1000:
            return JSONResponse(
                status_code=400,
                content={"error": "limit must be between 1 and 1000"},
            )
        if scope not in ("missing_only", "all"):
            return JSONResponse(
                status_code=400,
                content={"error": "scope must be 'missing_only' or 'all'"},
            )

        manager = get_auto_order_manager()
        selected_case_refs = {
            str(value or "").strip().upper()
            for value in case_refs
            if str(value or "").strip()
        }
        selected_board_dates = {
            str(value or "").strip()
            for value in board_dates
            if str(value or "").strip()
        }

        # Push the selected board dates into the Firestore query.  Filtering
        # after the limit used to return zero candidates whenever the selected
        # dates fell outside the first `limit` documents scanned.
        candidate_cases = manager._get_filtered_matters(
            filters,
            limit,
            board_dates=sorted(selected_board_dates) or None,
            scope=scope,
        )

        if selected_case_refs:
            candidate_cases = [
                case_data
                for case_data in candidate_cases
                if str(case_data.get("case_ref") or "").strip().upper()
                in selected_case_refs
            ]

        queued = 0
        skipped_not_due = 0
        queued_case_refs = []
        today = datetime.now().date()

        for case_data in candidate_cases:
            board_date = manager._parse_board_date(case_data.get("board_date"))
            case_ref = case_data.get("case_ref")
            case_id = case_data.get("id")

            if board_date and board_date > today:
                skipped_not_due += 1
                manager.case_store.transition_lifecycle(
                    case_ref,
                    "fetch_not_due",
                    reason=(
                        f"Order fetch is not due yet for board date {board_date.isoformat()}"
                    ),
                    metadata={"source": "jobs.fetch-orders", "case_id": case_id},
                    event_type="fetch_job_not_due",
                )
                continue

            manager.case_store.transition_lifecycle(
                case_ref,
                "fetch_queued",
                metadata={"source": "jobs.fetch-orders", "case_id": case_id},
                event_type="fetch_job_queued",
            )
            queued += 1
            queued_case_refs.append(case_ref)

        if queued:
            _wake_fetch_poll.set()

        return JSONResponse(
            content={
                "success": True,
                "scope": scope,
                "queued": queued,
                "skipped_not_due": skipped_not_due,
                "queued_case_refs": queued_case_refs,
                "selected_board_dates": sorted(selected_board_dates),
                "selected_case_refs": sorted(selected_case_refs),
            }
        )
    except Exception as e:
        logger.error(f"Error queueing fetch jobs: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to queue fetch jobs: {str(e)}"},
        )


@app.post("/jobs/analyze-orders", tags=["Auto Order Management"])
async def queue_analysis_jobs(
    request: Request, current_user=Depends(require_admin_active)
):
    """Queue analysis jobs for cases that already have order links."""
    try:
        db = firestore.client()
        body = await request.json()

        limit = int(body.get("limit", 100))
        days_back = body.get("days_back")
        case_refs = body.get("case_refs") or []
        board_dates = body.get("board_dates") or []
        scope = body.get("scope", "missing_only")

        if limit < 1 or limit > 1000:
            return JSONResponse(
                status_code=400,
                content={"error": "limit must be between 1 and 1000"},
            )
        if scope not in ("missing_only", "all"):
            return JSONResponse(
                status_code=400,
                content={"error": "scope must be 'missing_only' or 'all'"},
            )

        manager = get_auto_order_manager()
        queued = 0
        skipped = 0
        queued_case_refs = []
        selected_board_dates = {
            str(value or "").strip()
            for value in board_dates
            if str(value or "").strip()
        }

        candidate_rows = []
        if case_refs:
            for case_ref in case_refs:
                normalized_ref = str(case_ref or "").strip().upper()
                if not normalized_ref:
                    continue
                parts = normalized_ref.split("/")
                if len(parts) != 3:
                    continue
                query = (
                    db.collection("daily-boards")
                    .where("case_type", "==", parts[0])
                    .where("case_no", "==", parts[1])
                    .where("case_year", "==", parts[2])
                    .limit(1)
                )
                rows = list(query.stream())
                if rows:
                    candidate_rows.append(rows[0])
        elif selected_board_dates:
            # One equality query per selected date.  Filtering after a blanket
            # limit used to yield zero rows whenever the selected dates were not
            # among the first `limit * 4` documents scanned.
            seen_row_ids = set()
            for date_str in sorted(selected_board_dates):
                board_dt = manager._to_board_date_query_value(date_str)
                if not board_dt:
                    continue
                query = (
                    db.collection("daily-boards")
                    .where("board_date", "==", board_dt)
                    .limit(limit * 4)
                )
                for row in query.stream():
                    if row.id in seen_row_ids:
                        continue
                    seen_row_ids.add(row.id)
                    candidate_rows.append(row)
        else:
            query = db.collection("daily-boards")
            if days_back:
                start_date = datetime.now() - timedelta(days=int(days_back))
                start_datetime = datetime(
                    start_date.year, start_date.month, start_date.day, 0, 0, 0
                )
                query = query.where("board_date", ">=", start_datetime)
            candidate_rows = list(query.limit(limit * 4).stream())

        for row in candidate_rows:
            row_data = row.to_dict() or {}
            board_date_obj = manager._parse_board_date(row_data.get("board_date"))
            board_date_iso = board_date_obj.isoformat() if board_date_obj else None
            if selected_board_dates and board_date_iso not in selected_board_dates:
                continue

            case_ref = manager.case_store.build_case_ref(
                row_data.get("case_type"),
                row_data.get("case_no"),
                row_data.get("case_year"),
            )
            order_context = manager._get_case_order_context(case_ref)
            order_link = order_context.get("order_link")

            if not order_link:
                skipped += 1
                continue
            if scope == "missing_only" and order_context.get("order_status") in (
                "analysed",
                "manual_review_required",
            ):
                # Already analysed (or awaiting human review, which an
                # automatic re-analysis must not silently override).
                skipped += 1
                continue

            manager.case_store.transition_lifecycle(
                case_ref,
                "analysis_queued",
                metadata={"source": "jobs.analyze-orders", "case_id": row.id},
                event_type="analysis_job_queued",
            )
            queued += 1
            queued_case_refs.append(case_ref)

            if queued >= limit:
                break

        if queued:
            _wake_analysis_poll.set()

        return JSONResponse(
            content={
                "success": True,
                "scope": scope,
                "queued": queued,
                "skipped": skipped,
                "queued_case_refs": queued_case_refs,
                "selected_board_dates": sorted(selected_board_dates),
            }
        )
    except Exception as e:
        logger.error(f"Error queueing analysis jobs: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to queue analysis jobs: {str(e)}"},
        )


def _query_stuck_candidates(db, limit: int) -> List[Dict]:
    """Cases actually in one of STUCK_LIFECYCLE_STATUSES -- the same
    population /queue/status's needs_attention_count counts. One equality
    query per status (auto-indexed, no composite index needed), stopping
    once ``limit`` candidates have been collected across all four."""
    candidates: List[Dict] = []
    for status in STUCK_LIFECYCLE_STATUSES:
        if len(candidates) >= limit:
            break
        remaining = limit - len(candidates)
        docs = (
            db.collection("case-details")
            .where("lifecycle_status", "==", status)
            .limit(remaining)
            .stream()
        )
        for doc in docs:
            data = doc.to_dict() or {}
            data["id"] = doc.id
            data["lifecycle_status"] = status
            candidates.append(data)
    return candidates


@app.post("/jobs/retry-failed", tags=["Auto Order Management"])
async def retry_failed_cases(
    request: Request, current_user=Depends(require_admin_active)
):
    """Re-queue cases the pipeline could not finish on its own -- the same
    population the Dashboard's "N cases could not be completed
    automatically" banner counts (STUCK_LIFECYCLE_STATUSES). Nothing else
    ever moves a case out of these states: the poll loops only reclaim
    STALE *_in_progress cases, never a *_failed_* one.

    fetch_failed_* -> fetch queue (re-download from scratch).
    analysis_failed_* -> analysis queue if an order link is already on
    file; otherwise falls back to the fetch queue.

    Was previously drawn from AutoOrderManager._get_filtered_matters, an
    unfiltered daily-boards scan (scope="actionable" filters afterwards, but
    the underlying query has no where clause at all). With no orderBy,
    Firestore defaults to document-ID order, and daily-boards doc ids are
    date-prefixed -- so this deterministically returned the OLDEST slice of
    the entire collection's history, every single call, regardless of
    where the actually-stuck cases currently were. Clicking "Retry them"
    could look like it did nothing because it almost always was: scanning
    the same long-since-resolved old boards instead of the current
    failures, using a legacy order_status field the rest of the pipeline
    no longer treats as authoritative.

    Accepts optional ``board_dates`` (list of YYYY-MM-DD strings) and ``limit``
    (default 200, 1-1000) in the request body.
    """
    try:
        body = await request.json()
        board_dates = body.get("board_dates") or []
        limit = int(body.get("limit", 200))

        if limit < 1 or limit > 1000:
            return JSONResponse(
                status_code=400,
                content={"error": "limit must be between 1 and 1000"},
            )

        db = firestore.client()
        selected_board_dates = {
            str(v or "").strip() for v in board_dates if str(v or "").strip()
        }

        candidate_cases = _query_stuck_candidates(db, limit)
        case_store = get_auto_order_manager().case_store

        fetch_queued = 0
        analysis_queued = 0
        skipped = 0
        fetch_queued_refs: list = []
        analysis_queued_refs: list = []

        for case_data in candidate_cases:
            case_ref = case_data.get("case_ref")
            if not case_ref:
                continue

            board_date_iso = case_store._to_iso_date(case_data.get("latest_board_date"))
            if selected_board_dates and board_date_iso not in selected_board_dates:
                skipped += 1
                continue

            case_id = case_data.get("id")
            lifecycle_status = case_data.get("lifecycle_status", "")

            if lifecycle_status.startswith("fetch_failed"):
                case_store.transition_lifecycle(
                    case_ref,
                    "fetch_queued",
                    metadata={"source": "jobs.retry-failed", "case_id": case_id},
                    event_type="retry_fetch_queued",
                )
                fetch_queued += 1
                fetch_queued_refs.append(case_ref)
            else:
                order_link = case_data.get("latest_order_link")
                if not order_link:
                    case_store.transition_lifecycle(
                        case_ref,
                        "fetch_queued",
                        metadata={
                            "source": "jobs.retry-failed",
                            "case_id": case_id,
                            "reason": "no_order_link_for_analysis_retry",
                        },
                        event_type="retry_fetch_queued",
                    )
                    fetch_queued += 1
                    fetch_queued_refs.append(case_ref)
                else:
                    case_store.transition_lifecycle(
                        case_ref,
                        "analysis_queued",
                        metadata={
                            "source": "jobs.retry-failed",
                            "case_id": case_id,
                        },
                        event_type="retry_analysis_queued",
                    )
                    analysis_queued += 1
                    analysis_queued_refs.append(case_ref)

        if fetch_queued:
            _wake_fetch_poll.set()
        if analysis_queued:
            _wake_analysis_poll.set()

        return JSONResponse(
            content={
                "success": True,
                "fetch_queued": fetch_queued,
                "analysis_queued": analysis_queued,
                "skipped": skipped,
                "fetch_queued_refs": fetch_queued_refs,
                "analysis_queued_refs": analysis_queued_refs,
                "selected_board_dates": sorted(selected_board_dates),
            }
        )

    except Exception as e:
        logger.error(f"Error in retry-failed: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to retry failed cases: {str(e)}"},
        )


async def mark_case_for_manual_review(
    case_ref: str,
    request: Request,
    current_user=Depends(require_active_user),
):
    """Move a case to manual review workflow with audit metadata."""
    try:
        body = await request.json()
        normalized_case_ref = str(case_ref or "").strip().upper()
        if not normalized_case_ref:
            return JSONResponse(
                status_code=400, content={"error": "case_ref is required"}
            )

        case_store = get_auto_order_manager().case_store
        case_details = case_store.get_case_details(normalized_case_ref)
        if not case_details:
            return JSONResponse(status_code=404, content={"error": "Case not found"})

        reason = body.get("reason") or "Marked for manual review"
        notes = body.get("notes")
        actor_uid = current_user.get("uid")

        transition = case_store.transition_lifecycle(
            normalized_case_ref,
            "manual_review_required",
            reason=reason,
            metadata={
                "source": "manual-review",
                "actor_uid": actor_uid,
                "notes": notes,
            },
            event_type="manual_review_marked",
            force=True,
        )

        return JSONResponse(
            content={
                "success": True,
                "case_ref": normalized_case_ref,
                "transition": transition,
                "lifecycle": case_store.build_lifecycle_summary(normalized_case_ref),
            }
        )
    except Exception as e:
        logger.error(f"Error marking manual review for {case_ref}: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to mark manual review: {str(e)}"},
        )


@app.post("/cases/{case_ref:path}/manual-override", tags=["Auto Order Management"])
async def manual_override_case_outcome(
    case_ref: str,
    request: Request,
    current_user=Depends(require_admin_active),
):
    """Apply a reviewed final outcome and move case lifecycle to analysed."""
    try:
        body = await request.json()
        normalized_case_ref = str(case_ref or "").strip().upper()
        if not normalized_case_ref:
            return JSONResponse(
                status_code=400, content={"error": "case_ref is required"}
            )

        case_store = get_auto_order_manager().case_store
        case_details = case_store.get_case_details(normalized_case_ref)
        if not case_details:
            return JSONResponse(status_code=404, content={"error": "Case not found"})

        order_category = body.get("order_category")
        order_date = body.get("order_date")
        notes = body.get("notes") or "Manual override applied"
        actor_uid = current_user.get("uid")

        latest_link = case_details.get("latest_order_link")
        case_store.append_case_order(
            normalized_case_ref,
            {
                "order_status": "analysed",
                "order_category": order_category,
                "order_date": order_date,
                "order_link": latest_link,
                "order_analysis_timestamp": datetime.now().isoformat(),
                "order_manual_override": True,
                "order_manual_override_notes": notes,
                "order_manual_override_by": actor_uid,
            },
        )
        transition = case_store.transition_lifecycle(
            normalized_case_ref,
            "analysed",
            reason="Manual override completed",
            metadata={
                "source": "manual-override",
                "actor_uid": actor_uid,
                "notes": notes,
                "order_category": order_category,
                "order_date": order_date,
            },
            event_type="manual_override",
            force=True,
        )

        return JSONResponse(
            content={
                "success": True,
                "case_ref": normalized_case_ref,
                "transition": transition,
                "lifecycle": case_store.build_lifecycle_summary(normalized_case_ref),
            }
        )
    except Exception as e:
        logger.error(f"Error applying manual override for {case_ref}: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to apply manual override: {str(e)}"},
        )


@app.post("/cases/{case_ref:path}/reset", tags=["Auto Order Management"])
async def reset_case_orders(
    case_ref: str,
    current_user=Depends(require_admin),
):
    """Hard-reset a case: clear all order history and requeue every board entry.

    Wipes the orders array and latest_order_* fields from case-details, then
    resets every daily-boards entry for this case to fetch_queued so the
    background pipeline re-fetches every order PDF and uploads them to GCS.

    Use this when a case has stale BHC links that retries have not replaced,
    or whenever a clean slate is needed.
    """
    try:
        ensure_firebase()
        normalized = str(case_ref or "").strip().upper()
        if not normalized:
            return JSONResponse(
                status_code=400, content={"error": "case_ref is required"}
            )

        case_store = get_auto_order_manager().case_store
        case_store.reset_case_for_reprocessing(normalized)

        return JSONResponse(
            content={
                "success": True,
                "case_ref": normalized,
                "message": f"Case {normalized} reset and queued for re-fetch.",
            }
        )
    except Exception as exc:
        logger.error("reset_case_orders failed for %s: %s", case_ref, exc)
        return JSONResponse(
            status_code=500,
            content={"error": f"Reset failed: {str(exc)}"},
        )


@app.post("/auto-orders/process-case", tags=["Auto Order Management"])
async def process_single_case(request: Request, current_user=Depends(get_current_user)):
    """Process a single case for order download and analysis"""
    try:
        db = firestore.client()

        body = await request.json()
        case_id = body.get("case_id")
        case_ref = body.get("case_ref")
        board_date = body.get("board_date")

        if not case_id or not case_ref:
            return JSONResponse(
                status_code=400, content={"error": "case_id and case_ref are required"}
            )

        # Fetch existing case data from database
        case_doc = db.collection("daily-boards").document(case_id).get()

        if not case_doc.exists:
            return JSONResponse(status_code=404, content={"error": "Case not found"})

        case_data = case_doc.to_dict()

        # Enqueue asynchronously — do not block the request for up to 5 minutes.
        # The caller can poll GET /auto-orders/job-status/{case_id} for progress.
        auto_mgr = get_auto_order_manager()
        # force=True so retries always reset lifecycle regardless of current state
        # (e.g. fetch_in_progress from a previous stuck/timed-out attempt).
        auto_mgr.case_store.transition_lifecycle(
            case_ref,
            "fetch_queued",
            force=True,
            metadata={"source": "process_case_endpoint", "case_id": case_id},
            event_type="fetch_job_queued",
            extra_fields={
                "latest_board_date": auto_mgr.case_store._to_iso_date(
                    board_date or case_data.get("board_date")
                )
            },
        )
        _wake_fetch_poll.set()

        return JSONResponse(
            content={
                "success": True,
                "job_id": case_id,
                "status": "queued",
                "case_ref": case_ref,
                "message": "Case queued for processing. Poll /auto-orders/job-status/{case_id} for progress.",
            }
        )

    except Exception as e:
        logger.error(f"Error processing single case: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to process case: {str(e)}"}
        )


@app.get("/auto-orders/job-status/{doc_id}", tags=["Auto Order Management"])
async def get_job_status(doc_id: str, current_user=Depends(get_current_user)):
    """Poll the processing status of a single case queued via /auto-orders/process-case."""
    try:
        db = firestore.client()
        # doc_id is the daily-boards document ID; read it to get case_ref
        board_doc = db.collection("daily-boards").document(doc_id).get()
        if not board_doc.exists:
            return JSONResponse(status_code=404, content={"error": "Case not found"})
        board_data = board_doc.to_dict() or {}

        # Lifecycle status lives in case-details (keyed by case_ref with / → -)
        case_ref = board_data.get(
            "case_ref"
        ) or get_auto_order_manager().case_store.build_case_ref(
            board_data.get("case_type"),
            board_data.get("case_no"),
            board_data.get("case_year"),
        )
        case_details_id = case_ref.replace("/", "-")
        case_doc = db.collection("case-details").document(case_details_id).get()
        case_data = case_doc.to_dict() if case_doc.exists else {}

        lifecycle_status = case_data.get("lifecycle_status") or "board_ingested"
        updated_at = case_data.get("updated_at")

        # Surface the last lifecycle event so the UI can show the actual error
        events = case_data.get("lifecycle_events") or []
        last_event = events[-1] if events else {}

        # Check if the most recent order was stored with an expiring court URL
        # because the GCS upload failed (persisted in order payload).
        orders = case_data.get("orders") or []
        orders_with_link = [
            o for o in orders if isinstance(o, dict) and o.get("order_link")
        ]
        latest_order = orders_with_link[-1] if orders_with_link else {}
        gcs_upload_failed = bool(latest_order.get("gcs_upload_failed"))

        return JSONResponse(
            content={
                "doc_id": doc_id,
                "status": lifecycle_status,
                "error_reason": case_data.get("lifecycle_status_reason"),
                "last_event": (
                    {
                        "event_type": last_event.get("event_type"),
                        "reason": last_event.get("reason"),
                        "timestamp": last_event.get("timestamp"),
                    }
                    if last_event
                    else None
                ),
                "order_category": case_data.get("latest_order_category")
                or board_data.get("order_category"),
                "order_link": case_data.get("latest_order_link")
                or board_data.get("order_link"),
                "gcs_upload_failed": gcs_upload_failed,
                "updated_at": (
                    updated_at.isoformat()
                    if hasattr(updated_at, "isoformat")
                    else str(updated_at)
                    if updated_at
                    else None
                ),
            }
        )
    except Exception as e:
        logger.error(f"Error getting job status for {doc_id}: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to get job status: {str(e)}"}
        )


@app.post("/auto-orders/analyze-case/{case_id}", tags=["Auto Order Management"])
async def analyze_single_case(case_id: str, current_user=Depends(get_current_user)):
    """Analyze an already downloaded order for a case"""
    try:
        db = firestore.client()
        # Get case data from database
        case_doc = db.collection("daily-boards").document(case_id).get()
        if not case_doc.exists:
            return JSONResponse(status_code=404, content={"error": "Case not found"})

        case_data = case_doc.to_dict()
        case_ref = f"{case_data.get('case_type')}/{case_data.get('case_no')}/{case_data.get('case_year')}"
        case_details = (
            get_auto_order_manager().case_store.get_case_details(case_ref) or {}
        )
        latest_status = case_details.get("latest_order_status", "not_linked")
        latest_order_link = case_details.get("latest_order_link")

        if not latest_order_link:
            return JSONResponse(
                status_code=400,
                content={
                    "error": "No order available to analyze. Please download the order first."
                },
            )

        # If already analyzed, return existing analysis
        if latest_status == "analysed":
            orders = case_details.get("orders") or []
            latest_order = orders[-1] if orders and isinstance(orders[-1], dict) else {}
            return JSONResponse(
                content={
                    "success": True,
                    "message": "Order already analyzed",
                    "data": {
                        "order_category": case_details.get("latest_order_category")
                        or latest_order.get("order_category"),
                        "order_date": case_details.get("latest_order_date")
                        or latest_order.get("order_date"),
                        "order_petitioner": case_details.get("petitioner"),
                        "order_respondent": case_details.get("respondent"),
                        "government_pleader": case_details.get("government_pleader"),
                    },
                }
            )

        # Analyse the already-downloaded order.  This delegates to the same
        # _analyze_existing_order the analysis queue uses, so the per-row
        # "Analyse" button and POST /jobs/analyze-orders share one code path
        # (and one definition of "already analysed").
        manager = get_auto_order_manager()
        try:
            analysis = manager._analyze_existing_order(
                {
                    "id": case_id,
                    "case_ref": case_ref,
                    "order_link": latest_order_link,
                    "board_date": case_data.get("board_date"),
                    "order_status": latest_status,
                },
                {
                    "case_id": case_id,
                    "case_ref": case_ref,
                    "download_success": True,
                    "analysis_success": False,
                    "order_link": latest_order_link,
                    "analysis_data": None,
                    "error": None,
                    "retry_attempts": [],
                    "has_existing_order": True,
                },
            )

            if analysis.get("analysis_success"):
                data = dict(analysis.get("analysis_data") or {})
                data.pop("order_cases", None)
                return JSONResponse(content={"success": True, "data": data})

            # Stored link is stale or unreadable — re-fetch from the court and
            # analyse, which is exactly what _process_single_case already does.
            logger.warning(
                "analyze-case: stored link unusable for %s (%s). Re-fetching from court.",
                case_ref,
                analysis.get("error"),
            )
            fresh = manager._process_single_case(
                {
                    "id": case_id,
                    "case_ref": case_ref,
                    "case_type": case_data.get("case_type"),
                    "case_no": case_data.get("case_no"),
                    "case_year": case_data.get("case_year"),
                    "board_date": case_data.get("board_date"),
                }
            )
            if fresh.get("analysis_success"):
                fresh_data = dict(fresh.get("analysis_data") or {})
                fresh_data.pop("order_cases", None)
                return JSONResponse(
                    content={
                        "success": True,
                        "data": fresh_data,
                        "message": "Order re-downloaded and analysed successfully",
                    }
                )
            return JSONResponse(
                status_code=500,
                content={
                    "error": fresh.get("error")
                    or analysis.get("error")
                    or "Failed to analyse order"
                },
            )

        except Exception as e:
            logger.error(f"Unexpected error in download/analyze: {e}", exc_info=True)
            return JSONResponse(
                status_code=500, content={"error": f"Failed to analyze order: {str(e)}"}
            )

    except Exception as e:
        logger.error(f"Error in analyze-case: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to analyze case: {str(e)}"}
        )


@app.post("/auto-orders/bulk-process", tags=["Auto Order Management"])
async def bulk_process_orders(request: Request, current_user=Depends(get_current_user)):
    """Bulk process specific cases by IDs with configurable max sequences"""
    try:
        body = await request.json()
        case_ids = body.get("case_ids", [])

        if not case_ids:
            return JSONResponse(
                status_code=400, content={"error": "No case IDs provided"}
            )

        result = get_auto_order_manager().bulk_process_orders(case_ids)

        if result.get("success"):
            return JSONResponse(content=result)
        else:
            return JSONResponse(
                status_code=500, content={"error": result.get("error", "Unknown error")}
            )

    except Exception as e:
        logger.error(f"Error in bulk-process-orders: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to bulk process orders: {str(e)}"},
        )


@app.post("/auto-orders/upload-manual-order/{case_id}", tags=["Auto Order Management"])
async def upload_manual_order(
    case_id: str, file: UploadFile = File(...), current_user=Depends(get_current_user)
):
    """
    Upload a manual order PDF for a case and automatically analyze it
    This allows users to upload order PDFs when automatic download isn't available
    """
    try:
        db = firestore.client()
        # Verify it's a PDF
        if file.content_type != "application/pdf":
            return JSONResponse(
                status_code=400, content={"error": "File must be a PDF"}
            )

        # Get case data
        case_doc = db.collection("daily-boards").document(case_id).get()
        if not case_doc.exists:
            return JSONResponse(status_code=404, content={"error": "Case not found"})

        case_data = case_doc.to_dict()
        case_ref = f"{case_data.get('case_type')}/{case_data.get('case_no')}/{case_data.get('case_year')}"

        # Read PDF content
        pdf_content = await file.read()

        manager = get_auto_order_manager()
        # A manually uploaded order carries no court-API date, so the board date
        # is the best available order date.
        order_date_str = manager._normalise_order_date(
            case_data.get("board_date")
        ) or datetime.now().strftime("%Y-%m-%d")

        # Store the PDF permanently so the resulting link is actually viewable.
        # Falls back to a marker string when GCS is not configured.
        order_link = (
            manager._upload_order_to_gcs(pdf_content, case_ref, order_date_str)
            or f"manual_upload_{case_id}_{file.filename}"
        )

        # Analyse via the same path the automatic pipeline uses, which also
        # records the lifecycle transitions and writes the case-details order.
        analysis_result = manager._analyze_order_with_api_metadata(
            case_id=case_id,
            case_ref=case_ref,
            pdf_content=pdf_content,
            api_order_date=order_date_str,
            api_petitioner="",
            api_respondent="",
            order_link=order_link,
            board_date=case_data.get("board_date"),
        )

        if analysis_result.get("success"):
            # Propagate the link/category back to daily-boards, as the
            # automatic path does.
            manager._update_board_entries_for_case_date(
                case_ref,
                order_date_str,
                order_link,
                (analysis_result.get("data") or {}).get("order_category"),
            )

            return JSONResponse(
                content={
                    "success": True,
                    "message": "Order uploaded and analyzed successfully",
                    "case_id": case_id,
                    "case_ref": case_ref,
                    "filename": file.filename,
                    "analysis": {
                        "order_category": analysis_result["data"].get("order_category"),
                        "order_date": analysis_result["data"].get("order_date"),
                        "order_petitioner": analysis_result["data"].get(
                            "order_petitioner"
                        ),
                        "order_respondent": analysis_result["data"].get(
                            "order_respondent"
                        ),
                        "government_pleader": analysis_result["data"].get(
                            "government_pleader"
                        ),
                    },
                }
            )
        else:
            return JSONResponse(
                status_code=500,
                content={
                    "success": False,
                    "message": "Order uploaded but analysis failed",
                    "error": analysis_result.get("error"),
                },
            )

    except Exception as e:
        logger.error(f"Error uploading manual order: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to upload order: {str(e)}"}
        )


@app.post("/auto-orders/scheduled-retry", tags=["Auto Order Management"])
async def scheduled_retry_orders(
    days_back: int = Query(7, description="Number of days to look back"),
    limit: int = Query(100, description="Maximum cases to process"),
    current_user=Depends(get_current_user),
):
    """
    Scheduled endpoint for automatic retry of order downloads
    Can be called by Cloud Scheduler or cron job to automatically process cases without orders

    Use Case: After board upload, orders may not be available yet. This endpoint
    retries downloading orders for recent cases that don't have orders yet.
    """
    try:
        db = firestore.client()
        # Calculate date range
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days_back)

        query = (
            db.collection("daily-boards")
            .where("board_date", ">=", start_date.strftime("%Y-%m-%d"))
            .where("board_date", "<=", end_date.strftime("%Y-%m-%d"))
            .limit(limit * 3)
        )

        cases = query.get()

        case_list = []
        for case_doc in cases:
            case_data = case_doc.to_dict()
            case_ref = f"{case_data.get('case_type')}/{case_data.get('case_no')}/{case_data.get('case_year')}"
            status = (
                get_auto_order_manager()
                ._get_case_order_context(case_ref)
                .get("order_status", "not_linked")
            )
            if status not in {"not_linked", "order_failed", "order_analysis_failed"}:
                continue
            case_info = {
                "id": case_doc.id,
                "case_ref": case_ref,
                "board_date": case_data.get("board_date"),
            }
            case_list.append(case_info)
            if len(case_list) >= limit:
                break

        if not case_list:
            return JSONResponse(
                content={
                    "success": True,
                    "message": f"No cases without orders found in the last {days_back} days",
                    "cases_found": 0,
                }
            )

        case_store = get_auto_order_manager().case_store
        for case_info in case_list:
            case_store.transition_lifecycle(
                case_info["case_ref"],
                "fetch_queued",
                metadata={"source": "scheduled_retry", "case_id": case_info["id"]},
                event_type="fetch_job_queued",
                extra_fields={
                    "latest_board_date": case_store._to_iso_date(
                        case_info.get("board_date")
                    )
                },
            )
        _wake_fetch_poll.set()

        logger.info(f"Scheduled retry: Marked {len(case_list)} cases fetch_queued")

        return JSONResponse(
            content={
                "success": True,
                "message": f"Added {len(case_list)} cases to background processing queue",
                "cases_queued": len(case_list),
                "date_range": {
                    "start": start_date.strftime("%Y-%m-%d"),
                    "end": end_date.strftime("%Y-%m-%d"),
                },
            }
        )

    except Exception as e:
        logger.error(f"Error in scheduled retry: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Scheduled retry failed: {str(e)}"}
        )


@app.get("/admin/test-gcs", tags=["Admin Order Management"])
async def test_gcs_access(current_user=Depends(require_admin)):
    """Smoke-test GCS bucket read and write access from Cloud Run.

    Writes a tiny sentinel object, reads it back, then deletes it.
    Returns ``status: ok`` when the service account can reach the bucket,
    or ``status: error`` with the exception detail when it cannot.
    """
    mgr = get_auto_order_manager()
    if not mgr._gcs_bucket_name:
        return {"status": "skipped", "reason": "ORDER_PDF_BUCKET env var is not set"}
    try:
        from google.cloud import storage as gcs_storage

        client = gcs_storage.Client()
        bucket = client.bucket(mgr._gcs_bucket_name)
        blob = bucket.blob("__health-check__.txt")
        blob.upload_from_string(b"ok", content_type="text/plain")
        data = blob.download_as_bytes()
        blob.delete()
        return {
            "status": "ok",
            "bucket": mgr._gcs_bucket_name,
            "read_write": data == b"ok",
        }
    except Exception as exc:
        return {
            "status": "error",
            "bucket": mgr._gcs_bucket_name,
            "detail": str(exc),
        }


def _count_case_details_by_order_status(order_status: str) -> int:
    """Cheap Firestore .count() aggregation on the legacy latest_order_status
    field -- single-field equality, already auto-indexed, no composite index
    needed. See get_order_status_overview for why this replaced a per-case
    read loop."""
    try:
        db = firestore.client()
        return (
            db.collection("case-details")
            .where("latest_order_status", "==", order_status)
            .count()
            .get()[0][0]
            .value
        )
    except Exception as e:
        logger.error(f"Error counting latest_order_status={order_status}: {e}")
        return 0


@app.get("/admin/order-status-overview", tags=["Admin Order Management"])
async def get_order_status_overview(current_user=Depends(require_admin)):
    """Get overview of order statuses for admin dashboard.

    Was previously a full `daily-boards` collection scan with one extra
    synchronous case-details read PER row (N+1) to derive each row's
    status -- fine on a handful of test boards, but every one of those
    reads is blocking I/O with no `await`, so on a production-sized
    collection this ties up the single event loop thread for the entire
    scan. That doesn't just make this one request slow: it stalls every
    other request on the same Cloud Run instance, including the poll
    loops -- which is what made the Pipeline tab spin forever with
    nothing ever rendering above it.

    Was then rewritten to cheap .count() aggregations, but still broken
    down by the legacy order_status vocabulary (not_linked/linked/
    analysed/order_failed/order_analysis_failed) -- a different status
    language than the rest of the app (Dashboard, Search Orders) had
    already moved to (lifecycle_status's four plain-English buckets:
    waiting/working/ready/attention). Now driven entirely by cheap
    .count() aggregations per raw lifecycle_status value, bucketed with
    Board.simple_status_for -- the exact same function Search Orders'
    status column and the Dashboard's simple-status filter already use, so
    this table can't drift from what those show.
    """
    try:
        db = firestore.client()
        total_cases = db.collection("case-details").count().get()[0][0].value

        bucket_counts = {k: 0 for k in SIMPLE_STATUS_KEYS}
        counted = 0
        for status in ALL_LIFECYCLE_STATUSES:
            n = _count_lifecycle_status(status)
            counted += n
            bucket_counts[simple_status_for(status)] += n
        # lifecycle_status is absent on cases predating this field (a
        # Firestore equality query can't match a missing field), so
        # whatever's left over defaults to "waiting" -- the same default
        # simple_status_for itself falls back to for an empty status.
        bucket_counts["waiting"] += max(0, total_cases - counted)

        return JSONResponse(
            content={
                "success": True,
                "total_cases": total_cases,
                "status_counts": bucket_counts,
                # Not yet done and not currently moving -- "working" cases
                # are already in flight (see the Processing Queue card),
                # so they don't belong in a "needs action" count.
                "pending_processing": bucket_counts["waiting"]
                + bucket_counts["attention"],
            }
        )

    except Exception as e:
        logger.error(f"Error getting order status overview: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get order status overview: {str(e)}"},
        )


@app.get("/admin/review-queue", tags=["Admin Order Management"])
async def get_admin_review_queue(current_user=Depends(require_admin)):
    """Return cases in the manual_review_required lifecycle state."""
    try:
        db = firestore.client()
        docs = (
            db.collection("case-details")
            .where("lifecycle_status", "==", "manual_review_required")
            .stream()
        )
        cases = []
        for doc in docs:
            data = doc.to_dict() or {}
            data.setdefault("id", doc.id)
            cases.append(data)
        return JSONResponse(content=cases)
    except Exception as e:
        logger.error(f"Error fetching review queue: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to fetch review queue: {str(e)}"},
        )


@app.post("/admin/orders/{doc_id}/ai-suggestion", tags=["Admin Order Management"])
async def admin_ai_review_suggestion(doc_id: str, current_user=Depends(require_admin)):
    """LLM read of a manual-review case's order text, with a rationale --
    offered as a drafted suggestion alongside the regex classifier's own
    result, never applied automatically. The reviewer still picks one of
    the three category buttons themselves; this only saves them from
    opening the PDF blind to figure out why the case was flagged.

    Requires GEMINI_API_KEY. The review queue works fully without it --
    this endpoint just isn't available, and the UI treats a 501 here as
    "no suggestion available" rather than an error."""
    try:
        api_key = os.environ.get("GEMINI_API_KEY")
        if not api_key:
            return JSONResponse(
                status_code=501,
                content={
                    "error": "AI suggestions are not configured (GEMINI_API_KEY not set)."
                },
            )

        case_ref = doc_id.replace("-", "/")
        manager = get_auto_order_manager()
        order_link = manager._get_case_order_context(case_ref).get("order_link")
        if not order_link:
            return JSONResponse(
                status_code=404,
                content={"error": "No order PDF on file for this case."},
            )

        pdf_response = requests.get(order_link, timeout=30)
        pdf_response.raise_for_status()
        analysis = manager.order_analyzer.analyze_order_document(
            f"{doc_id}.pdf", pdf_response.content
        )

        from review_copilot import ReviewCopilotError, call_gemini

        try:
            suggestion = call_gemini(analysis.order_text, api_key)
        except ReviewCopilotError as e:
            return JSONResponse(
                status_code=502, content={"error": f"AI suggestion failed: {e}"}
            )

        return JSONResponse(
            content={
                "doc_id": doc_id,
                "case_ref": case_ref,
                "category": suggestion.get("category"),
                "confidence": suggestion.get("confidence"),
                "rationale": suggestion.get("rationale"),
            }
        )
    except Exception as e:
        logger.error(f"Error getting AI suggestion for {doc_id}: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get AI suggestion: {str(e)}"},
        )


@app.post("/admin/orders/{doc_id}/override", tags=["Admin Order Management"])
async def admin_override_order_category(
    doc_id: str, request: Request, current_user=Depends(require_admin)
):
    """Override the order category for a case in the manual review queue."""
    try:
        body = await request.json()
        order_category = body.get("order_category")
        if not order_category:
            return JSONResponse(
                status_code=400, content={"error": "order_category is required"}
            )

        # doc_id is the case-details document id, i.e. case_ref with "/" -> "-".
        case_ref = doc_id.replace("-", "/")
        manager = get_auto_order_manager()

        result = manager.case_store.apply_category_override(
            case_ref,
            order_category,
            actor_uid=current_user.get("uid"),
            notes=body.get("notes"),
        )
        if not result.get("success"):
            return JSONResponse(
                status_code=404,
                content={"error": result.get("error") or "Case not found"},
            )

        case_ref = result.get("case_ref") or case_ref
        # Propagate to daily-boards so the board view and the "analysed" counts
        # agree with the correction.
        order_date = result.get("order_date") or result.get("board_date")
        if order_date:
            try:
                manager._update_board_entries_for_case_date(
                    case_ref, order_date, None, order_category
                )
            except Exception as propagate_error:
                logger.warning(
                    "override: could not propagate to daily-boards for %s: %s",
                    case_ref,
                    propagate_error,
                )

        return JSONResponse(
            content={
                "success": True,
                "doc_id": doc_id,
                "case_ref": case_ref,
                "order_category": order_category,
                "previous_category": result.get("previous_category"),
            }
        )
    except Exception as e:
        logger.error(f"Error overriding order category for {doc_id}: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to override order category: {str(e)}"},
        )


@app.post("/admin/bulk-order-processing", tags=["Admin Order Management"])
async def admin_bulk_order_processing(
    request: Request, current_user=Depends(require_admin)
):
    """
    Admin endpoint to trigger bulk order processing for cases with specific order status.
    Marks matching cases fetch_queued for the fetch_poll_loop to pick up.

    Request body:
    {
        "order_statuses": ["not_linked", "order_failed"],  // Which statuses to process
        "limit": 100,  // Maximum cases to process (1-1000)
        "days_back": 30  // Only process cases from last N days (optional)
    }

    Note: Cases with "unknown" or missing status are automatically normalized to "not_linked"

    Was previously its own daily-boards scan with a per-candidate blocking
    case-details read (N+1) and no upper bound on `limit` -- on a large
    historical backlog this could take a very long time per request and,
    same as the old /admin/order-status-overview, block the whole event
    loop while it ran. Now delegates to _get_filtered_matters (the same,
    limit-bounded candidate selector /jobs/fetch-orders already uses).
    """
    try:
        body = await request.json()
        order_statuses = body.get(
            "order_statuses", ["not_linked", "linked", "order_failed"]
        )
        limit = int(body.get("limit", 100))
        days_back = body.get("days_back")

        if limit < 1 or limit > 1000:
            return JSONResponse(
                status_code=400,
                content={"error": "limit must be between 1 and 1000"},
            )

        filters = {}
        if days_back:
            filters["date_from"] = (
                datetime.now() - timedelta(days=days_back)
            ).strftime("%Y-%m-%d")

        manager = get_auto_order_manager()
        candidate_cases = manager._get_filtered_matters(
            filters, limit, order_statuses=set(order_statuses)
        )

        if not candidate_cases:
            return JSONResponse(
                content={
                    "success": True,
                    "message": f"No cases found with statuses {order_statuses}",
                    "cases_queued": 0,
                }
            )

        case_store = manager.case_store
        for case_data in candidate_cases:
            case_store.transition_lifecycle(
                case_data["case_ref"],
                "fetch_queued",
                metadata={
                    "source": "admin_bulk_processing",
                    "case_id": case_data.get("id"),
                },
                event_type="fetch_job_queued",
                extra_fields={
                    "latest_board_date": case_store._to_iso_date(
                        case_data.get("board_date")
                    )
                },
            )
        _wake_fetch_poll.set()

        logger.info(
            f"Admin bulk processing: Marked {len(candidate_cases)} cases fetch_queued"
        )

        return JSONResponse(
            content={
                "success": True,
                "message": f"Marked {len(candidate_cases)} cases fetch_queued for background processing",
                "cases_queued": len(candidate_cases),
                "statuses_processed": order_statuses,
            }
        )

    except Exception as e:
        logger.error(f"Error in admin bulk order processing: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Admin bulk processing failed: {str(e)}"},
        )


# Queue Management Endpoints

# The lifecycle_status values that mean the pipeline could not finish a case
# on its own and it needs a human-triggered retry -- nothing else ever moves
# a case out of these (the poll loops only reclaim STALE *_in_progress
# cases, never a *_failed_* one). Named once and reused by both
# /queue/status's needs_attention_count and /jobs/retry-failed's candidate
# selection so the "N cases could not be completed automatically" banner and
# the button that's supposed to clear it can never target different
# populations.
STUCK_LIFECYCLE_STATUSES = (
    "fetch_failed_retryable",
    "fetch_failed_terminal",
    "analysis_failed_retryable",
    "analysis_failed_terminal",
)

# Every lifecycle_status a case can be in (case_data_store.py's
# ALLOWED_LIFECYCLE_TRANSITIONS), queried individually and bucketed with
# Board.simple_status_for to build /admin/order-status-overview's
# waiting/working/ready/attention breakdown -- the same four buckets
# Search Orders' status column and the Dashboard already show, so this
# table can't present a different status language from the rest of the app.
ALL_LIFECYCLE_STATUSES = (
    "board_ingested",
    "fetch_not_due",
    "fetch_queued",
    "fetch_in_progress",
    "fetch_succeeded",
    "analysis_queued",
    "analysis_in_progress",
    "analysed",
    "fetch_failed_retryable",
    "fetch_failed_terminal",
    "analysis_failed_retryable",
    "analysis_failed_terminal",
    "manual_review_required",
)


# Single shared Firestore doc every poll-loop tick writes its timestamp to.
# Replaces the old in-process _last_fetch_poll_tick/_last_analysis_poll_tick
# globals: those lived in ONE Cloud Run instance's memory, but /queue/status
# can be answered by any of up to 10 instances, and there is no reason the
# instance that happens to answer a given request is the same one whose poll
# loop just ticked. In practice it almost never was, so
# "fetch/analysis_processing_active" reported False almost unconditionally
# regardless of whether the pipeline was actually running -- this is why the
# Dashboard always showed the queue as inactive. A shared doc that any
# instance can overwrite and any instance can read fixes that: "last writer
# wins" is exactly the semantics "was ANY instance's loop ticking recently"
# needs.
_POLL_HEARTBEAT_COLLECTION = "system-health"
_POLL_HEARTBEAT_DOC_ID = "poll-loops"


def _write_poll_heartbeat(field: str) -> None:
    try:
        firestore.client().collection(_POLL_HEARTBEAT_COLLECTION).document(
            _POLL_HEARTBEAT_DOC_ID
        ).set({field: datetime.now().isoformat()}, merge=True)
    except Exception as e:  # noqa: BLE001 -- must never crash a poll loop tick
        logger.warning(f"Failed to write poll heartbeat {field}: {e}")


def _poll_loop_is_active(field: str) -> bool:
    try:
        doc = (
            firestore.client()
            .collection(_POLL_HEARTBEAT_COLLECTION)
            .document(_POLL_HEARTBEAT_DOC_ID)
            .get()
        )
        last_tick_iso = (doc.to_dict() or {}).get(field)
        if not last_tick_iso:
            return False
        last_tick = datetime.fromisoformat(last_tick_iso)
        return (
            datetime.now() - last_tick
        ).total_seconds() < QUEUE_POLL_INTERVAL_SECONDS * 4
    except Exception as e:
        logger.warning(f"Failed to read poll heartbeat {field}: {e}")
        return False


def _count_lifecycle_status(status: str) -> int:
    """Cheap Firestore .count() aggregation on a single-field equality query
    -- already auto-indexed, no composite index needed."""
    try:
        db = firestore.client()
        return (
            db.collection("case-details")
            .where("lifecycle_status", "==", status)
            .count()
            .get()[0][0]
            .value
        )
    except Exception as e:
        logger.error(f"Error counting lifecycle_status={status}: {e}")
        return 0


def _count_stale_in_progress() -> int:
    """Cases genuinely stuck: sitting at fetch_in_progress or
    analysis_in_progress longer than STALE_IN_PROGRESS_MINUTES with nothing
    apparently touching them -- almost always a worker (an instance CPU-
    throttled or torn down mid-run) that claimed the case and then never
    got to finish it. These counts stay small (bounded by how many cases a
    poll tick claims at once, not the historical backlog), so streaming the
    actual docs to check each one's staleness -- not just count() -- is
    cheap here, unlike a full-collection scan."""
    stale = 0
    try:
        db = firestore.client()
        case_store = get_auto_order_manager().case_store
        for status in ("fetch_in_progress", "analysis_in_progress"):
            docs = (
                db.collection("case-details")
                .where("lifecycle_status", "==", status)
                .stream()
            )
            for doc in docs:
                data = doc.to_dict() or {}
                if case_store._is_stale(
                    data.get("lifecycle_status_updated_at"),
                    STALE_IN_PROGRESS_MINUTES,
                ):
                    stale += 1
    except Exception as e:
        logger.error(f"Error counting stale in-progress cases: {e}")
    return stale


def _get_distributed_queue_metrics() -> Dict:
    """The actual, durable queue backlog -- lifecycle_status counts across
    all Cloud Run instances, since there is no in-memory queue to inspect."""
    fetch_pending = sum(
        _count_lifecycle_status(s)
        for s in ("fetch_queued", "fetch_in_progress", "fetch_failed_retryable")
    )
    analysis_pending = sum(
        _count_lifecycle_status(s)
        for s in (
            "analysis_queued",
            "analysis_in_progress",
            "analysis_failed_retryable",
        )
    )
    return {
        "fetch_pending_cases": fetch_pending,
        "analysis_pending_cases": analysis_pending,
    }


@app.get("/queue/status", tags=["Queue Management"])
async def get_queue_status(current_user=Depends(get_current_user)):
    """Get status of fetch and analysis processing, driven entirely by
    persisted lifecycle_status counts (the durable, shared source of truth
    across every Cloud Run instance)."""
    try:
        import time as _time

        if (
            _time.time() - _queue_status_cache["ts"] < 30
            and _queue_status_cache["data"]
        ):
            return JSONResponse(content=_queue_status_cache["data"])

        distributed_metrics = _get_distributed_queue_metrics()
        fetch_queue_size = _count_lifecycle_status("fetch_queued")
        analysis_queue_size = _count_lifecycle_status("analysis_queued")
        fetch_in_progress_count = _count_lifecycle_status("fetch_in_progress")
        analysis_in_progress_count = _count_lifecycle_status("analysis_in_progress")
        review_count = _count_lifecycle_status("manual_review_required")

        # Cases the pipeline could not finish on its own.  The fetch/analyse
        # pipeline runs automatically after upload, so the only thing a user
        # ever needs to act on is this number — it drives the single "needs
        # attention" action on the Dashboard.
        stuck_count = sum(_count_lifecycle_status(s) for s in STUCK_LIFECYCLE_STATUSES)

        fetch_active = _poll_loop_is_active("fetch_last_tick")
        analysis_active = _poll_loop_is_active("analysis_last_tick")
        stale_in_progress_count = _count_stale_in_progress()

        # One combined view of "is the pipeline doing something", matching
        # how it actually behaves: analysis runs inline right after a
        # successful fetch for the normal case, so fetch and analysis are
        # almost always one worker turn per case, not two independent
        # queues. They're only ever tracked separately because they have
        # different retry/timeout characteristics. total_queued/
        # total_in_progress/pipeline_active let the UI show one number and
        # one status instead of two queues that make a single pipeline look
        # like two disconnected ones.
        #
        # pipeline_active is heartbeat-only -- a case merely sitting at
        # fetch_in_progress does NOT mean a worker is actively touching it
        # right now (that's exactly what an orphaned case, one claimed by an
        # instance that was then throttled or torn down before finishing,
        # looks like: total_in_progress > 0 forever, nothing moving).
        # stale_in_progress_count is the honest signal for that: cases
        # claimed but not touched in over STALE_IN_PROGRESS_MINUTES. The
        # poll loop already reclaims these automatically on its next tick --
        # this number existing at all past a tick or two means the loop
        # itself isn't getting a chance to run.
        total_queued = fetch_queue_size + analysis_queue_size
        total_in_progress = fetch_in_progress_count + analysis_in_progress_count
        pipeline_active = fetch_active or analysis_active

        # Every request here is a chance to nudge the loops awake -- if the
        # process is CPU-throttled between requests (Cloud Run's default),
        # the loops' own timer can't reliably fire, but handling THIS
        # request already means the process has CPU right now. Piggybacks
        # pipeline progress on whatever traffic the app already gets
        # (anyone with the Dashboard open polls this every ~15s) instead of
        # relying solely on a timer that may never get scheduled.
        if total_queued > 0 or stale_in_progress_count > 0:
            _wake_fetch_poll.set()
            _wake_analysis_poll.set()

        result = {
            "fetch_queue_size": fetch_queue_size,
            "analysis_queue_size": analysis_queue_size,
            "fetch_in_progress_count": fetch_in_progress_count,
            "analysis_in_progress_count": analysis_in_progress_count,
            "total_queued": total_queued,
            "total_in_progress": total_in_progress,
            "stale_in_progress_count": stale_in_progress_count,
            "fetch_pending_cases": distributed_metrics.get("fetch_pending_cases", 0),
            "analysis_pending_cases": distributed_metrics.get(
                "analysis_pending_cases", 0
            ),
            "review_queue_count": review_count,
            "needs_attention_count": stuck_count,
            "distributed_metrics": distributed_metrics,
            "fetch_processing_active": fetch_active,
            "analysis_processing_active": analysis_active,
            "pipeline_active": pipeline_active,
            "status": "active" if pipeline_active else "inactive",
            "message": (
                f"{total_queued} queued, {total_in_progress} in progress"
                if total_queued > 0 or total_in_progress > 0
                else "Nothing pending"
            ),
        }
        _queue_status_cache["ts"] = _time.time()
        _queue_status_cache["data"] = result
        return JSONResponse(content=result)

    except Exception as e:
        logger.error(f"Error getting queue status: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to get queue status: {str(e)}"}
        )


@app.post("/queue/restart", tags=["Queue Management"])
async def restart_queue_processing(current_user=Depends(require_admin)):
    """Wake both poll loops immediately instead of waiting out their
    interval (admin only). The loops themselves run continuously for the
    life of the process — there is nothing to actually "restart"."""
    try:
        _wake_fetch_poll.set()
        _wake_analysis_poll.set()

        return JSONResponse(
            content={
                "success": True,
                "message": "Fetch and analysis poll loops woken",
                "fetch_processing_active": _poll_loop_is_active("fetch_last_tick"),
                "analysis_processing_active": _poll_loop_is_active(
                    "analysis_last_tick"
                ),
            }
        )

    except Exception as e:
        logger.error(f"Error restarting queue processing: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to restart queue processing: {str(e)}"},
        )


@app.get("/queue/detail", tags=["Queue Management"])
def get_queue_detail(
    limit: int = Query(50, description="Maximum cases to return, oldest first"),
    current_user=Depends(require_admin),
):
    """The actual list of cases currently queued or in-progress, not just an
    aggregate count -- so an admin can see *which* specific cases are
    affected and how long each has been sitting there, instead of only a
    number."""
    try:
        limit = max(1, min(limit, 200))
        db = firestore.client()
        active_statuses = (
            "fetch_queued",
            "fetch_in_progress",
            "analysis_queued",
            "analysis_in_progress",
        )
        in_progress_statuses = {"fetch_in_progress", "analysis_in_progress"}

        cases = []
        for status in active_statuses:
            docs = (
                db.collection("case-details")
                .where("lifecycle_status", "==", status)
                .limit(limit)
                .stream()
            )
            for doc in docs:
                data = doc.to_dict() or {}
                updated_at = data.get("lifecycle_status_updated_at")
                age_seconds = None
                is_stale = False
                if updated_at:
                    try:
                        age_seconds = (
                            datetime.now() - datetime.fromisoformat(updated_at)
                        ).total_seconds()
                        is_stale = (
                            status in in_progress_statuses
                            and age_seconds >= STALE_IN_PROGRESS_MINUTES * 60
                        )
                    except ValueError:
                        pass
                cases.append(
                    {
                        "doc_id": doc.id,
                        "case_ref": data.get("case_ref"),
                        "board_date": data.get("latest_board_date"),
                        "status": status,
                        "updated_at": updated_at,
                        "age_seconds": age_seconds,
                        "stale": is_stale,
                    }
                )

        # Oldest first (None sorts last) so the cases that have waited
        # longest -- the ones most likely to need attention -- lead the list.
        cases.sort(
            key=lambda c: c["age_seconds"] if c["age_seconds"] is not None else -1,
            reverse=True,
        )

        return JSONResponse(
            content={
                "cases": cases[:limit],
                "total_returned": min(len(cases), limit),
                "stale_after_minutes": STALE_IN_PROGRESS_MINUTES,
            }
        )
    except Exception as e:
        logger.error(f"Error getting queue detail: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to get queue detail: {str(e)}"}
        )


@app.get("/admin/queue-health", tags=["Queue Management"])
async def get_queue_health(current_user=Depends(require_admin)):
    """An actual diagnosis of the failed/stuck cases, not just a count --
    groups failures by normalized reason to tell "one flaky case" apart
    from "a systemic problem" (e.g. the court portal changed), and flags
    cases stuck in a claim/retry loop without ever reaching a terminal
    status. Safe to hit on a schedule (e.g. Cloud Scheduler) -- read-only,
    just Firestore equality queries on already-indexed fields."""
    try:
        from queue_health import FAILED_STATUSES, diagnose

        db = firestore.client()
        cases = []
        for status in FAILED_STATUSES:
            for doc in (
                db.collection("case-details")
                .where("lifecycle_status", "==", status)
                .limit(200)
                .stream()
            ):
                data = doc.to_dict() or {}
                data.setdefault("lifecycle_status", status)
                cases.append(data)

        report = diagnose(cases)
        return JSONResponse(content=report)
    except Exception as e:
        logger.error(f"Error getting queue health: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to get queue health: {str(e)}"}
        )


# User Matter Mapping Endpoints
@app.get("/user-matters/my-matters", tags=["User Matter Mapping"])
async def get_my_matters(
    limit: int = Query(100, description="Maximum number of matters to return"),
    current_user=Depends(get_current_user),
):
    """Get matters linked to the current logged-in user"""
    try:
        user_id = current_user.get("uid")
        matches = get_user_matter_matcher().find_user_matters(user_id, limit)

        # Convert dataclass objects to dictionaries
        matters_data = []
        for match in matches:
            matters_data.append(
                {
                    "case_id": match.case_id,
                    "case_ref": match.case_ref,
                    "match_source": match.match_source,
                    "match_field": match.match_field,
                    "matched_text": match.matched_text,
                    "confidence_score": match.confidence_score,
                    "role_type": match.role_type,
                    "board_date": match.board_date,
                }
            )

        return JSONResponse(
            content={
                "user_id": user_id,
                "total_matches": len(matters_data),
                "matters": matters_data,
            }
        )

    except Exception as e:
        logger.error(f"Error getting user matters: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to get user matters: {str(e)}"}
        )


@app.post("/admin/remap-user-matters", tags=["Admin"])
async def admin_remap_user_matters(
    start_date: str = Query(..., description="Board date range start, YYYY-MM-DD"),
    end_date: str = Query(..., description="Board date range end, YYYY-MM-DD"),
    limit: int = Query(500, description="Maximum board rows to remap (1-2000)"),
    current_user=Depends(require_admin_active),
):
    """Re-run AGP name matching over already-analysed board rows.

    auto_map_case_to_users is otherwise only ever called from the two
    poll-loop success paths, so a case that finished analysis without being
    mapped is never revisited -- it is already `analysed`, so nothing will
    reprocess it, and its matters stay missing from that AGP's bill forever
    with no error anywhere. That is exactly what happened to every case
    processed while the poll loops were passing a case-details doc id into a
    lookup that needed a daily-boards one.

    Writes are idempotent: mapping keys are
    {user}_{case}_{source}_{field} written with merge=True, so re-running
    over a range that is already correct is a no-op rather than a duplicate.
    """
    try:
        from AutoOrderManager import AutoOrderManager

        limit = max(1, min(limit, 2000))
        db = firestore.client()

        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            return JSONResponse(
                status_code=400,
                content={"error": "start_date and end_date must be YYYY-MM-DD"},
            )
        if start_dt > end_dt:
            return JSONResponse(
                status_code=400,
                content={"error": "start_date must not be after end_date"},
            )

        rows = (
            db.collection("daily-boards")
            .where("board_date", ">=", start_dt)
            .where("board_date", "<=", end_dt)
            .limit(limit)
            .stream()
        )

        remapped = 0
        failed = 0
        for row in rows:
            data = row.to_dict() or {}
            case_ref = AutoOrderManager.build_case_ref_from_data(data)
            if not case_ref:
                continue
            try:
                # row.id IS the daily-boards doc id the matcher needs -- this
                # endpoint reads the board collection directly, so there is no
                # id-shape translation to get wrong.
                await auto_map_case_to_users(
                    row.id, {"case_ref": case_ref, "board_date": data.get("board_date")}
                )
                remapped += 1
            except Exception as remap_error:
                failed += 1
                logger.error(f"Remap failed for {case_ref}: {remap_error}")

        logger.info(
            f"Remap user matters {start_date}..{end_date}: "
            f"{remapped} remapped, {failed} failed"
        )
        return JSONResponse(
            content={
                "success": True,
                "date_range": {"start": start_date, "end": end_date},
                "rows_remapped": remapped,
                "rows_failed": failed,
                "limit_applied": limit,
                "truncated": remapped + failed >= limit,
            }
        )
    except Exception as e:
        logger.error(f"Error remapping user matters: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to remap: {str(e)}"}
        )


@app.post("/admin/repair-order-board-dates", tags=["Admin"])
async def admin_repair_order_board_dates(
    limit: int = Query(200, description="Max case-details docs to scan (1-1000)"),
    start_after: Optional[str] = Query(
        None, description="Doc id to resume pagination from (from a prior response)"
    ),
    current_user=Depends(require_admin_active),
):
    """Re-tag stored order entries whose board_date disagrees with their own
    order_date -- the write-side fix for the two order-linking bugs (analysis
    path used the case's latest board date instead of the order's own date;
    manual linking wrote no date at all).

    Search Orders resolves which order to show against a board row by
    matching orders[].board_date to that row's own date
    (Board._hydrate_with_case_details), so a wrong board_date here is why an
    order can show against the wrong hearing, or not show at all. Only that
    fix changes future writes -- entries already stored wrong stay wrong
    until repaired, since their case is already `analysed` and nothing
    reprocesses it.

    Pure data fix: rewrites board_date to match order_date on existing
    orders[] entries. No re-download, no portal call, no daily-boards write
    needed -- _update_board_entries_for_case_date already keyed off the
    order's own date, so daily-boards was never wrong, only case-details'
    orders[] array was.

    Paginated and idempotent: call repeatedly, passing next_start_after from
    the previous response, until it comes back null.
    """
    try:
        limit = max(1, min(limit, 1000))
        db = firestore.client()

        query = db.collection("case-details").order_by("__name__").limit(limit)
        if start_after:
            start_doc = db.collection("case-details").document(start_after).get()
            if start_doc.exists:
                query = query.start_after(start_doc)

        docs = list(query.stream())
        docs_updated = 0
        entries_fixed = 0
        last_id = None

        for doc in docs:
            last_id = doc.id
            data = doc.to_dict() or {}
            orders = data.get("orders")
            if not isinstance(orders, list) or not orders:
                continue

            changed = False
            fixed_orders = []
            for order in orders:
                if not isinstance(order, dict):
                    fixed_orders.append(order)
                    continue
                order_date = order.get("order_date")
                if order_date and order.get("board_date") != order_date:
                    order = {**order, "board_date": order_date}
                    changed = True
                    entries_fixed += 1
                fixed_orders.append(order)

            if changed:
                doc.reference.update({"orders": fixed_orders})
                docs_updated += 1

        logger.info(
            f"Repair order board_dates: scanned={len(docs)} "
            f"docs_updated={docs_updated} entries_fixed={entries_fixed}"
        )
        return JSONResponse(
            content={
                "success": True,
                "docs_scanned": len(docs),
                "docs_updated": docs_updated,
                "entries_fixed": entries_fixed,
                "next_start_after": last_id if len(docs) == limit else None,
            }
        )
    except Exception as e:
        logger.error(f"Error repairing order board dates: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Repair failed: {str(e)}"}
        )


@app.get("/user-matters/pending-confirmations", tags=["User Matter Mapping"])
async def get_pending_matter_confirmations(current_user=Depends(get_current_user)):
    """Matches that fell just short of the auto-accept threshold -- 'is this
    you?' candidates the current user can confirm or dismiss, instead of a
    real matter assignment being silently missed (roadmap #9)."""
    try:
        user_id = current_user.get("uid")
        db = firestore.client()
        docs = (
            db.collection("user-matter-pending-confirmations")
            .where("user_id", "==", user_id)
            .where("status", "==", "pending")
            .stream()
        )
        pending = []
        for doc in docs:
            data = doc.to_dict() or {}
            data["id"] = doc.id
            pending.append(data)

        return JSONResponse(content={"pending": pending, "total": len(pending)})
    except Exception as e:
        logger.error(f"Error getting pending matter confirmations: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get pending confirmations: {str(e)}"},
        )


def _load_owned_pending_confirmation(db, confirmation_id: str, uid: str):
    """Fetch a pending-confirmation doc and verify it belongs to ``uid``.
    Returns (doc_ref, data) or (None, error_response)."""
    doc_ref = db.collection("user-matter-pending-confirmations").document(
        confirmation_id
    )
    snapshot = doc_ref.get()
    if not snapshot.exists:
        return None, JSONResponse(
            status_code=404, content={"error": "Confirmation not found"}
        )
    data = snapshot.to_dict() or {}
    if data.get("user_id") != uid:
        # Deliberately the same 404 as "not found" rather than 403 -- a user
        # should not be able to tell someone else's pending confirmation
        # exists at all.
        return None, JSONResponse(
            status_code=404, content={"error": "Confirmation not found"}
        )
    return doc_ref, data


@app.post(
    "/user-matters/pending-confirmations/{confirmation_id}/confirm",
    tags=["User Matter Mapping"],
)
async def confirm_pending_matter(
    confirmation_id: str, current_user=Depends(get_current_user)
):
    """User confirms a near-miss match really is them -- creates the real
    matter mapping, same shape as an auto-accepted one."""
    try:
        db = firestore.client()
        uid = current_user.get("uid")
        doc_ref, data = _load_owned_pending_confirmation(db, confirmation_id, uid)
        if doc_ref is None:
            return data  # the error JSONResponse

        mapping_key = (
            f"{data['user_id']}_{data['case_id']}_{data['match_source']}_"
            f"{data['match_field']}"
        )
        db.collection("user-case-mappings").document(mapping_key).set(
            {
                "user_id": data["user_id"],
                "case_id": data["case_id"],
                "case_ref": data.get("case_ref"),
                "match_source": data["match_source"],
                "match_field": data["match_field"],
                "matched_text": data.get("matched_text"),
                "confidence_score": data.get("confidence_score"),
                "role_type": data.get("role_type"),
                "board_date": data.get("board_date"),
                "mapped_at": firestore.SERVER_TIMESTAMP,
                "auto_mapped": False,
                "confirmed_by_user": True,
            },
            merge=True,
        )
        doc_ref.set(
            {"status": "confirmed", "resolved_at": firestore.SERVER_TIMESTAMP},
            merge=True,
        )
        return JSONResponse(content={"success": True, "case_ref": data.get("case_ref")})
    except Exception as e:
        logger.error(f"Error confirming pending matter {confirmation_id}: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to confirm match: {str(e)}"}
        )


@app.post(
    "/user-matters/pending-confirmations/{confirmation_id}/reject",
    tags=["User Matter Mapping"],
)
async def reject_pending_matter(
    confirmation_id: str, current_user=Depends(get_current_user)
):
    """User says a near-miss match isn't them -- no mapping is created."""
    try:
        db = firestore.client()
        uid = current_user.get("uid")
        doc_ref, data = _load_owned_pending_confirmation(db, confirmation_id, uid)
        if doc_ref is None:
            return data  # the error JSONResponse

        doc_ref.set(
            {"status": "rejected", "resolved_at": firestore.SERVER_TIMESTAMP},
            merge=True,
        )
        return JSONResponse(content={"success": True})
    except Exception as e:
        logger.error(f"Error rejecting pending matter {confirmation_id}: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to reject match: {str(e)}"}
        )


@app.get("/user-matters/summary", tags=["User Matter Mapping"])
async def get_my_matters_summary(current_user=Depends(get_current_user)):
    """Get summary statistics of matters for the current user"""
    try:
        user_id = current_user.get("uid")
        summary = get_user_matter_matcher().get_matters_summary(user_id)

        return JSONResponse(content={"user_id": user_id, "summary": summary})

    except Exception as e:
        logger.error(f"Error getting matters summary: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get matters summary: {str(e)}"},
        )


@app.get("/user-matters/role-config", tags=["User Matter Mapping"])
async def get_user_role_config(current_user=Depends(get_current_user)):
    """Get current user's role configuration"""
    try:
        user_id = current_user.get("uid")
        user_role = get_user_matter_matcher().get_user_role_config(user_id)

        if not user_role:
            return JSONResponse(
                content={
                    "user_id": user_id,
                    "role_configured": False,
                    "message": "No role configuration found. Please configure your legal role and name variations.",
                }
            )

        return JSONResponse(
            content={
                "user_id": user_id,
                "role_configured": True,
                "role_config": {
                    "role_type": user_role.role_type,
                    "full_name": user_role.full_name,
                    "name_variations": user_role.name_variations,
                    "pattern_keywords": user_role.pattern_keywords,
                    "confidence_threshold": user_role.confidence_threshold,
                },
            }
        )

    except Exception as e:
        logger.error(f"Error getting user role config: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to get role config: {str(e)}"}
        )


@app.post("/user-matters/configure-role", tags=["User Matter Mapping"])
async def configure_user_role(request: Request, current_user=Depends(get_current_user)):
    """Configure user's legal role and name variations for matter matching"""
    try:
        user_id = current_user.get("uid")
        body = await request.json()

        # Validate required fields
        role_type = body.get("role_type")
        full_name = body.get("full_name")

        if not role_type or not full_name:
            return JSONResponse(
                status_code=400,
                content={"error": "role_type and full_name are required"},
            )

        # Valid role types
        valid_roles = ["AGP", "GP", "Addl_GP", "B_Pnl", "State_Advocate", "AG"]
        if role_type not in valid_roles:
            return JSONResponse(
                status_code=400,
                content={
                    "error": f"Invalid role_type. Must be one of: {', '.join(valid_roles)}"
                },
            )

        # Generate name variations if not provided
        name_variations = body.get("name_variations", [])
        if not name_variations:
            name_variations = get_user_matter_matcher().generate_name_variations(
                full_name
            )

        # Create user role configuration
        user_role = UserRole(
            role_type=role_type,
            full_name=full_name,
            name_variations=name_variations,
            pattern_keywords=body.get("pattern_keywords", []),
            confidence_threshold=body.get("confidence_threshold", 0.75),
        )

        # Save configuration
        success = get_user_matter_matcher().save_user_role_config(user_id, user_role)

        if success:
            return JSONResponse(
                content={
                    "success": True,
                    "message": "Role configuration saved successfully",
                    "user_id": user_id,
                    "role_config": {
                        "role_type": user_role.role_type,
                        "full_name": user_role.full_name,
                        "name_variations": user_role.name_variations,
                        "pattern_keywords": user_role.pattern_keywords,
                        "confidence_threshold": user_role.confidence_threshold,
                    },
                }
            )
        else:
            return JSONResponse(
                status_code=500, content={"error": "Failed to save role configuration"}
            )

    except Exception as e:
        logger.error(f"Error configuring user role: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to configure role: {str(e)}"}
        )


@app.post("/user-matters/generate-name-variations", tags=["User Matter Mapping"])
async def generate_name_variations(
    request: Request, current_user=Depends(get_current_user)
):
    """Generate name variations for a given full name (helper endpoint)"""
    try:
        body = await request.json()
        full_name = body.get("full_name")

        if not full_name:
            return JSONResponse(
                status_code=400, content={"error": "full_name is required"}
            )

        variations = get_user_matter_matcher().generate_name_variations(full_name)

        return JSONResponse(
            content={
                "full_name": full_name,
                "name_variations": variations,
                "total_variations": len(variations),
            }
        )

    except Exception as e:
        logger.error(f"Error generating name variations: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to generate name variations: {str(e)}"},
        )


# Order Management Center Endpoints
@app.get("/orders/overview-stats", tags=["Order Management"])
async def get_order_overview_stats(current_user=Depends(get_current_user)):
    """Get comprehensive overview statistics for Order Management Center"""
    try:
        import time as _time

        if (
            _time.time() - _overview_stats_cache["ts"] < 120
            and _overview_stats_cache["data"]
        ):
            return JSONResponse(content=_overview_stats_cache["data"])

        ensure_firebase()
        db = firestore.client()

        # Use Firestore count() aggregation — reads 0 documents, billed as 1 read each
        total_cases = db.collection("daily-boards").count().get()[0][0].value
        analysed = _count_case_details_by_order_status("analysed")
        order_failed = _count_case_details_by_order_status("order_failed")
        order_analysis_failed = _count_case_details_by_order_status(
            "order_analysis_failed"
        )
        not_linked = _count_case_details_by_order_status("not_linked")
        total_case_details = db.collection("case-details").count().get()[0][0].value

        # "Fetched" (Step 2, the workflow strip's "order PDFs downloaded") is
        # cases whose PDF actually downloaded -- order_analysis_failed still
        # means the download itself succeeded (only the read failed after),
        # order_failed means it didn't. Was previously
        # total_case_details - not_linked, which counted order_failed cases
        # as "downloaded" too.
        cases_with_orders = total_case_details - not_linked - order_failed
        cases_without_orders = not_linked + order_failed

        # "Analysed" (Step 3, "read and categorised") is a case-level
        # percentage of unique cases, not a percentage of board-row
        # appearances -- fetch/analysis happens once per case_ref regardless
        # of how many board dates it's listed on. Was previously
        # cases_with_orders / total_cases, which (a) mixed a case-details
        # numerator with a daily-boards denominator, different units for a
        # case listed more than once, and (b) counted every downloaded-but-
        # not-yet-analysed and every failed case as "analysed", which is why
        # this number tracked "Fetch orders" almost exactly instead of
        # measuring analysis progress at all.
        analysis_completion_rate = round(
            (analysed / total_case_details * 100) if total_case_details > 0 else 0, 1
        )

        result = {
            "total_cases": total_cases,
            "cases_with_orders": cases_with_orders,
            "cases_without_orders": cases_without_orders,
            "analysis_completion_rate": analysis_completion_rate,
            "recent_successful_analyses": analysed,
            "recent_failed_analyses": order_failed + order_analysis_failed,
            "last_updated": datetime.now().isoformat(),
        }
        _overview_stats_cache["ts"] = _time.time()
        _overview_stats_cache["data"] = result
        return JSONResponse(content=result)

    except Exception as e:
        logger.error(f"Error getting order overview stats: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get overview stats: {str(e)}"},
        )


_LIFECYCLE_ACTION_LABELS = {
    "analysed": ("Analysis Complete", "success"),
    "fetch_succeeded": ("Order Downloaded", "success"),
    "fetch_failed_terminal": ("Download Failed", "error"),
    "fetch_failed_retryable": ("Download Failed (Retrying)", "warning"),
    "analysis_failed_terminal": ("Analysis Failed", "error"),
    "analysis_failed_retryable": ("Analysis Failed (Retrying)", "warning"),
    "manual_review_required": ("Manual Review Required", "warning"),
    "fetch_in_progress": ("Fetching Order", "info"),
    "analysis_in_progress": ("Analysing Order", "info"),
    "fetch_queued": ("Queued for Download", "info"),
    "analysis_queued": ("Queued for Analysis", "info"),
}


@app.get("/orders/recent-activity", tags=["Order Management"])
def get_recent_activity(
    limit: int = Query(20, description="Number of recent activities to return"),
    current_user=Depends(get_current_user),
):
    """Get recent order processing activity from case-details lifecycle events."""
    try:
        db = firestore.client()
        docs = (
            db.collection("case-details")
            .order_by("updated_at", direction=firestore.Query.DESCENDING)
            .limit(limit)
            .stream()
        )
        recent_activity = []
        for doc in docs:
            data = doc.to_dict() or {}
            lifecycle_status = str(data.get("lifecycle_status") or "")
            action, status = _LIFECYCLE_ACTION_LABELS.get(
                lifecycle_status, ("Status Update", "info")
            )
            case_type = data.get("case_type") or ""
            case_no = str(data.get("case_no") or "")
            case_year = str(data.get("case_year") or "")
            if case_type and case_no and case_year:
                case_ref = f"{case_type}/{case_no}/{case_year}"
            else:
                case_ref = doc.id.replace("-", "/", 2)
            updated_at = data.get("updated_at")
            timestamp = (
                updated_at.isoformat()
                if hasattr(updated_at, "isoformat")
                else datetime.now().isoformat()
            )
            recent_activity.append(
                {
                    "timestamp": timestamp,
                    "action": action,
                    "case_ref": case_ref,
                    "status": status,
                    "lifecycle_status": lifecycle_status,
                }
            )
        return JSONResponse(content=recent_activity)
    except Exception as e:
        logger.error(f"Error getting recent activity: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get recent activity: {str(e)}"},
        )


@app.get("/orders/pdf/{doc_id}", tags=["Order Management"])
async def get_order_pdf(doc_id: str):
    """Serve a court order PDF with automatic GCS upgrade.

    Deliberately unauthenticated. Every caller is a plain <a href> opened in
    a new tab (Table.jsx, BillGeneration.jsx, CaseDetailModal.jsx) and a
    browser navigation carries no Authorization header, so requiring auth
    here 401s every "View PDF" link in the app. Doc ids are predictable
    (YYYY-MM-DD-TYPE-NO-YEAR), so the stored corpus is enumerable by anyone
    who guesses them -- accepted because these are public court records.
    To close that off, the fix is client-side: fetch with
    authenticatedFetch and open a blob URL, at all three call sites.

    - GCS URLs: fetched via service-account credentials and streamed back
      (no public bucket access required; Cloud Run ADC authenticates).
    - Live court URLs: stream PDF to client and upgrade the stored link to GCS
      in the background so the next access is served from GCS.
    - Expired court URLs: queue re-fetch via AutoOrderManager, return 503.

    No authentication required — court PDFs are public documents, consistent
    with the current behaviour where court URLs are opened directly.
    """
    try:
        ensure_firebase()
        db = firestore.client()
        doc = db.collection("daily-boards").document(doc_id).get()

        case_data: dict = {}
        order_link = ""

        if doc.exists:
            case_data = doc.to_dict() or {}
            # Prefer the board entry's own order_link — written by
            # _update_board_entries_for_case_date so it is date-specific.
            order_link = (case_data.get("order_link") or "").strip()

            if not order_link:
                # Fall back to case-details (covers board entries pre-dating
                # the per-entry order_link write).
                _ct = case_data.get("case_type", "")
                _cn = str(case_data.get("case_no") or "")
                _cy = str(case_data.get("case_year") or "")
                if _ct and _cn and _cy:
                    # doc_id format: YYYY-MM-DD-TYPE-NO-YEAR → extract date portion
                    _order_date = (
                        doc_id[:10] if len(doc_id) > 10 and doc_id[10] == "-" else ""
                    )
                    _details_id = f"{_ct}-{_cn}-{_cy}"
                    _details_snap = (
                        db.collection("case-details").document(_details_id).get()
                    )
                    if _details_snap.exists:
                        _details = _details_snap.to_dict() or {}
                        # Try to find the order matching this board entry's date
                        if _order_date:
                            for _o in _details.get("orders") or []:
                                if isinstance(_o, dict) and _o.get("order_link"):
                                    if (
                                        str(_o.get("order_date", ""))[:10]
                                        == _order_date
                                    ):
                                        order_link = _o["order_link"].strip()
                                        break
                        if not order_link:
                            order_link = (
                                _details.get("latest_order_link") or ""
                            ).strip()
                        if not order_link:
                            for _o in reversed(_details.get("orders") or []):
                                if isinstance(_o, dict) and _o.get("order_link"):
                                    order_link = _o["order_link"].strip()
                                    break
        else:
            # Board entry not found — doc_id may be a constructed ID for a historical
            # order that has no board entry.  Parse: YYYY-MM-DD-{case-details-id}
            # e.g. "2025-07-15-WP-2316-2026" → date "2025-07-15", details "WP-2316-2026"
            if len(doc_id) > 11 and doc_id[10] == "-":
                _order_date = doc_id[:10]
                _details_id = doc_id[11:]
                _details_snap = (
                    db.collection("case-details").document(_details_id).get()
                )
                if _details_snap.exists:
                    _details = _details_snap.to_dict() or {}
                    for _o in _details.get("orders") or []:
                        if isinstance(_o, dict) and _o.get("order_link"):
                            if str(_o.get("order_date", ""))[:10] == _order_date:
                                order_link = _o["order_link"].strip()
                                break
                    if not order_link:
                        order_link = (_details.get("latest_order_link") or "").strip()
            if not order_link:
                raise HTTPException(status_code=404, detail="Case not found")

        if not order_link:
            raise HTTPException(
                status_code=404, detail="No order link stored for this case"
            )

        # GCS URL: download via service-account credentials and stream back.
        # Public bucket access is not required — Cloud Run ADC authenticates
        # transparently, so the bucket can stay private.
        if order_link.startswith("https://storage.googleapis.com"):
            try:
                from google.cloud import storage as gcs_storage

                # Parse  https://storage.googleapis.com/{bucket}/{blob_path}
                without_prefix = order_link[len("https://storage.googleapis.com/") :]
                bucket_name, _, blob_name = without_prefix.partition("/")
                client = gcs_storage.Client()
                pdf_bytes = (
                    client.bucket(bucket_name).blob(blob_name).download_as_bytes()
                )
                return Response(
                    content=pdf_bytes,
                    media_type="application/pdf",
                    headers={
                        "Content-Disposition": f'inline; filename="order-{doc_id}.pdf"'
                    },
                )
            except Exception as gcs_err:
                logger.warning(
                    "get_order_pdf: GCS download failed for doc_id=%s: %s",
                    doc_id,
                    gcs_err,
                )
                # Blob missing or inaccessible — queue a re-fetch so the PDF is
                # retrieved from the court API and re-uploaded to GCS, exactly as
                # we do for expired court URLs.
                _case_type = case_data.get("case_type", "")
                _case_no = str(case_data.get("case_no") or "")
                _case_year = str(case_data.get("case_year") or "")
                if _case_type and _case_no:
                    _refetch_data = {**case_data, "id": doc_id}
                    if not _refetch_data.get("case_ref") and _case_type and _case_year:
                        _refetch_data[
                            "case_ref"
                        ] = f"{_case_type}/{_case_no}/{_case_year}"
                    loop = asyncio.get_event_loop()
                    loop.run_in_executor(
                        executor,
                        get_auto_order_manager()._process_single_case,
                        _refetch_data,
                    )
                return JSONResponse(
                    status_code=503,
                    content={
                        "error": "order_link_expired",
                        "message": (
                            "Order PDF is unavailable. The system is re-fetching it — "
                            "please try again in a few minutes."
                        ),
                        "doc_id": doc_id,
                    },
                )

        # Court URL: attempt download
        try:
            resp = requests.get(order_link, timeout=15)
            resp.raise_for_status()
            pdf_bytes = resp.content
            if not pdf_bytes or b"%PDF" not in pdf_bytes[:10]:
                raise ValueError("Response is not a valid PDF")
        except Exception:
            # Expired or unreachable — queue re-fetch and tell the client to retry
            case_type = case_data.get("case_type", "")
            case_no = str(case_data.get("case_no") or "")
            case_year = str(case_data.get("case_year") or "")
            if case_type and case_no:
                _refetch_data = {**case_data, "id": doc_id}
                if not _refetch_data.get("case_ref") and case_type and case_year:
                    _refetch_data["case_ref"] = f"{case_type}/{case_no}/{case_year}"
                loop = asyncio.get_event_loop()
                loop.run_in_executor(
                    executor,
                    get_auto_order_manager()._process_single_case,
                    _refetch_data,
                )
            return JSONResponse(
                status_code=503,
                content={
                    "error": "order_link_expired",
                    "message": (
                        "Order link has expired. The system is re-fetching it — "
                        "please try again in a few minutes."
                    ),
                    "doc_id": doc_id,
                },
            )

        # Court URL still live: serve the PDF and upgrade to GCS in the background
        case_type = case_data.get("case_type", "")
        case_no = str(case_data.get("case_no") or "")
        case_year = str(case_data.get("case_year") or "")
        case_ref = f"{case_type}/{case_no}/{case_year}"
        # Use the actual order date from case-details to match the blob name used
        # during the original upload. Fall back through: latest_order_date →
        # the matching order entry's date → board_date → today (last resort, avoids
        # creating an orphan GCS blob with today's date that can't be linked back).
        _raw_order_date = _details.get("latest_order_date")
        if not _raw_order_date:
            # Walk orders to find the entry whose order_link matches the URL we just
            # served — that gives us the correct date for the GCS blob name.
            for _o in reversed(_details.get("orders") or []):
                if isinstance(_o, dict) and _o.get("order_link") == order_link:
                    _raw_order_date = _o.get("order_date")
                    break
        if not _raw_order_date:
            _raw_order_date = case_data.get("board_date") or datetime.now().strftime(
                "%Y-%m-%d"
            )
        # Firestore Timestamps stringify as "YYYY-MM-DD HH:MM:SS"; strip the time
        # component so GCS blob names don't contain spaces.
        order_date = str(_raw_order_date).split(" ")[0].split("T")[0]

        def _upgrade_to_gcs(
            _pdf: bytes, _case_ref: str, _order_date: str, _doc_id: str
        ) -> None:
            mgr = get_auto_order_manager()
            if not mgr._gcs_bucket_name:
                logger.debug(
                    "get_order_pdf: GCS upgrade skipped for doc_id=%s — "
                    "ORDER_PDF_BUCKET not configured",
                    _doc_id,
                )
                return
            gcs_url = mgr._upload_order_to_gcs(_pdf, _case_ref, _order_date)
            if not gcs_url:
                logger.warning(
                    "get_order_pdf: GCS upload failed for doc_id=%s case_ref=%s — "
                    "order_link remains as court URL. Run GET /admin/test-gcs to diagnose.",
                    _doc_id,
                    _case_ref,
                )
                return
            try:
                mgr.case_store.append_case_order(
                    _case_ref, {"order_link": gcs_url, "order_date": _order_date}
                )
                logger.info(
                    "get_order_pdf: upgraded order_link to GCS for doc_id=%s", _doc_id
                )
            except Exception as _e:
                logger.warning(
                    "get_order_pdf: Firestore update after GCS upload failed: %s", _e
                )

        asyncio.get_event_loop().run_in_executor(
            executor, _upgrade_to_gcs, pdf_bytes, case_ref, order_date, doc_id
        )

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="order-{doc_id}.pdf"'},
        )

    except HTTPException:
        raise
    except Exception as exc:
        logger.error("get_order_pdf failed for doc_id=%s: %s", doc_id, exc)
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/admin/gcs-bucket-info", tags=["Admin"])
async def gcs_bucket_info(current_user=Depends(require_admin)):
    """Return GCS bucket metadata so admins can verify no lifecycle rules are deleting blobs.

    If the bucket has a lifecycle rule that deletes objects after N days, order PDFs
    would silently disappear, causing the proxy to return 503 even for GCS URLs.
    """
    try:
        from google.cloud import storage as gcs_storage

        mgr = get_auto_order_manager()
        bucket_name = mgr._gcs_bucket_name
        if not bucket_name:
            return JSONResponse(
                status_code=400,
                content={"error": "ORDER_PDF_BUCKET env var not set on this instance"},
            )
        client = gcs_storage.Client()
        bucket = client.get_bucket(bucket_name)
        rules = []
        if bucket.lifecycle_rules:
            for rule in bucket.lifecycle_rules:
                rules.append(rule)
        return JSONResponse(
            content={
                "bucket": bucket_name,
                "location": bucket.location,
                "storage_class": bucket.storage_class,
                "lifecycle_rules": rules,
                "lifecycle_rule_count": len(rules),
                "versioning_enabled": bucket.versioning_enabled,
                "diagnosis": (
                    "No lifecycle rules — blobs are retained indefinitely."
                    if not rules
                    else f"WARNING: {len(rules)} lifecycle rule(s) found — "
                    "they may be deleting order PDFs. Check rules above."
                ),
            }
        )
    except Exception as exc:
        logger.error("gcs_bucket_info failed: %s", exc)
        return JSONResponse(
            status_code=500,
            content={"error": str(exc)},
        )


@app.post("/orders/migrate-to-gcs", tags=["Order Management"])
async def migrate_orders_to_gcs(
    limit: int = Query(100, description="Max docs to process per call (max 500)"),
    current_user=Depends(require_admin),
):
    """Admin: backfill existing court order URLs to permanent GCS URLs.

    Scans ``daily-boards`` documents that hold expiring court URLs, downloads
    each PDF, uploads it to the configured GCS bucket, and updates Firestore.
    Run repeatedly with the default ``limit`` until ``skipped`` == ``total_scanned``.
    """
    mgr = get_auto_order_manager()
    if not mgr._gcs_bucket_name:
        raise HTTPException(
            status_code=400,
            detail=(
                "ORDER_PDF_BUCKET is not configured. "
                "Set the environment variable and redeploy before running the backfill."
            ),
        )

    limit = min(limit, 500)
    db = firestore.client()
    migrated = skipped = failed = 0

    # Firestore has no "not starts-with" filter; over-fetch and filter in Python
    docs = list(
        db.collection("daily-boards")
        .where("order_link", "!=", "")
        .limit(limit * 5)
        .stream()
    )

    for doc in docs:
        if migrated + failed >= limit:
            break

        data = doc.to_dict() or {}
        order_link = (data.get("order_link") or "").strip()

        if not order_link or order_link.startswith("https://storage.googleapis.com"):
            skipped += 1
            continue

        case_type = data.get("case_type", "")
        case_no = str(data.get("case_no") or "")
        case_year = str(data.get("case_year") or "")
        case_ref = f"{case_type}/{case_no}/{case_year}"
        order_date = str(data.get("latest_order_date") or data.get("board_date") or "")

        try:
            resp = requests.get(order_link, timeout=15)
            resp.raise_for_status()
            pdf_bytes = resp.content
            if not pdf_bytes or b"%PDF" not in pdf_bytes[:10]:
                raise ValueError("Not a valid PDF")

            gcs_url = mgr._upload_order_to_gcs(pdf_bytes, case_ref, order_date)
            if not gcs_url:
                raise ValueError("GCS upload returned None")

            doc.reference.update({"order_link": gcs_url})
            mgr.case_store.append_case_order(
                case_ref, {"order_link": gcs_url, "order_date": order_date}
            )
            migrated += 1
        except Exception as exc:
            logger.warning("migrate-to-gcs failed for %s: %s", doc.id, exc)
            failed += 1

    return {
        "migrated": migrated,
        "skipped": skipped,
        "failed": failed,
        "total_scanned": len(docs),
    }


# Bill Generation Endpoints
@app.get("/bills/generate", tags=["Bill Generation"])
async def generate_bill_data(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    user_name: Optional[str] = Query(
        None, description="User's full name to generate bill for (admin only)"
    ),
    current_user=Depends(get_current_user),
):
    """Generate bill data for logged-in user or specific user (admin only) based on date range"""
    try:
        user_id = current_user.get("uid")
        is_admin = get_user_manager().is_admin(user_id)

        # Initialize Firestore client
        db = firestore.client()

        # Parse dates
        from datetime import datetime

        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")

        bill_entries = []
        case_ids = set()

        # Admin can generate bill for any user, non-admin only for themselves
        if user_name and not is_admin:
            return JSONResponse(
                status_code=403,
                content={
                    "error": "Only administrators can generate bills for other users"
                },
            )

        # Determine which cases to include
        if user_name:
            # Admin generating bill for specific user - use ENHANCED fuzzy matching
            logger.info(f"Admin {user_id} generating bill for user: {user_name}")

            # Step 1: Collect all unique AGP names with PRIORITY ORDER
            # Priority: 1) government_pleader (from order analysis)
            #          2) respondent_lawyer (from board data)
            #          3) additional_respondent_lawyers (from board data)
            # Restrict to the requested date range so we don't scan the entire
            # collection — avoids timeouts on large deployments.
            boards_ref = db.collection("daily-boards")
            all_cases = (
                boards_ref.where("board_date", ">=", start_dt)
                .where("board_date", "<=", end_dt)
                .stream()
            )

            unique_agp_names = set()
            cases_by_agp: Dict[str, List[Any]] = {}

            for case_doc in all_cases:
                case_data = case_doc.to_dict()
                case_id = case_doc.id

                # PRIORITY 1: government_pleader (from order analysis)
                # This is the MOST ACCURATE source as it's extracted from the actual court order
                government_pleader = case_data.get("government_pleader")
                has_order_agp = False

                if government_pleader:
                    # Handle both string and list formats
                    if isinstance(government_pleader, str):
                        # Single string - treat as one name
                        agp_name = government_pleader.strip()
                        if agp_name:
                            unique_agp_names.add(agp_name)
                            if agp_name not in cases_by_agp:
                                cases_by_agp[agp_name] = []
                            cases_by_agp[agp_name].append((case_id, case_data))
                            has_order_agp = True
                    elif isinstance(government_pleader, list):
                        # List of names
                        for agp_name in government_pleader:
                            agp_name = str(agp_name).strip() if agp_name else ""
                            if agp_name:
                                unique_agp_names.add(agp_name)
                                if agp_name not in cases_by_agp:
                                    cases_by_agp[agp_name] = []
                                cases_by_agp[agp_name].append((case_id, case_data))
                                has_order_agp = True

                # PRIORITY 2 & 3: Only use board data if NO government_pleader exists
                # This ensures government_pleader from order analysis takes precedence
                if not has_order_agp:
                    # Source 2: respondent_lawyer (from board data)
                    respondent_lawyer = case_data.get("respondent_lawyer", "").strip()
                    if respondent_lawyer:
                        unique_agp_names.add(respondent_lawyer)
                        if respondent_lawyer not in cases_by_agp:
                            cases_by_agp[respondent_lawyer] = []
                        cases_by_agp[respondent_lawyer].append((case_id, case_data))

                    # Source 3: additional_respondent_lawyers (from board data)
                    additional_lawyers = case_data.get(
                        "additional_respondent_lawyers", []
                    )
                    if additional_lawyers and isinstance(additional_lawyers, list):
                        for lawyer_name in additional_lawyers:
                            lawyer_name = lawyer_name.strip().rstrip(",")
                            if lawyer_name:
                                unique_agp_names.add(lawyer_name)
                                if lawyer_name not in cases_by_agp:
                                    cases_by_agp[lawyer_name] = []
                                cases_by_agp[lawyer_name].append((case_id, case_data))

            logger.info(
                f"📚 Collected {len(unique_agp_names)} unique AGP names (PRIORITY: government_pleader > respondent_lawyer > additional_respondent_lawyers)"
            )

            # Log sample of AGP names for debugging
            agp_names_list = sorted(list(unique_agp_names))
            logger.info(f"📝 Sample AGP names (first 10): {agp_names_list[:10]}")
            logger.info(f"📝 Sample AGP names (last 10): {agp_names_list[-10:]}")

            # Step 2: Use ENHANCED fuzzy matching with initials support
            # Changed: Instead of finding only the BEST match, find ALL AGP names that match with >= 50% confidence
            user_manager = get_user_manager()
            threshold = 0.50

            # Find ALL matching AGP names with scores >= 50% in one efficient pass
            all_matching_agps = user_manager.match_user_name_to_all_agps(
                user_name, list(unique_agp_names), threshold=threshold
            )

            if all_matching_agps:
                # Step 3: Collect cases from ALL matching AGP names
                matched_cases = []
                matched_variants = []
                for agp_variant, confidence in all_matching_agps:
                    if agp_variant in cases_by_agp:
                        variant_cases = cases_by_agp[agp_variant]
                        matched_cases.extend(variant_cases)
                        matched_variants.append(agp_variant)
                        logger.info(
                            f"   📁 '{agp_variant}' ({confidence:.0%}): {len(variant_cases)} cases"
                        )

                # Use the best match for display purposes
                matched_agp = all_matching_agps[0][0] if all_matching_agps else None
                confidence = all_matching_agps[0][1] if all_matching_agps else 0.0

                logger.info(
                    f"📊 Collected {len(matched_variants)} AGP variants matching '{user_name}'"
                )
                logger.info(f"📁 Total cases across all variants: {len(matched_cases)}")

                # Track filtering for debugging
                date_filtered = 0
                duplicate_filtered = 0

                for case_id, case_data in matched_cases:
                    board_date_raw = case_data.get("board_date")

                    if board_date_raw:
                        try:
                            # Handle both Firestore Timestamp and string formats
                            if isinstance(board_date_raw, str):
                                board_date = datetime.strptime(
                                    board_date_raw, "%Y-%m-%d"
                                )
                                board_date_str = board_date_raw
                            else:
                                # Firestore DatetimeWithNanoseconds object - convert to naive datetime
                                board_date = board_date_raw.replace(tzinfo=None)
                                board_date_str = board_date.strftime("%Y-%m-%d")

                            # Check if case falls within date range
                            if (
                                start_dt <= board_date <= end_dt
                                and case_id not in case_ids
                            ):
                                case_ids.add(case_id)

                                # Determine fee and result based on order analysis
                                fee_info = calculate_case_fee(
                                    case_data, board_date=board_date_str
                                )

                                # Extract parties information
                                parties = extract_parties_info(case_data)

                                bill_entry = {
                                    "id": case_id,
                                    "date": board_date_str,
                                    "case_detail": f"{case_data.get('case_type')}/{case_data.get('case_no')}/{case_data.get('case_year')}",
                                    "case_type": case_data.get("case_type", ""),
                                    "case_no": case_data.get("case_no", ""),
                                    "case_year": case_data.get("case_year", ""),
                                    "parties_name": parties,
                                    "results": fee_info["result"],
                                    "fees_rs": fee_info["fee"],
                                    "order_link": fee_info.get("order_link"),
                                    "order_category": fee_info.get("order_category"),
                                    "agp_name": matched_agp,  # Show the actual AGP name from data
                                    "user_name": user_name,  # Show the selected user name
                                    "name_match_confidence": round(confidence, 3),
                                    # The UI reads confidence_score; the other
                                    # bill-entry path already emitted that name,
                                    # so this one did too — meaning its
                                    # low-confidence warning never fired.
                                    "confidence_score": round(confidence, 3),
                                    "order_category_confidence": fee_info.get(
                                        "order_category_confidence"
                                    ),
                                    "editable": True,
                                }
                                bill_entries.append(bill_entry)
                        except ValueError:
                            logger.warning(
                                f"Invalid date format for case {case_id}: {board_date_str}"
                            )
                            continue

                logger.info(
                    f"📊 Filtering summary: {len(matched_cases)} total cases → {len(bill_entries)} included"
                )
                logger.info(f"   - Date range: {start_date} to {end_date}")
                logger.info(
                    f"   - Cases outside date range: {len(matched_cases) - len(bill_entries)}"
                )
                logger.info(
                    f"✅ Found {len(bill_entries)} bill entries for user '{user_name}'"
                )
            else:
                # No AGP name in the date range matched the requested user name.
                # bill_entries stays empty — return a valid empty bill rather than
                # a 400, so the UI shows "0 entries" instead of an error banner.
                logger.warning(
                    "No AGP name in %s–%s matched '%s' above 50%% threshold",
                    start_date,
                    end_date,
                    user_name,
                )
        else:
            # Non-admin or admin generating their own bill - use user-case-mappings
            mappings_ref = db.collection("user-case-mappings")
            query = mappings_ref.where("user_id", "==", user_id)
            mappings = query.stream()

            for mapping_doc in mappings:
                mapping_data = mapping_doc.to_dict()
                case_id = mapping_data.get("case_id")

                # Get case details from daily-boards
                case_ref = db.collection("daily-boards").document(case_id)
                case_doc = case_ref.get()

                if case_doc.exists:
                    case_data = case_doc.to_dict()
                    board_date_raw = case_data.get("board_date")

                    if board_date_raw:
                        try:
                            # Handle both Firestore Timestamp and string formats
                            if isinstance(board_date_raw, str):
                                board_date = datetime.strptime(
                                    board_date_raw, "%Y-%m-%d"
                                )
                                board_date_str = board_date_raw
                            else:
                                # Firestore DatetimeWithNanoseconds object - convert to naive datetime
                                board_date = board_date_raw.replace(tzinfo=None)
                                board_date_str = board_date.strftime("%Y-%m-%d")

                            # Check if case falls within date range
                            if (
                                start_dt <= board_date <= end_dt
                                and case_id not in case_ids
                            ):
                                case_ids.add(case_id)

                                # Determine fee and result based on order analysis
                                fee_info = calculate_case_fee(
                                    case_data, board_date=board_date_str
                                )

                                # Extract parties information
                                parties = extract_parties_info(case_data)

                                bill_entry = {
                                    "id": case_id,
                                    "date": board_date_str,
                                    "case_detail": f"{case_data.get('case_type')}/{case_data.get('case_no')}/{case_data.get('case_year')}",
                                    "case_type": case_data.get("case_type", ""),
                                    "case_no": case_data.get("case_no", ""),
                                    "case_year": case_data.get("case_year", ""),
                                    "parties_name": parties,
                                    "results": fee_info["result"],
                                    "fees_rs": fee_info["fee"],
                                    "order_link": fee_info.get("order_link"),
                                    "order_category": fee_info.get("order_category"),
                                    "confidence_score": mapping_data.get(
                                        "confidence_score", 0.0
                                    ),
                                    "order_category_confidence": fee_info.get(
                                        "order_category_confidence"
                                    ),
                                    "match_source": mapping_data.get("match_source"),
                                    "agp_name": case_data.get("agp_name", "N/A"),
                                    "editable": True,
                                }
                                bill_entries.append(bill_entry)

                        except ValueError:
                            logger.warning(
                                f"Invalid date format for case {case_id}: {board_date_str}"
                            )
                            continue

        # Sort by date
        bill_entries.sort(key=lambda x: x["date"])

        # Add debug information
        response_data = {
            "user_id": user_id,
            "user_name": user_name if user_name else "self",
            "date_range": {"start": start_date, "end": end_date},
            "total_entries": len(bill_entries),
            "total_fees": sum(entry["fees_rs"] for entry in bill_entries),
            "bill_entries": bill_entries,
        }

        # Add matching debug info for admin fuzzy matching
        if user_name and "matched_agp" in locals() and matched_agp is not None:
            response_data["debug_info"] = {
                "requested_name": user_name,
                "matched_agp_name": matched_agp,
                "match_confidence": round(confidence, 3),
                "total_cases_for_agp": len(cases_by_agp.get(matched_agp, [])),
                "cases_in_date_range": len(bill_entries),
            }

        return JSONResponse(content=response_data)

    except Exception as e:
        logger.error(f"Error generating bill data: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to generate bill data: {str(e)}"},
        )


def generate_bill_number_safe(db, user_id: str, year: int) -> tuple:
    """Generate unique bill number with year and sequence for a user (transaction-safe)"""
    try:
        # Use a counter document per user per year to ensure atomic increments
        counter_id = f"{user_id}_{year}"
        counter_ref = db.collection("bill-counters").document(counter_id)

        # Use Firestore transaction to atomically increment counter
        @firestore.transactional
        def increment_counter(transaction):
            counter_doc = counter_ref.get(transaction=transaction)

            if counter_doc.exists:
                current_seq = counter_doc.to_dict().get("sequence", 0)
                next_sequence = current_seq + 1
            else:
                # First bill for this user in this year
                next_sequence = 1

            # Update counter atomically
            transaction.set(
                counter_ref,
                {
                    "user_id": user_id,
                    "year": year,
                    "sequence": next_sequence,
                    "last_updated": firestore.SERVER_TIMESTAMP,
                },
            )

            return next_sequence

        # Execute transaction
        transaction = db.transaction()
        next_sequence = increment_counter(transaction)

        # Format: BILL/YEAR/SEQUENCE (e.g., BILL/2025/001)
        bill_number = f"BILL/{year}/{next_sequence:03d}"

        logger.info(f"✨ Generated bill number: {bill_number} for user {user_id}")
        return bill_number, next_sequence

    except Exception as e:
        logger.error(f"Error generating bill number: {e}")
        # Fallback to timestamp-based number (should rarely happen)
        import time

        timestamp_seq = int(time.time()) % 10000
        return f"BILL/{year}/{timestamp_seq:04d}", timestamp_seq


def generate_month_description(start_date: str, end_date: str) -> str:
    """Generate month description from date range (e.g., 'January 2025 - March 2025')"""
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")

        # Same month
        if start_dt.month == end_dt.month and start_dt.year == end_dt.year:
            return start_dt.strftime("%B %Y")

        # Different months, same year
        if start_dt.year == end_dt.year:
            return f"{start_dt.strftime('%B')} - {end_dt.strftime('%B %Y')}"

        # Different years
        return f"{start_dt.strftime('%B %Y')} - {end_dt.strftime('%B %Y')}"
    except Exception as e:
        logger.error(f"Error generating month description: {e}")
        return f"{start_date} to {end_date}"


def _previously_billed_keys(db, user_id: str) -> set:
    """(case_ref, date) pairs already present in this user's saved bills --
    the one QA check that needs I/O, kept outside qa_check_bill so that
    function stays pure and fully unit-testable."""
    keys = set()
    for bill_doc in (
        db.collection("user-bills").where("user_id", "==", user_id).stream()
    ):
        for entry in (bill_doc.to_dict() or {}).get("entries") or []:
            case_ref = entry.get("case_detail") or entry.get("case_ref")
            date = entry.get("date")
            if case_ref and date:
                keys.add((case_ref, date))
    return keys


def _run_bill_qa(db, user_id: str, bill_entries: List[Dict]) -> Dict:
    """Shared by POST /bills/qa-check (advisory preview) and POST /bills/save
    (enforced gate) so the two can never disagree about what counts as a
    problem."""
    from AutoOrderManager import AutoOrderManager
    from bill_qa import qa_check_bill

    return qa_check_bill(
        bill_entries,
        previously_billed_keys=_previously_billed_keys(db, user_id),
        review_confidence_threshold=AutoOrderManager.REVIEW_CONFIDENCE_THRESHOLD,
    )


def _recompute_entry_fees(bill_entries: List[Dict]) -> Tuple[List[Dict], int]:
    """Recompute every entry's fee from the canonical fee schedule instead of
    trusting the value the browser posted.

    /bills/save used to persist client-supplied ``fees_rs`` verbatim and sum
    them for ``total_fees`` -- so a stale tab, an edited request, or a UI bug
    could put a number on a government-bound bill that no server-side rule
    ever agreed to. Entries whose ``results`` isn't in the schedule (custom or
    unrecognised outcomes) keep their submitted fee: there is no canonical
    value to substitute, and qa_check_bill deliberately doesn't flag them
    either.

    Returns (entries, total_fees) with entries copied, never mutated in place.
    """
    from bill_qa import FEE_SCHEDULE

    recomputed: List[Dict] = []
    total = 0
    for entry in bill_entries:
        entry = dict(entry)
        expected = FEE_SCHEDULE.get(entry.get("results"))
        if expected is not None:
            entry["fees_rs"] = expected
        try:
            total += int(entry.get("fees_rs") or 0)
        except (TypeError, ValueError):
            pass
        recomputed.append(entry)
    return recomputed, total


@app.post("/bills/qa-check", tags=["Bill Generation"])
async def bill_qa_check(request: Request, current_user=Depends(get_current_user)):
    """Roadmap #5: a second pair of eyes on a bill before it's saved --
    the bill export is the one artifact that leaves the building and goes
    to a government body, and everything upstream (a classification
    error, a fuzzy-match miss, a manual fee edit) can ride silently into
    it. Advisory only, never blocks: returns which entries triggered
    which check and why, the caller decides what to do about it.

    POST body: {"bill_entries": [...]} (the same shape /bills/save takes)
    """
    try:
        db = firestore.client()
        user_id = current_user.get("uid")
        body = await request.json()
        bill_entries = body.get("bill_entries", [])

        report = _run_bill_qa(db, user_id, bill_entries)
        return JSONResponse(content=report)
    except Exception as e:
        logger.error(f"Error running bill QA check: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to check bill: {str(e)}"}
        )


@app.post("/bills/save", tags=["Bill Generation"])
async def save_bill_entries(request: Request, current_user=Depends(get_current_user)):
    """Save bill entries with unique bill number and year for logged-in user"""
    try:
        db = firestore.client()
        user_id = current_user.get("uid")
        body = await request.json()

        bill_entries = body.get("bill_entries", [])
        bill_metadata = body.get("metadata", {})

        # This endpoint used to persist whatever the browser posted, summing
        # client-supplied fees_rs into total_fees with no server-side check at
        # all -- on the one artifact that leaves the building for a government
        # body. The QA logic to catch fee/category mismatches and duplicate
        # billing already existed (bill_qa.py) but was only ever run
        # advisorily, by a separate endpoint the UI could simply not call.
        # Now it gates the write: blocking issues are refused unless the
        # caller explicitly overrides, and the override is recorded on the
        # bill so the decision is auditable rather than invisible.
        qa_report = _run_bill_qa(db, user_id, bill_entries)
        override_qa = bool(body.get("override_qa"))
        if not qa_report["ok"] and not override_qa:
            return JSONResponse(
                status_code=400,
                content={
                    "error": "Bill did not pass validation",
                    "qa_report": qa_report,
                    "hint": (
                        "Fix the flagged entries, or resubmit with "
                        "override_qa=true to save anyway (the override is "
                        "recorded on the bill)."
                    ),
                },
            )

        # Fees come from the canonical schedule, not the request body.
        bill_entries, recomputed_total = _recompute_entry_fees(bill_entries)

        # Get date range from metadata (frontend sends startDate/endDate)
        date_range = bill_metadata.get("date_range", {})
        start_date = date_range.get("startDate", date_range.get("start", ""))
        end_date = date_range.get("endDate", date_range.get("end", ""))

        # Determine bill year from date range (use end date year)
        current_year = datetime.now().year
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                current_year = end_dt.year
            except ValueError:
                pass

        # Generate unique bill number (transaction-safe to prevent duplicates)
        bill_number, bill_sequence = generate_bill_number_safe(
            db, user_id, current_year
        )

        # Generate month description
        month_description = (
            generate_month_description(start_date, end_date)
            if start_date and end_date
            else ""
        )

        # Create a bill document with bill number and year
        bill_data = {
            "user_id": user_id,
            "bill_number": bill_number,
            "bill_year": current_year,
            "bill_sequence": bill_sequence,
            "month_description": month_description,
            "created_at": firestore.SERVER_TIMESTAMP,
            "updated_at": firestore.SERVER_TIMESTAMP,
            "metadata": bill_metadata,
            "entries": bill_entries,
            "total_entries": len(bill_entries),
            "total_fees": recomputed_total,
            "qa_summary": qa_report["summary_lines"],
        }
        if override_qa and not qa_report["ok"]:
            bill_data["qa_override"] = {
                "overridden_by": user_id,
                "at": datetime.now().isoformat(),
                "report": qa_report,
            }

        # Save to user-bills collection
        bill_ref = db.collection("user-bills").document()
        bill_ref.set(bill_data)
        bill_id = bill_ref.id

        logger.info(
            f"✅ Bill saved: {bill_number} for user {user_id}, {month_description}"
        )

        return JSONResponse(
            content={
                "success": True,
                "bill_id": bill_id,
                "bill_number": bill_number,
                "bill_year": current_year,
                "month_description": month_description,
                "total_entries": len(bill_entries),
                "total_fees": bill_data["total_fees"],
            }
        )

    except Exception as e:
        logger.error(f"Error saving bill entries: {e}")
        import traceback

        traceback.print_exc()
        return JSONResponse(
            status_code=500, content={"error": f"Failed to save bill entries: {str(e)}"}
        )


@app.get("/bills/my-bills", tags=["Bill Generation"])
async def get_my_bills(
    limit: int = Query(20, description="Maximum number of bills to return"),
    user_id_filter: Optional[str] = Query(
        None, description="Filter by user ID (admin only)"
    ),
    current_user=Depends(get_current_user),
):
    """Get saved bills - logged-in user's bills or all bills (admin only)"""
    try:
        db = firestore.client()
        user_id = current_user.get("uid")
        is_admin = get_user_manager().is_admin(user_id)

        bills_ref = db.collection("user-bills")

        # Admin can view all bills or filter by specific user
        if is_admin and user_id_filter:
            # Admin viewing specific user's bills
            query = (
                bills_ref.where("user_id", "==", user_id_filter)
                .order_by("created_at", direction=firestore.Query.DESCENDING)
                .limit(limit)
            )
            target_user_id = user_id_filter
        elif is_admin and not user_id_filter:
            # Admin viewing all bills
            query = bills_ref.order_by(
                "created_at", direction=firestore.Query.DESCENDING
            ).limit(limit)
            target_user_id = "all"
        else:
            # Regular user - only their own bills
            query = (
                bills_ref.where("user_id", "==", user_id)
                .order_by("created_at", direction=firestore.Query.DESCENDING)
                .limit(limit)
            )
            target_user_id = user_id

        bills = query.stream()

        bills_list = []
        for bill_doc in bills:
            bill_data = bill_doc.to_dict()
            bill_data["id"] = bill_doc.id

            # Convert timestamps
            if "created_at" in bill_data and bill_data["created_at"]:
                bill_data["created_at"] = bill_data["created_at"].isoformat()
            if "updated_at" in bill_data and bill_data["updated_at"]:
                bill_data["updated_at"] = bill_data["updated_at"].isoformat()

            bills_list.append(bill_data)

        return JSONResponse(
            content={
                "user_id": target_user_id,
                "is_admin": is_admin,
                "bills": bills_list,
                "total_bills": len(bills_list),
            }
        )

    except Exception as e:
        logger.error(f"Error getting user bills: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to get user bills: {str(e)}"}
        )


def calculate_case_fee(case_data: Dict, board_date: Optional[str] = None) -> Dict:
    """Calculate fee and result based on order analysis.

    Returns a dict with keys: result, fee, order_link, order_category.
    order_link and order_category are populated when the order has been
    analysed; both are None when the case has no linked order.

    When *board_date* (YYYY-MM-DD) is provided, the order whose board_date
    or order_date matches is used for fee calculation and the returned
    order_link.  This prevents a later hearing's order from being shown
    against an earlier bill entry for the same case.
    """
    case_ref = f"{case_data.get('case_type', '')}/{case_data.get('case_no', '')}/{case_data.get('case_year', '')}"
    try:
        case_details = (
            get_auto_order_manager().case_store.get_case_details(case_ref) or {}
        )

        orders = case_details.get("orders") or []

        # When board_date is given, only use an order that specifically belongs
        # to that hearing date.  Do NOT fall back to a different date's order —
        # the bill entry for May 6 must not show April 30's order.
        # Without board_date (legacy callers) fall back to the last analysed order.
        target_order: Dict = {}
        if board_date and orders:
            for o in reversed(orders):
                if not isinstance(o, dict):
                    continue
                if (
                    o.get("board_date") == board_date
                    or o.get("order_date") == board_date
                ):
                    target_order = o
                    break
            # No date-specific match → show no link for this hearing
        else:
            # No board_date context (legacy callers) — use last analysed order
            for o in reversed(orders):
                if isinstance(o, dict) and o.get("order_status") == "analysed":
                    target_order = o
                    break

        # If we still have nothing, the case is not yet analysed
        if not target_order or target_order.get("order_status") != "analysed":
            return {
                "result": "*ADJOURNED*",
                "fee": 1250,
                "order_link": None,
                "order_category": None,
            }

        order_link = target_order.get("order_link") or None
        order_category = (target_order.get("order_category") or "").upper()
        # How sure the classifier was. Surfaced on the bill so a low-confidence
        # categorisation — which sets the fee — is visible before submission.
        category_confidence = target_order.get("order_category_confidence")

        # Fee calculation logic based on order category alone. order_category
        # is the single source of truth here -- it must never be widened back
        # to also match against order text/reason strings: order_analyzer
        # never persists order_text to Firestore (only category + confidence
        # + a few metadata fields), so a text-matching branch here would
        # either be permanently dead or, the moment text does get persisted,
        # silently start letting stray prose in the order body pick the fee
        # instead of the classifier's actual category.
        if "DISPOSED" in order_category:
            return {
                "result": "WP DISPOSED OF",
                "fee": 2500,
                "order_link": order_link,
                "order_category": order_category,
                "order_category_confidence": category_confidence,
            }

        # Check for heard & adjourned (middle fee)
        elif "HEARD" in order_category and "ADJOURNED" in order_category:
            return {
                "result": "HEARD & ADJN.",
                "fee": 1875,
                "order_link": order_link,
                "order_category": order_category,
                "order_category_confidence": category_confidence,
            }

        # Check for simple adjournment (lowest fee)
        elif "ADJOURNED" in order_category:
            return {
                "result": "ADJOURNED",
                "fee": 1250,
                "order_link": order_link,
                "order_category": order_category,
                "order_category_confidence": category_confidence,
            }

        # Default
        else:
            if "HEARD" in order_category:
                return {
                    "result": "HEARD & ADJN.",
                    "fee": 1875,
                    "order_link": order_link,
                    "order_category": order_category,
                    "order_category_confidence": category_confidence,
                }
            else:
                return {
                    "result": "*ADJOURNED*",
                    "fee": 1250,
                    "order_link": None,
                    "order_category": None,
                }

    except Exception as e:
        # Bills the same floor rate as a legitimate "no order on file" (there
        # is no better fee to guess) but order_category distinguishes the two
        # causes -- None means genuinely no order was ever matched; this
        # sentinel means the calculation itself crashed and needs
        # investigating, not just review. Previously both cases returned an
        # identical dict, so a code fault was indistinguishable from a normal
        # gap on the bill and in bill_qa's flagged-entries output.
        logger.error(
            f"Error calculating case fee for case_ref={case_ref}: {e}", exc_info=True
        )
        return {
            "result": "*ADJOURNED*",
            "fee": 1250,
            "order_link": None,
            "order_category": "CALCULATION_ERROR",
        }


def extract_parties_info(case_data: Dict) -> str:
    """Extract parties information from case data (format: Petitioner vs Respondent)"""
    try:
        case_ref = f"{case_data.get('case_type', '')}/{case_data.get('case_no', '')}/{case_data.get('case_year', '')}"
        case_details = (
            get_auto_order_manager().case_store.get_case_details(case_ref) or {}
        )
        petitioner = str(case_details.get("petitioner") or "").strip()
        respondent = str(case_details.get("respondent") or "").strip()
        if petitioner and respondent:
            return f"{petitioner} Versus {respondent}"

        # Fallback to case reference
        return f"Matter in {case_ref}"

    except Exception as e:
        logger.error(f"Error extracting parties info: {e}")
        return "Parties information not available"


# Excel column width units -> approximate wrapped-line count.
#
# Row heights below are computed, not copied from the reference bill: the
# reference's own per-row heights were baked in for one specific bill's exact
# text lengths (long AGP name, specific party names) and would misalign for
# any other bill. Instead each row's height is derived from its own text,
# calibrated against the reference file's Parties Name column, which is the
# one column with real length variation (case refs, results, and the header
# rows are short and effectively fixed). Three known (char-count, line-count)
# pairs from that column ("...YADAV Versus...ORS." = 74 chars/3 lines,
# "...SADIQUE Versus...ANR" = 91 chars/4 lines, "...PATIL...Versus...DEPT." =
# 134 chars/6 lines) imply roughly 22-25 characters per line at that column's
# 58.57-unit width; CHARS_PER_UNIT below uses the low end of that range so the
# estimate rounds lines UP rather than risking clipped text — a bit of extra
# whitespace is a fine trade for never cutting off a party's name.
_BILL_LINE_HEIGHT_PT = 23.25  # the reference bill's own single-line row height
_BILL_CHARS_PER_UNIT = 0.38
_BILL_CELL_PADDING_UNITS = 1.0


def _bill_lines_needed(text: str, width_units: float) -> int:
    usable = max(width_units - _BILL_CELL_PADDING_UNITS, 1.0)
    chars_per_line = max(1, int(usable * _BILL_CHARS_PER_UNIT))
    total = 0
    for segment in str(text or "").split("\n"):
        total += max(1, -(-len(segment) // chars_per_line))  # ceil division
    return max(total, 1)


def _bill_row_height(cells_and_widths) -> float:
    """cells_and_widths: iterable of (text, width_units) sharing a row.

    Height is driven by whichever cell needs the most lines, since all cells
    in an Excel row share one height.
    """
    lines = [_bill_lines_needed(text, width) for text, width in cells_and_widths]
    return max(lines or [1]) * _BILL_LINE_HEIGHT_PT


def build_bill_workbook(entries, agp_name: str, bill_number: str, period_str: str):
    """Build the AGP fee bill workbook.

    Layout matches the office's reference bill format: 6 columns (SR. NO. /
    DATE / CASE DETAILS / RESULTS / PARTIES NAME / FEES (RS.)), Times New
    Roman 18pt throughout with bold reserved for the column-header row, a
    thin border around every cell including the merged header block, and a
    footer with a live GROSS AMOUNT total plus a signature block.

    Deliberately does NOT compute "Fees Earn Due To Ceiling", "TOTAL TDS 10%"
    or "NET AMOUNT" — those depend on the AGP's annual fee ceiling and
    cumulative prior claims, which this system does not track, so they are
    left blank for manual completion, exactly as the reference bill does.

    Pure function: no Firestore, no auth, no I/O. Returns an openpyxl
    Workbook: callers save/stream it however they need.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Bill"

    body_font = Font(name="Times New Roman", size=18)
    bold_font = Font(name="Times New Roman", size=18, bold=True)
    thin_side = Side(style="thin")
    border_thin = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)

    column_widths = {"A": 10, "B": 20.43, "C": 28.86, "D": 29.0, "E": 58.57, "F": 15.0}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    full_width = sum(column_widths.values())

    def box_border(cell_range: str):
        """Thin border on every cell in a merged range: top/bottom on all of
        them, left/right only on the range's outer edge — otherwise a merged
        cell only carries a border on its anchor cell and the box looks
        broken. Matches how the reference bill borders its header rows."""
        rows = ws[cell_range]
        if not isinstance(rows[0], tuple):
            rows = (rows,)
        n_cols = len(rows[0])
        for row in rows:
            for c_idx, cell in enumerate(row):
                cell.border = Border(
                    left=thin_side if c_idx == 0 else None,
                    right=thin_side if c_idx == n_cols - 1 else None,
                    top=thin_side,
                    bottom=thin_side,
                )

    def merged_row(row_num, text, col_range, align, bold=False):
        ws.merge_cells(f"A{row_num}:F{row_num}" if col_range is None else col_range)
        first_cell_ref = (col_range or f"A{row_num}:F{row_num}").split(":")[0]
        cell = ws[first_cell_ref]
        cell.value = text
        cell.font = bold_font if bold else body_font
        cell.alignment = align
        box_border(col_range or f"A{row_num}:F{row_num}")
        return cell

    current_row = 1

    # --- Header block --------------------------------------------------
    title_text = f"STATEMENT OF PROFESSIONAL FEES BILL OF AGP {agp_name.upper()}"
    merged_row(current_row, title_text, f"A{current_row}:F{current_row}", center_align)
    ws.row_dimensions[current_row].height = _bill_row_height([(title_text, full_width)])
    current_row += 1

    subtitle_text = (
        "A.S.(WRIT CELL),HIGH COURT, MUMBAI FOR CONDUCTING WRIT MATTERS ETC."
    )
    merged_row(
        current_row, subtitle_text, f"A{current_row}:F{current_row}", center_align
    )
    ws.row_dimensions[current_row].height = _bill_row_height(
        [(subtitle_text, full_width)]
    )
    current_row += 1

    gr_text = (
        "SANCTIONED VIDE:/ GOVERNMENT OF MAHARASHTRA\n"
        "LAW AND JUDICIARY DEPARTMENT,\n"
        "GOVERNMENT RESOLUTION NO. MEETING/GPH/2023/C.R.29/D/14,\n"
        "DATED/30TH OCTOBER, 2023"
    )
    merged_row(current_row, gr_text, f"A{current_row}:F{current_row}", center_align)
    ws.row_dimensions[current_row].height = _bill_row_height([(gr_text, full_width)])
    current_row += 1

    # Period (left) + bill number (right) on one row. The reference embeds
    # both in a single string, right-shifted with manually counted padding
    # spaces calibrated to that one bill's exact text lengths — it would
    # misalign for any other period or bill number. Two properly aligned
    # cells give the same look and stay correct at any length.
    months_text = f"MONTHS : {period_str}"
    ws.merge_cells(f"A{current_row}:D{current_row}")
    a_cell = ws[f"A{current_row}"]
    a_cell.value = months_text
    a_cell.font = body_font
    a_cell.alignment = left_align
    ws.merge_cells(f"E{current_row}:F{current_row}")
    e_cell = ws[f"E{current_row}"]
    e_cell.value = bill_number
    e_cell.font = body_font
    e_cell.alignment = right_align
    box_border(f"A{current_row}:D{current_row}")
    box_border(f"E{current_row}:F{current_row}")
    width_ad = sum(column_widths[c] for c in "ABCD")
    width_ef = sum(column_widths[c] for c in "EF")
    ws.row_dimensions[current_row].height = max(
        _bill_row_height([(months_text, width_ad)]),
        _bill_row_height([(str(bill_number), width_ef)]),
    )
    current_row += 1

    declaration_text = (
        "DECLARATION : I hereby certify that the below mentioned matters were "
        "allotted to me by the Government Pleader, I personally appeared in "
        "the below mentioned matters. The below mentioned entries/information "
        "given in above columns are true and correct to the best of my "
        "knowledge and belief. I further certify that nothing is suppressed "
        "by me. Also, the fees which is claimed in bill no. "
        f"{bill_number} has not been claimed by me earlier."
    )
    merged_row(
        current_row, declaration_text, f"A{current_row}:F{current_row}", left_align
    )
    ws.row_dimensions[current_row].height = _bill_row_height(
        [(declaration_text, full_width)]
    )
    current_row += 1

    # Column headers — 6 columns. CASE TYPE/NO/YEAR collapse into one CASE
    # DETAILS column (e.g. "WP/10000/2017"), matching the reference; that
    # value is already available verbatim as entry["case_detail"].
    headers = [
        "SR. NO.",
        "DATE",
        "CASE DETAILS",
        "RESULTS",
        "PARTIES NAME",
        "FEES (RS.)",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border_thin
    # Deterministic content (always these 6 short words) — a fixed height
    # matching the reference is safe here, no text-length variance to guard.
    ws.row_dimensions[current_row].height = 60.0
    current_row += 1

    # --- Data rows -------------------------------------------------------
    first_data_row = current_row
    for idx, entry in enumerate(entries, 1):
        case_detail = str(entry.get("case_detail") or "").strip()
        if not case_detail:
            # Fallback for older saved bills predating the case_detail field.
            case_detail = "/".join(
                p
                for p in (
                    str(entry.get("case_type") or "").strip(),
                    str(entry.get("case_no") or "").strip(),
                    str(entry.get("case_year") or "").strip(),
                )
                if p
            )

        date_str = entry.get("date", "")
        try:
            date_value = datetime.strptime(date_str, "%Y-%m-%d")
        except (ValueError, TypeError):
            date_value = str(date_str) if date_str else ""

        results = str(entry.get("results", "") or "")
        parties_name = str(entry.get("parties_name", "") or "")

        fees_rs = entry.get("fees_rs", 0)
        try:
            fees_rs = float(fees_rs) if fees_rs is not None else 0.0
        except (ValueError, TypeError):
            fees_rs = 0.0

        row_values = [idx, date_value, case_detail, results, parties_name, fees_rs]
        for col_num, value in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_num, value=value)
            cell.font = body_font
            cell.border = border_thin
            cell.alignment = center_align
        if isinstance(date_value, datetime):
            ws.cell(row=current_row, column=2).number_format = "dd/mmm/yyyy"

        # Only CASE DETAILS and PARTIES NAME have meaningful length variance
        # (SR. NO./DATE/RESULTS/FEES are always short, fixed-vocabulary
        # values that never wrap in their columns) — estimating wrap on those
        # too would systematically over-count lines for the common case
        # ("HEARD & ADJN." alone looks like it needs 2 lines at column D's
        # width by character count, when in practice it always fits on 1).
        ws.row_dimensions[current_row].height = _bill_row_height(
            [
                (case_detail, column_widths["C"]),
                (parties_name, column_widths["E"]),
            ]
        )
        current_row += 1

    last_data_row = current_row - 1

    # --- Footer: totals + signature ---------------------------------------
    # Fees Earn Due To Ceiling / TOTAL TDS 10% / NET AMOUNT are left blank —
    # see the docstring: this system doesn't track the inputs needed to
    # compute them, so guessing would be worse than leaving them for the AGP
    # or their accountant to fill in by hand, which is what the reference
    # bill does too.
    def footer_label(row, text):
        cell = ws[f"D{row}"]
        cell.value = text
        cell.font = body_font
        cell.alignment = center_align
        cell.border = border_thin
        ws.merge_cells(f"E{row}:F{row}")
        box_border(f"E{row}:F{row}")
        ws.row_dimensions[row].height = _BILL_LINE_HEIGHT_PT

    ws[f"D{current_row}"] = "GROSS AMOUNT"
    ws[f"D{current_row}"].font = body_font
    ws[f"D{current_row}"].alignment = center_align
    ws[f"D{current_row}"].border = border_thin
    ws.merge_cells(f"E{current_row}:F{current_row}")
    gross_cell = ws[f"E{current_row}"]
    gross_cell.value = f"=SUM(F{first_data_row}:F{last_data_row})"
    gross_cell.font = bold_font
    gross_cell.alignment = center_align
    box_border(f"E{current_row}:F{current_row}")
    ws.row_dimensions[current_row].height = _BILL_LINE_HEIGHT_PT
    current_row += 1

    footer_label(current_row, "Fees Earn Due To Ceiling")
    current_row += 1

    tds_note = (
        "As per G. R. F. D. No. आयकर १००७/प्र. क्र. १०५/कोषा प्र. ५ "
        "मंत्रालय मुंबई Dated \n22-02-2008"
    )
    ws.merge_cells(f"A{current_row}:C{current_row + 1}")
    note_cell = ws[f"A{current_row}"]
    note_cell.value = tds_note
    note_cell.font = body_font
    note_cell.alignment = center_align
    box_border(f"A{current_row}:C{current_row + 1}")
    footer_label(current_row, "TOTAL TDS 10%")
    current_row += 1

    footer_label(current_row, "NET AMOUNT")
    current_row += 2  # spacer

    ws[f"A{current_row}"] = "VERIFIED & CORRECT"
    ws[f"A{current_row}"].font = body_font
    current_row += 2  # spacer

    ws[f"A{current_row}"] = "PLACE:"
    ws[f"A{current_row}"].font = body_font
    ws[f"B{current_row}"] = "Mumbai"
    ws[f"B{current_row}"].font = body_font
    ws[f"E{current_row}"] = f"({agp_name.strip().upper()})"
    ws[f"E{current_row}"].font = body_font
    current_row += 1

    ws[f"A{current_row}"] = "DATE:"
    ws[f"A{current_row}"].font = body_font
    date_cell = ws[f"B{current_row}"]
    date_cell.value = datetime.now()
    date_cell.number_format = "dd/mmm/yyyy"
    date_cell.font = body_font
    date_cell.alignment = left_align
    # Correcting the reference bill's "Assisstant Governtment Pleader" typo —
    # this session already standardized AGP terminology across the app.
    ws[f"E{current_row}"] = "Assistant Government Pleader"
    ws[f"E{current_row}"].font = body_font

    return wb


@app.get("/bills/export/excel", tags=["Bill Generation"])
async def export_bill_excel(
    bill_id: str = Query(None, description="Bill ID to export"),
    start_date: str = Query(None, description="Start date for generating fresh export"),
    end_date: str = Query(None, description="End date for generating fresh export"),
    user_name: Optional[str] = Query(
        None, description="User name for bill header (admin only)"
    ),
    current_user=Depends(get_current_user),
):
    """Export bill data as Excel format matching AGP bill specification"""
    try:
        import io
        from datetime import datetime

        db = firestore.client()
        user_id = current_user.get("uid")
        is_admin = get_user_manager().is_admin(user_id)

        # Get bill data - either from saved bill or generate fresh
        agp_name = "ASSISTANT GOVERNMENT PLEADER"
        bill_number = f"BILL/{datetime.now().strftime('%m')}/{datetime.now().year}"

        if bill_id:
            # Export saved bill
            bill_ref = db.collection("user-bills").document(bill_id)
            bill_doc = bill_ref.get()

            if not bill_doc.exists:
                return JSONResponse(
                    status_code=404, content={"error": "Bill not found"}
                )

            bill_data = bill_doc.to_dict()
            if bill_data.get("user_id") != user_id and not is_admin:
                return JSONResponse(status_code=403, content={"error": "Access denied"})

            entries = bill_data.get("entries", [])
            metadata = bill_data.get("metadata", {})
            agp_name = entries[0].get("agp_name", agp_name) if entries else agp_name
            bill_number = bill_data.get("bill_number", bill_number)
            filename = f"bill_{bill_id}.xlsx"

        elif start_date and end_date:
            # Generate fresh export
            response = await generate_bill_data(
                start_date, end_date, user_name, current_user
            )
            if response.status_code != 200:
                return response

            response_data = json.loads(response.body.decode())
            entries = response_data.get("bill_entries", [])
            metadata = {"date_range": {"start": start_date, "end": end_date}}

            # Get AGP name from entries or debug info
            if entries and entries[0].get("agp_name"):
                agp_name = entries[0].get("agp_name")
            elif "debug_info" in response_data and response_data["debug_info"].get(
                "matched_agp_name"
            ):
                agp_name = response_data["debug_info"]["matched_agp_name"]

            filename = f"AGP_Bill_{start_date}_to_{end_date}.xlsx"

        else:
            return JSONResponse(
                status_code=400,
                content={
                    "error": "Either bill_id or both start_date and end_date are required"
                },
            )

        if not entries:
            return JSONResponse(
                status_code=404, content={"error": "No bill entries found"}
            )

        # Parse dates for header
        date_range = metadata.get("date_range", {})
        start_dt = datetime.strptime(date_range.get("start", start_date), "%Y-%m-%d")
        end_dt = datetime.strptime(date_range.get("end", end_date), "%Y-%m-%d")
        period_str = (
            f"{start_dt.strftime('%B %Y').upper()} - {end_dt.strftime('%B %Y').upper()}"
        )

        wb = build_bill_workbook(entries, agp_name, bill_number, period_str)

        # Save to BytesIO and extract raw bytes for response
        output = io.BytesIO()
        wb.save(output)

        # Return as downloadable file with proper headers
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )

    except Exception as e:
        logger.error(f"Error exporting bill to Excel: {e}")
        import traceback

        traceback.print_exc()
        return JSONResponse(
            status_code=500, content={"error": f"Failed to export bill: {str(e)}"}
        )


@app.get("/bills/{bill_id}", tags=["Bill Generation"])
async def get_bill_details(bill_id: str, current_user=Depends(get_current_user)):
    """Get details of a specific saved bill - admin can view any bill"""
    try:
        db = firestore.client()
        user_id = current_user.get("uid")
        is_admin = get_user_manager().is_admin(user_id)

        bill_ref = db.collection("user-bills").document(bill_id)
        bill_doc = bill_ref.get()

        if not bill_doc.exists:
            return JSONResponse(status_code=404, content={"error": "Bill not found"})

        bill_data = bill_doc.to_dict()

        # Check ownership - admin can view any bill, regular user only their own
        if not is_admin and bill_data.get("user_id") != user_id:
            return JSONResponse(status_code=403, content={"error": "Access denied"})

        bill_data["id"] = bill_doc.id

        # Convert timestamps
        if "created_at" in bill_data and bill_data["created_at"]:
            bill_data["created_at"] = bill_data["created_at"].isoformat()
        if "updated_at" in bill_data and bill_data["updated_at"]:
            bill_data["updated_at"] = bill_data["updated_at"].isoformat()

        return JSONResponse(content=bill_data)

    except Exception as e:
        logger.error(f"Error getting bill details: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to get bill details: {str(e)}"}
        )


@app.delete("/bills/{bill_id}", tags=["Bill Generation"])
async def delete_bill(bill_id: str, current_user=Depends(get_current_user)):
    """Delete a saved bill - admin can delete any bill"""
    try:
        db = firestore.client()
        user_id = current_user.get("uid")
        is_admin = get_user_manager().is_admin(user_id)

        bill_ref = db.collection("user-bills").document(bill_id)
        bill_doc = bill_ref.get()

        if not bill_doc.exists:
            return JSONResponse(status_code=404, content={"error": "Bill not found"})

        bill_data = bill_doc.to_dict()

        # Check ownership - admin can delete any bill, regular user only their own
        if not is_admin and bill_data.get("user_id") != user_id:
            return JSONResponse(status_code=403, content={"error": "Access denied"})

        # Delete the bill
        bill_ref.delete()

        return JSONResponse(
            content={"success": True, "message": "Bill deleted successfully"}
        )

    except Exception as e:
        logger.error(f"Error deleting bill: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Failed to delete bill: {str(e)}"}
        )


@app.get("/scraper/status", tags=["Case Orders"])
async def scraper_status(current_user: dict = Depends(require_admin_active)):
    """Return current Bombay High Court scraper configuration and provider status."""
    _ = current_user  # Explicitly keep dependency for admin-only access.
    scraper = get_court_scraper()
    return JSONResponse(content=scraper.get_scraper_config())


@app.post("/scraper/configure", tags=["Case Orders"])
async def scraper_configure(
    provider: Optional[str] = None,
    current_user: dict = Depends(require_admin_active),
):
    """Update scraper provider settings at runtime without redeploying the backend."""
    _ = current_user  # Explicitly keep dependency for admin-only access.
    scraper = get_court_scraper()
    try:
        updated = scraper.configure_scraper(
            provider=provider,
        )
        return JSONResponse(
            content={
                "message": "Scraper configuration updated",
                **updated,
            }
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/scraper/test-case", tags=["Case Orders"])
async def scraper_test_case(
    request: Request,
    current_user: dict = Depends(require_admin_active),
):
    """Run a live scrape for a case reference and return full diagnostics.

    POST body: {"case_ref": "WP/3434/2026", "date": "2026-05-30"}
    Returns the raw provider result including court_orders, case_details,
    the provider sequence that ran, and each attempt's duration and status.
    Useful for diagnosing download failures without triggering a full pipeline run.
    """
    _ = current_user
    import time as _time

    body = await request.json()
    case_ref = str(body.get("case_ref") or "").strip().upper()
    date = body.get("date")

    if not case_ref:
        return JSONResponse(status_code=400, content={"error": "case_ref is required"})

    scraper = get_court_scraper()
    started = _time.time()

    loop = asyncio.get_event_loop()
    try:
        diagnostics = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                lambda: scraper._fetch_with_provider(
                    case_ref=case_ref,
                    date=date,
                    bench="mumbai",
                    include_diagnostics=True,
                ),
            ),
            timeout=120.0,
        )
    except asyncio.TimeoutError:
        return JSONResponse(
            status_code=504,
            content={
                "error": "Scraper timed out after 120 seconds",
                "case_ref": case_ref,
            },
        )
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"error": str(exc), "case_ref": case_ref},
        )

    elapsed_ms = int((_time.time() - started) * 1000)
    result = diagnostics.get("result") or {}
    court_orders = result.get("court_orders") or []
    return JSONResponse(
        content={
            "case_ref": case_ref,
            "date_filter": date,
            "elapsed_ms": elapsed_ms,
            "provider": diagnostics.get("provider"),
            "provider_sequence": diagnostics.get("provider_sequence"),
            "provider_attempts": diagnostics.get("provider_attempts"),
            "found": bool(result),
            "orders_found": len(court_orders),
            "court_orders": court_orders,
            "case_details": result.get("case_details"),
            "source": result.get("source"),
        }
    )


@app.post("/admin/portal-health-check", tags=["Case Orders"])
async def portal_health_check(
    request: Request,
    current_user: dict = Depends(require_admin_active),
):
    """Roadmap #3: run a canary case through BOTH scraper providers
    (HTTP and Playwright -- independent extraction paths) and diagnose
    whether zero-orders-found looks like a genuine "no orders yet" or a
    portal change nobody would otherwise notice until orders silently
    stop flowing. Safe to run on a schedule (e.g. Cloud Scheduler hitting
    this endpoint) -- read-only against the court portal, same scraper
    calls /scraper/test-case already makes for one provider at a time.

    POST body: {"case_ref": "WP/3434/2026", "date": "2026-05-30",
                "expected_min_orders": 1}
    """
    _ = current_user
    import time as _time

    from portal_health import call_llm_for_diagnosis, diagnose_probe

    body = await request.json()
    case_ref = str(body.get("case_ref") or "").strip().upper()
    date = body.get("date")
    bench = body.get("bench", "mumbai")
    expected_min_orders = int(body.get("expected_min_orders", 1))

    if not case_ref:
        return JSONResponse(status_code=400, content={"error": "case_ref is required"})

    scraper = get_court_scraper()
    started = _time.time()

    loop = asyncio.get_event_loop()
    try:
        provider_matrix = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                lambda: scraper._probe_provider_matrix(case_ref, date, bench),
            ),
            timeout=180.0,
        )
    except asyncio.TimeoutError:
        return JSONResponse(
            status_code=504,
            content={
                "error": "Portal health check timed out after 180 seconds",
                "case_ref": case_ref,
            },
        )
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"error": str(exc), "case_ref": case_ref},
        )

    report = diagnose_probe(
        provider_matrix, case_ref=case_ref, expected_min_orders=expected_min_orders
    )
    report["elapsed_ms"] = int((_time.time() - started) * 1000)

    api_key = os.environ.get("GEMINI_API_KEY")
    if report["likely_drift"] and api_key:
        llm_diagnosis = call_llm_for_diagnosis(report, api_key)
        if llm_diagnosis:
            report["llm_diagnosis"] = llm_diagnosis

    return JSONResponse(content=report)


async def _run_assistant_tool(
    tool_name: str, tool_args: Dict[str, Any], current_user: Dict[str, Any]
) -> Dict[str, Any]:
    """Executes one of assistant.TOOL_DECLARATIONS by calling the SAME
    endpoint function the corresponding REST route uses -- no separate
    query logic to keep in sync, and current_user (hence the asking
    user's uid) is threaded through every call. Every one of these reads
    only; none can save a bill, override a category, or queue a job."""
    if tool_name == "get_queue_status":
        response = await get_queue_status(current_user=current_user)
        return json.loads(response.body)

    if tool_name == "get_bill_preview":
        response = await generate_bill_data(
            start_date=tool_args.get("start_date"),
            end_date=tool_args.get("end_date"),
            user_name=None,
            current_user=current_user,
        )
        data = json.loads(response.body)
        entries = data.get("bill_entries") or []
        outcome_counts: Dict[str, int] = {}
        for entry in entries:
            result = entry.get("results") or "unknown"
            outcome_counts[result] = outcome_counts.get(result, 0) + 1
        # Trimmed for the model: aggregate stats only, not the full entry
        # list -- keeps the tool response small and avoids the model
        # verbatim-dumping a long table back into the chat.
        return {
            "date_range": data.get("date_range"),
            "total_entries": data.get("total_entries"),
            "total_fees": data.get("total_fees"),
            "outcome_counts": outcome_counts,
        }

    if tool_name == "get_my_saved_bills":
        limit = int(tool_args.get("limit") or 5)
        response = await get_my_bills(
            limit=limit, user_id_filter=None, current_user=current_user
        )
        data = json.loads(response.body)
        bills = data.get("bills") or []
        return {
            "total_saved_bills": len(bills),
            "bills": [
                {
                    "bill_number": b.get("bill_number"),
                    "month_description": b.get("month_description"),
                    "total_entries": b.get("total_entries"),
                    "total_fees": b.get("total_fees"),
                }
                for b in bills[:limit]
            ],
        }

    if tool_name == "get_pending_matter_confirmations":
        response = await get_pending_matter_confirmations(current_user=current_user)
        return json.loads(response.body)

    raise ValueError(f"Unknown tool: {tool_name}")


@app.post("/assistant/ask", tags=["Assistant"])
async def assistant_ask(request: Request, current_user=Depends(get_current_user)):
    """Roadmap #8: a natural-language front door to data that already
    exists behind an API call -- "how many of my cases need attention",
    "what would my October bill look like" -- instead of navigating
    Table, filters, then Bills to find the same answer.

    Deliberately last in the roadmap and narrowly scoped, per the
    roadmap's own sequencing note ("a multiplier on a pipeline that's
    already trustworthy, not a fix for one that isn't"): every tool is
    READ-ONLY (see assistant.py) and scoped to the asking user's own
    uid -- this can look things up, never save a bill, override a
    category, or queue a job.

    POST body: {"question": "...", "history": [{"role": "user"|"assistant",
    "text": "..."}]}
    """
    try:
        api_key = os.environ.get("GEMINI_API_KEY")
        if not api_key:
            return JSONResponse(
                status_code=501,
                content={
                    "error": "Assistant is not configured (GEMINI_API_KEY not set)."
                },
            )

        body = await request.json()
        question = str(body.get("question") or "").strip()
        if not question:
            return JSONResponse(
                status_code=400, content={"error": "question is required"}
            )
        # Keep only recent turns -- this is a lightweight front door, not a
        # long-running conversation; also bounds the prompt/cost per call.
        history = (body.get("history") or [])[-6:]

        from assistant import AssistantError, ask

        async def tool_executor(
            tool_name: str, tool_args: Dict[str, Any]
        ) -> Dict[str, Any]:
            return await _run_assistant_tool(tool_name, tool_args, current_user)

        try:
            result = await ask(question, history, tool_executor, api_key)
        except AssistantError as e:
            return JSONResponse(
                status_code=502, content={"error": f"Assistant failed: {e}"}
            )

        return JSONResponse(content=result)
    except Exception as e:
        logger.error(f"Error in assistant_ask: {e}")
        return JSONResponse(
            status_code=500, content={"error": f"Assistant failed: {str(e)}"}
        )


# Cloud Run entry point - uvicorn will run the app directly
# For Cloud Functions deployment, use a separate functions_entry.py file
