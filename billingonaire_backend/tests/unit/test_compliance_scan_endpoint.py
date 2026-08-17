"""Tests for POST /compliance/scan (main.compliance_scan)."""

import io
import sys
import types
from unittest.mock import MagicMock, Mock, patch

import pytest
from fastapi import HTTPException

# Stub spaCy before any main import so the import-time crash is avoided,
# same pattern as test_order_pdf_endpoints.py.
if "spacy" not in sys.modules:
    _spacy_stub = types.ModuleType("spacy")
    _spacy_matcher_stub = types.ModuleType("spacy.matcher")

    class _Matcher:  # pragma: no cover
        pass

    _spacy_matcher_stub.Matcher = _Matcher
    _spacy_stub.matcher = _spacy_matcher_stub
    sys.modules["spacy"] = _spacy_stub
    sys.modules["spacy.matcher"] = _spacy_matcher_stub

import main


def _row(
    case_ref="WP/1/2026",
    board_date="2026-07-08",
    order_date="2026-07-08",
    order_category="HEARD_AND_ADJOURNED",
    order_link="https://storage.example/order.pdf",
    order_text_url="https://storage.googleapis.com/bucket/order.txt",
    cached_directives=None,
):
    order_entry = {
        "board_date": board_date,
        "order_date": order_date,
        "order_link": order_link,
        "order_text_url": order_text_url,
    }
    if cached_directives is not None:
        order_entry["compliance_directives"] = cached_directives
    return {
        "case_ref": case_ref,
        "board_date": board_date,
        "order_date": order_date,
        "order_category": order_category,
        "order_link": order_link,
        "order_history": [order_entry],
    }


def _mock_user_manager(is_admin=False, agp_filter="Pooja Deshpande"):
    um = MagicMock()
    um.is_admin.return_value = is_admin
    um.get_user_agp_filter.return_value = agp_filter
    return um


def _mock_board(rows):
    board_instance = MagicMock()
    board_instance.getData.return_value = rows
    board_cls = Mock(return_value=board_instance)
    return board_cls, board_instance


def _mock_auto_mgr(
    downloaded_text="AGP appeared. File reply affidavit by 13th August.",
):
    mgr = MagicMock()
    mgr._download_gcs_text.return_value = downloaded_text
    # Real CaseDataStore._to_iso_date is a pure string-format normaliser;
    # the portal-lookup phase calls it on dates that are already
    # YYYY-MM-DD in these fixtures, so a pass-through is a faithful stand-in
    # without needing every test to configure it itself.
    mgr.case_store._to_iso_date.side_effect = lambda v: v
    # _row() fixtures don't set order_petitioner/order_respondent, so every
    # eligible row looks "missing parties" to the portal-lookup phase. Default
    # the live lookup to "not found on the portal" so tests that aren't
    # exercising that phase don't pick up a MagicMock as a fake result;
    # tests that DO want to exercise it override court_scraper explicitly.
    mgr.court_scraper._fetch_with_provider.return_value = None
    return mgr


@pytest.mark.asyncio
async def test_non_admin_cannot_scan_for_another_user(monkeypatch):
    monkeypatch.setattr(
        main, "get_user_manager", lambda: _mock_user_manager(is_admin=False)
    )
    with pytest.raises(HTTPException) as exc_info:
        await main.compliance_scan(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name="Someone Else",
            current_user_with_profile={"uid": "u1"},
        )
    assert exc_info.value.status_code == 403


@pytest.mark.asyncio
async def test_adjourned_rows_are_skipped_entirely(monkeypatch):
    rows = [_row(order_category="ADJOURNED")]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: _mock_auto_mgr())
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    with patch("compliance_extractor.extract_directives") as extract:
        result = await main.compliance_scan(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name=None,
            current_user_with_profile={"uid": "u1"},
        )

    extract.assert_not_called()
    assert result["count"] == 0
    assert result["disposed_count"] == 0


@pytest.mark.asyncio
async def test_response_reports_scan_transparency_counts(monkeypatch):
    """The scan should surface how much of the AGP's matter universe it
    actually examined -- total matters in range, how many were eligible
    (HEARD_AND_ADJOURNED/DISPOSED_OFF) vs skipped as ADJOURNED -- so the UI
    can show "scanned N of M" instead of leaving that invisible."""
    rows = [
        _row(case_ref="WP/1/2026", order_category="ADJOURNED"),
        _row(case_ref="WP/2/2026", order_category="ADJOURNED"),
        _row(case_ref="WP/3/2026", order_category="HEARD_AND_ADJOURNED"),
        _row(case_ref="WP/4/2026", order_category="DISPOSED_OFF"),
    ]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: _mock_auto_mgr())
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    with patch("compliance_extractor.extract_directives", return_value=[]):
        result = await main.compliance_scan(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name=None,
            current_user_with_profile={"uid": "u1"},
        )

    assert result["total_matters"] == 4
    assert result["orders_scanned"] == 2
    assert result["adjourned_skipped"] == 2
    assert result["disposed_count"] == 1


@pytest.mark.asyncio
async def test_heard_and_adjourned_order_gets_scanned_and_cached(monkeypatch):
    rows = [_row(order_category="HEARD_AND_ADJOURNED")]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    mgr = _mock_auto_mgr()
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: mgr)
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    directives = [
        {
            "directive_type": "FILE_REPLY_AFFIDAVIT",
            "description": "file reply affidavit by 13th August",
            "deadline_date": "2026-08-13",
        }
    ]
    with patch(
        "compliance_extractor.extract_directives", return_value=directives
    ) as extract:
        result = await main.compliance_scan(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name=None,
            current_user_with_profile={"uid": "u1"},
        )

    extract.assert_called_once()
    assert result["count"] == 1
    assert result["newly_scanned"] == 1
    assert result["results"][0]["directives"] == directives
    mgr.case_store.set_order_compliance_directives.assert_called_once_with(
        "WP/1/2026",
        "https://storage.example/order.pdf",
        "2026-07-08",
        directives,
    )


@pytest.mark.asyncio
async def test_cached_directives_skip_a_second_llm_call(monkeypatch):
    cached = [
        {
            "directive_type": "FILE_REPLY_AFFIDAVIT",
            "description": "already extracted",
            "deadline_date": "2026-08-13",
        }
    ]
    rows = [_row(order_category="HEARD_AND_ADJOURNED", cached_directives=cached)]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    mgr = _mock_auto_mgr()
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: mgr)
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    with patch("compliance_extractor.extract_directives") as extract:
        result = await main.compliance_scan(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name=None,
            current_user_with_profile={"uid": "u1"},
        )

    extract.assert_not_called()
    mgr.case_store.set_order_compliance_directives.assert_not_called()
    assert result["newly_scanned"] == 0
    assert result["results"][0]["directives"] == cached


@pytest.mark.asyncio
async def test_disposed_case_is_reported_even_without_a_directive(monkeypatch):
    rows = [_row(order_category="DISPOSED_OFF", cached_directives=[])]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: _mock_auto_mgr())
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    result = await main.compliance_scan(
        start_date="2026-07-01",
        end_date="2026-07-31",
        user_name=None,
        current_user_with_profile={"uid": "u1"},
    )

    assert result["disposed_count"] == 1
    assert result["count"] == 1
    assert result["results"][0]["order_category"] == "DISPOSED_OFF"


@pytest.mark.asyncio
async def test_legacy_category_spelling_is_recognised(monkeypatch):
    rows = [_row(order_category="WP DISPOSED OF", cached_directives=[])]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: _mock_auto_mgr())
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    result = await main.compliance_scan(
        start_date="2026-07-01",
        end_date="2026-07-31",
        user_name=None,
        current_user_with_profile={"uid": "u1"},
    )

    assert result["disposed_count"] == 1
    assert result["results"][0]["order_category"] == "DISPOSED_OFF"


@pytest.mark.asyncio
async def test_without_gemini_key_disposed_cases_still_show_but_no_extraction_runs(
    monkeypatch,
):
    rows = [
        _row(case_ref="WP/1/2026", order_category="DISPOSED_OFF"),
        _row(case_ref="WP/2/2026", order_category="HEARD_AND_ADJOURNED"),
    ]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: _mock_auto_mgr())
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)

    with patch("compliance_extractor.extract_directives") as extract:
        result = await main.compliance_scan(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name=None,
            current_user_with_profile={"uid": "u1"},
        )

    extract.assert_not_called()
    assert result["ai_available"] is False
    assert result["disposed_count"] == 1
    # DISPOSED_OFF still reported (no directive needed to show it); the
    # HEARD_AND_ADJOURNED row has no cached directives and none could be
    # extracted, so it is correctly left out of the results.
    assert result["count"] == 1
    assert result["results"][0]["order_category"] == "DISPOSED_OFF"


@pytest.mark.asyncio
async def test_extraction_failure_is_counted_and_does_not_crash_the_scan(monkeypatch):
    rows = [_row(order_category="HEARD_AND_ADJOURNED")]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: _mock_auto_mgr())
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    with patch(
        "compliance_extractor.extract_directives", side_effect=RuntimeError("timeout")
    ):
        result = await main.compliance_scan(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name=None,
            current_user_with_profile={"uid": "u1"},
        )

    assert result["llm_errors"] == 1
    assert result["count"] == 0


@pytest.mark.asyncio
async def test_admin_can_scan_for_a_named_agp(monkeypatch):
    rows = [_row(order_category="DISPOSED_OFF", cached_directives=[])]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(
        main, "get_user_manager", lambda: _mock_user_manager(is_admin=True)
    )
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: _mock_auto_mgr())
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    await main.compliance_scan(
        start_date="2026-07-01",
        end_date="2026-07-31",
        user_name="Pooja Deshpande",
        current_user_with_profile={"uid": "admin1"},
    )

    board_instance.getData.assert_called_once_with(
        {"startDate": "2026-07-01", "endDate": "2026-07-31"}, "Pooja Deshpande"
    )


@pytest.mark.asyncio
async def test_row_with_no_date_matched_order_entry_is_skipped(monkeypatch):
    row = _row(order_category="HEARD_AND_ADJOURNED")
    # Corrupt the only order_history entry's board_date so it no longer
    # matches the row's own board_date -- simulates stale/mismatched data.
    row["order_history"][0]["board_date"] = "2020-01-01"
    board_cls, board_instance = _mock_board([row])
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: _mock_auto_mgr())
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    with patch("compliance_extractor.extract_directives") as extract:
        result = await main.compliance_scan(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name=None,
            current_user_with_profile={"uid": "u1"},
        )

    extract.assert_not_called()
    assert result["count"] == 0


@pytest.mark.asyncio
async def test_concurrent_extraction_resolves_multiple_uncached_orders(monkeypatch):
    """Uncached orders across a busy range must each still get their own
    directives and cache write, even though extraction now runs
    concurrently rather than one row at a time."""
    rows = [
        _row(case_ref="WP/1/2026", order_category="HEARD_AND_ADJOURNED"),
        _row(case_ref="WP/2/2026", order_category="HEARD_AND_ADJOURNED"),
        _row(case_ref="WP/3/2026", order_category="DISPOSED_OFF"),
    ]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    mgr = _mock_auto_mgr()
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: mgr)
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    def fake_extract(text, api_key, order_date=None, model=None):
        return [
            {
                "directive_type": "FILE_REPLY_AFFIDAVIT",
                "description": f"directive for text starting {text[:10]}",
                "deadline_date": "2026-08-13",
            }
        ]

    with patch(
        "compliance_extractor.extract_directives", side_effect=fake_extract
    ) as extract:
        result = await main.compliance_scan(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name=None,
            current_user_with_profile={"uid": "u1"},
        )

    assert extract.call_count == 3
    assert result["newly_scanned"] == 3
    assert mgr.case_store.set_order_compliance_directives.call_count == 3
    assert {r["case_ref"] for r in result["results"]} == {
        "WP/1/2026",
        "WP/2/2026",
        "WP/3/2026",
    }


@pytest.mark.asyncio
async def test_export_excel_returns_a_workbook_matching_the_scan(monkeypatch):
    from openpyxl import load_workbook

    directives = [
        {
            "directive_type": "FILE_REPLY_AFFIDAVIT",
            "description": "file reply affidavit by 13th August",
            "deadline_date": "2026-08-13",
        }
    ]
    rows = [
        _row(case_ref="WP/1/2026", order_category="HEARD_AND_ADJOURNED"),
        _row(case_ref="WP/2/2026", order_category="DISPOSED_OFF"),
    ]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: _mock_auto_mgr())
    monkeypatch.setenv("GEMINI_API_KEY", "test-key")

    with patch("compliance_extractor.extract_directives", return_value=directives):
        response = await main.export_compliance_excel(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name=None,
            current_user_with_profile={
                "uid": "u1",
                "profile": {"full_name": "Pooja Deshpande"},
            },
        )

    assert (
        response.media_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "Compliance_Report_2026-07-01_to_2026-07-31.xlsx" in (
        response.headers.get("content-disposition") or ""
    )

    wb = load_workbook(io.BytesIO(response.body))
    ws = wb.active
    all_text = "\n".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value
    )
    assert "POOJA DESHPANDE" in all_text
    assert "WP/1/2026" in all_text
    assert "WP/2/2026" in all_text
    assert "file reply affidavit by 13th August" in all_text


@pytest.mark.asyncio
async def test_export_excel_includes_petitioner_respondent_columns(monkeypatch):
    from openpyxl import load_workbook

    row = _row(case_ref="WP/3/2026", order_category="DISPOSED_OFF")
    row["order_petitioner"] = "Alice Petitioner"
    row["order_respondent"] = "State of Maharashtra"
    board_cls, board_instance = _mock_board([row])
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: _mock_auto_mgr())
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)

    response = await main.export_compliance_excel(
        start_date="2026-07-01",
        end_date="2026-07-31",
        user_name=None,
        current_user_with_profile={"uid": "u1"},
    )

    wb = load_workbook(io.BytesIO(response.body))
    ws = wb.active
    all_text = "\n".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value
    )
    assert "Alice Petitioner" in all_text
    assert "State of Maharashtra" in all_text
    assert "Petitioner" in all_text  # header
    assert "Respondent" in all_text  # header


@pytest.mark.asyncio
async def test_export_excel_denies_non_admin_requesting_another_user(monkeypatch):
    monkeypatch.setattr(
        main, "get_user_manager", lambda: _mock_user_manager(is_admin=False)
    )
    with pytest.raises(HTTPException) as exc_info:
        await main.export_compliance_excel(
            start_date="2026-07-01",
            end_date="2026-07-31",
            user_name="Someone Else",
            current_user_with_profile={"uid": "u1"},
        )
    assert exc_info.value.status_code == 403


# ---------------------------------------------------------------------------
# Live portal lookup (petitioner/respondent/stage/disposal_date backfill)
# ---------------------------------------------------------------------------


def _enriched_portal_result(
    petitioner="Alice Petitioner",
    respondent="State of Maharashtra",
    portal_case_status="DISPOSED",
    disposal_date="12/05/2026",
    case_orders=None,
):
    return {
        "petitioner": petitioner,
        "respondent": respondent,
        "portal_case_status": portal_case_status,
        "disposal_date": disposal_date,
        "case_orders": case_orders
        if case_orders is not None
        else [{"date": "2026-07-08", "download_link": "x", "stage": "Final Hearing"}],
    }


@pytest.mark.asyncio
async def test_portal_lookup_skipped_when_parties_already_present(monkeypatch):
    row = _row(order_category="DISPOSED_OFF")
    row["order_petitioner"] = "Existing Petitioner"
    row["order_respondent"] = "Existing Respondent"
    board_cls, board_instance = _mock_board([row])
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    mgr = _mock_auto_mgr()
    mgr.court_scraper = MagicMock()
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: mgr)
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)

    result = await main.compliance_scan(
        start_date="2026-07-01",
        end_date="2026-07-31",
        user_name=None,
        current_user_with_profile={"uid": "u1"},
    )

    mgr.court_scraper._fetch_with_provider.assert_not_called()
    assert result["portal_checked"] == 0
    assert result["results"][0]["petitioner"] == "Existing Petitioner"
    assert result["results"][0]["respondent"] == "Existing Respondent"


@pytest.mark.asyncio
async def test_portal_lookup_backfills_missing_parties_and_stage(monkeypatch):
    row = _row(
        case_ref="WP/1/2026",
        order_category="DISPOSED_OFF",
        order_date="2026-07-08",
    )
    board_cls, board_instance = _mock_board([row])
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    mgr = _mock_auto_mgr()
    mgr.court_scraper = MagicMock()
    mgr.court_scraper._fetch_with_provider.return_value = {"raw": "provider-result"}
    mgr.court_scraper._enrich_case_orders_result.return_value = (
        _enriched_portal_result()
    )
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: mgr)
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)

    result = await main.compliance_scan(
        start_date="2026-07-01",
        end_date="2026-07-31",
        user_name=None,
        current_user_with_profile={"uid": "u1"},
    )

    mgr.court_scraper._fetch_with_provider.assert_called_once_with(
        case_ref="WP/1/2026", date=None, bench="mumbai"
    )
    assert result["portal_checked"] == 1
    assert result["portal_check_errors"] == 0
    row_result = result["results"][0]
    assert row_result["petitioner"] == "Alice Petitioner"
    assert row_result["respondent"] == "State of Maharashtra"
    assert row_result["portal_case_status"] == "DISPOSED"
    assert row_result["portal_disposal_date"] == "12/05/2026"
    assert row_result["portal_stage"] == "Final Hearing"

    mgr.case_store.update_case_portal_status.assert_called_once()
    call_kwargs = mgr.case_store.update_case_portal_status.call_args.kwargs
    assert call_kwargs["petitioner"] == "Alice Petitioner"
    assert call_kwargs["stage_by_date"] == {"2026-07-08": "Final Hearing"}


@pytest.mark.asyncio
async def test_portal_lookup_respects_the_per_scan_cap(monkeypatch):
    rows = [
        _row(case_ref=f"WP/{i}/2026", order_category="HEARD_AND_ADJOURNED")
        for i in range(3)
    ]
    board_cls, board_instance = _mock_board(rows)
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    mgr = _mock_auto_mgr()
    mgr.court_scraper = MagicMock()
    mgr.court_scraper._fetch_with_provider.return_value = {"raw": "x"}
    mgr.court_scraper._enrich_case_orders_result.return_value = _enriched_portal_result(
        case_orders=[]
    )
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: mgr)
    monkeypatch.setattr(main, "MAX_PORTAL_LOOKUPS_PER_SCAN", 1)
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)

    result = await main.compliance_scan(
        start_date="2026-07-01",
        end_date="2026-07-31",
        user_name=None,
        current_user_with_profile={"uid": "u1"},
    )

    assert mgr.court_scraper._fetch_with_provider.call_count == 1
    assert result["portal_checked"] == 1
    assert result["portal_check_capped"] == 2


@pytest.mark.asyncio
async def test_portal_lookup_failure_is_counted_not_raised(monkeypatch):
    row = _row(order_category="DISPOSED_OFF")
    board_cls, board_instance = _mock_board([row])
    monkeypatch.setattr(main, "Board", board_cls)
    monkeypatch.setattr(main, "get_user_manager", lambda: _mock_user_manager())
    mgr = _mock_auto_mgr()
    mgr.court_scraper = MagicMock()
    mgr.court_scraper._fetch_with_provider.side_effect = RuntimeError("portal down")
    monkeypatch.setattr(main, "get_auto_order_manager", lambda: mgr)
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)

    result = await main.compliance_scan(
        start_date="2026-07-01",
        end_date="2026-07-31",
        user_name=None,
        current_user_with_profile={"uid": "u1"},
    )

    assert result["portal_check_errors"] == 1
    assert result["portal_checked"] == 0
    # A DISPOSED_OFF row still gets reported even though the portal lookup
    # failed -- the failure must not take down the whole scan.
    assert result["count"] == 1
